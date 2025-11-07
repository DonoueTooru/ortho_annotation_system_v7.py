import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser, simpledialog
from PIL import Image, ImageTk, ImageDraw, ImageFont
import json
import csv
import pandas as pd
from datetime import datetime
import shutil
import math
import numpy as np
import cv2
import copy
from pathlib import Path
import re
from io import BytesIO

# cairosvgは不要になりました（PNG/JPG直接読み込みに変更）


try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None

try:
    from PIL.ExifTags import TAGS, GPSTAGS
except ImportError:
    TAGS = {}
    GPSTAGS = {}

import difflib

# 管理レベルおよび拡張出力用の定数とユーティリティ
MANAGEMENT_LEVELS = ['S', 'A', 'B', 'N']

DEFECT_V2_HEADERS = [
    'エリア(工区)','接続箱No.','接続箱','回路No.','アレイ番号','モジュール場所',
    '不良分類','管理レベル','シリアルナンバー(交換前)','報告書番号','離線した日','離線場所',
    'シリアルナンバー(交換後)','モジュール','列1','送電完了日','報告書番号2','備考',
    '報告年月日','採番'
]

def normalize_management_level(value):
    if not value:
        return 'S'
    v = str(value).upper().strip()
    return v if v in MANAGEMENT_LEVELS else 'S'

class ODMImageSelector:
    def __init__(self, parent, annotation, image_type, webodm_path, callback, app_ref=None, rjpeg_folder=None):
        self.parent = parent
        self.annotation = annotation
        self.image_type = image_type
        self.webodm_path = webodm_path
        self.callback = callback
        self.app_ref = app_ref  # OrthoImageAnnotationSystem インスタンス参照（色設定取得用）
        self.rjpeg_folder = rjpeg_folder  # R-JPEG画像フォルダ
        
        self.coverage_image = None
        self.coverage_image_path = None
        self.ortho_image_size = None
        self.image_positions = []
        self.selected_image_path = None
        self.zoom_factor = 1.0
        
        self.create_selector_window()
        self.load_webodm_assets()

    def create_selector_window(self):
        """ODM画像選択ウィンドウを作成"""
        self.window = tk.Toplevel(self.parent)
        self.window.title(f"ODM画像選択 - {self.image_type}")
        self.window.geometry("1000x800")
        self.window.transient(self.parent)
        self.window.grab_set()
        
        # デバッグログ有効化フラグ（コンソール出力）
        self.debug = True
        
        main_frame = ttk.Frame(self.window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        info_frame = ttk.LabelFrame(main_frame, text="選択情報", padding=5)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        self.info_label = ttk.Label(info_frame, text="WebODMアセットを読み込み中...")
        self.info_label.pack()
        
        # サムネイルプレビュー枠
        preview_frame = ttk.LabelFrame(main_frame, text="プレビュー", padding=5)
        preview_frame.pack(fill=tk.X, pady=(0, 10))
        self.preview_label = tk.Label(preview_frame, text="画像を選択するとプレビューを表示します", anchor="center")
        self.preview_label.pack(fill=tk.X)
        self._odm_preview_imgtk = None  # 参照保持用
        
        image_frame = ttk.LabelFrame(main_frame, text="カバレッジ画像", padding=5)
        image_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas_frame = ttk.Frame(image_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas = tk.Canvas(canvas_frame, bg="white")
        v_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        h_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        
        self.canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.canvas.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)
        
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<MouseWheel>", self.on_mouse_wheel)
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(button_frame, text="R-JPEGフォルダ選択", command=self.select_rjpeg_folder).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="選択", command=self.confirm_selection).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(button_frame, text="キャンセル", command=self.window.destroy).pack(side=tk.RIGHT)

    def select_rjpeg_folder(self):
        """R-JPEG画像フォルダを選択してマッチング実行"""
        folder_path = filedialog.askdirectory(title="R-JPEG画像フォルダを選択")
        if folder_path:
            self.rjpeg_folder = folder_path
            self.match_rjpeg_images()
            self.display_coverage_image()
            self.update_info()

    def debug_log(self, msg: str):
        """簡易デバッグ出力（コンソール）"""
        try:
            if getattr(self, 'debug', True):
                print(f"[DEBUG][ODM] {msg}")
        except Exception:
            pass

    def update_preview(self, image_path):
        """プレビューを更新（最大380x240）"""
        try:
            if not image_path or not os.path.exists(image_path):
                self.preview_label.config(text="プレビューなし", image="")
                self._odm_preview_imgtk = None
                return
            img = Image.open(image_path)
            max_w, max_h = 380, 240
            w, h = img.size
            scale = min(1.0, min(max_w / float(w), max_h / float(h)))
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            img = img.resize(new_size, Image.Resampling.LANCZOS)
            imgtk = ImageTk.PhotoImage(img)
            self.preview_label.config(image=imgtk, text="")
            self._odm_preview_imgtk = imgtk
            self.debug_log(f"preview updated size={new_size} for {os.path.basename(image_path)}")
        except Exception as e:
            self.preview_label.config(text=f"プレビュー失敗: {e}", image="")
            self._odm_preview_imgtk = None
            self.debug_log(f"preview failed: {e}")

    def get_gps_coordinates(self, image_path):
        """画像からGPS座標を取得"""
        try:
            img = Image.open(image_path)
            exif_data = img._getexif()
            if not exif_data:
                return None
            
            gps_info = {}
            for tag, value in exif_data.items():
                tag_name = TAGS.get(tag, tag)
                if tag_name == 'GPSInfo':
                    for gps_tag in value:
                        gps_tag_name = GPSTAGS.get(gps_tag, gps_tag)
                        gps_info[gps_tag_name] = value[gps_tag]
            
            if not gps_info:
                return None
                
            # GPS座標を10進数に変換
            lat = self.convert_to_degrees(gps_info.get('GPSLatitude'))
            lon = self.convert_to_degrees(gps_info.get('GPSLongitude'))
            
            if lat is None or lon is None:
                return None
            
            if gps_info.get('GPSLatitudeRef') == 'S':
                lat = -lat
            if gps_info.get('GPSLongitudeRef') == 'W':
                lon = -lon
                
            return (lat, lon)
        except Exception as e:
            self.debug_log(f"GPS取得エラー ({os.path.basename(image_path)}): {e}")
            return None

    def convert_to_degrees(self, value):
        """GPS座標を度分秒から10進数に変換"""
        try:
            if not value or len(value) < 3:
                return None
            d, m, s = value
            return float(d) + float(m) / 60 + float(s) / 3600
        except Exception:
            return None

    def get_image_timestamp(self, image_path):
        """画像のEXIF撮影日時を取得"""
        try:
            img = Image.open(image_path)
            exif_data = img._getexif()
            if exif_data:
                for tag, value in exif_data.items():
                    if TAGS.get(tag) == 'DateTimeOriginal':
                        return datetime.strptime(str(value), '%Y:%m:%d %H:%M:%S')
            return None
        except Exception as e:
            self.debug_log(f"タイムスタンプ取得エラー ({os.path.basename(image_path)}): {e}")
            return None

    def match_rjpeg_images(self):
        """R-JPEG画像とWebODM座標を複合マッチング（GPS→ファイル名→タイムスタンプ）"""
        if not self.rjpeg_folder or not os.path.isdir(self.rjpeg_folder):
            self.debug_log("R-JPEGフォルダが指定されていません")
            return
        
        self.debug_log(f"R-JPEGマッチング開始: {self.rjpeg_folder}")
        
        # R-JPEG画像ファイルを収集
        rjpeg_files = []
        for root, dirs, files in os.walk(self.rjpeg_folder):
            for f in files:
                if f.lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff')):
                    rjpeg_files.append(os.path.join(root, f))
        
        self.debug_log(f"R-JPEG画像数: {len(rjpeg_files)}")
        
        if not rjpeg_files:
            self.debug_log("R-JPEG画像が見つかりません")
            return
        
        # WebODM画像位置情報のバックアップ
        original_positions = list(self.image_positions)
        
        # 各WebODM画像位置に対して最適なR-JPEG画像を見つける
        matched_positions = []
        
        for pos in original_positions:
            webodm_filename = pos['filename']
            webodm_path = pos['path']
            webodm_x = pos['x']
            webodm_y = pos['y']
            
            # WebODM画像の情報取得
            webodm_gps = self.get_gps_coordinates(webodm_path) if os.path.exists(webodm_path) else None
            webodm_time = self.get_image_timestamp(webodm_path) if os.path.exists(webodm_path) else None
            webodm_basename = os.path.splitext(os.path.basename(webodm_filename))[0].lower()
            
            best_match = None
            best_score = 0
            best_reason = ""
            
            for rjpeg_path in rjpeg_files:
                score = 0
                reasons = []
                
                # 1. GPSマッチング（最優先・最も高得点）
                rjpeg_gps = self.get_gps_coordinates(rjpeg_path)
                if webodm_gps and rjpeg_gps:
                    # 距離計算（簡易的なユークリッド距離）
                    distance = math.sqrt(
                        (webodm_gps[0] - rjpeg_gps[0])**2 + 
                        (webodm_gps[1] - rjpeg_gps[1])**2
                    )
                    # 閾値: 0.0001度以内（約11m以内）なら高得点
                    if distance < 0.0001:
                        gps_score = 100 - (distance * 100000)  # 距離が近いほど高得点
                        score += gps_score
                        reasons.append(f"GPS一致(距離:{distance:.6f}度, +{gps_score:.1f}点)")
                    elif distance < 0.001:  # 約110m以内なら低得点
                        gps_score = 50 - (distance * 10000)
                        score += gps_score
                        reasons.append(f"GPS近似(距離:{distance:.6f}度, +{gps_score:.1f}点)")
                
                # 2. ファイル名マッチング（中優先）
                rjpeg_basename = os.path.splitext(os.path.basename(rjpeg_path))[0].lower()
                similarity = difflib.SequenceMatcher(None, webodm_basename, rjpeg_basename).ratio()
                if similarity > 0.6:  # 類似度60%以上
                    name_score = similarity * 30  # 最大30点
                    score += name_score
                    reasons.append(f"名前類似度:{similarity:.2%}(+{name_score:.1f}点)")
                
                # 3. タイムスタンプマッチング（補助的）
                rjpeg_time = self.get_image_timestamp(rjpeg_path)
                if webodm_time and rjpeg_time:
                    time_diff = abs((webodm_time - rjpeg_time).total_seconds())
                    if time_diff < 60:  # 60秒以内
                        time_score = max(0, 20 - (time_diff / 3))  # 最大20点
                        score += time_score
                        reasons.append(f"時刻差:{time_diff:.0f}秒(+{time_score:.1f}点)")
                    elif time_diff < 300:  # 5分以内
                        time_score = max(0, 10 - (time_diff / 30))
                        score += time_score
                        reasons.append(f"時刻差:{time_diff:.0f}秒(+{time_score:.1f}点)")
                
                # 最高スコア更新
                if score > best_score:
                    best_score = score
                    best_match = rjpeg_path
                    best_reason = ", ".join(reasons)
            
            # マッチング結果
            if best_match and best_score > 20:  # 閾値: 20点以上
                matched_positions.append({
                    'filename': os.path.basename(best_match),
                    'path': best_match,
                    'x': webodm_x,
                    'y': webodm_y,
                    'webodm_original': webodm_filename,
                    'match_score': best_score,
                    'match_reason': best_reason
                })
                self.debug_log(f"マッチ: {os.path.basename(webodm_filename)} -> {os.path.basename(best_match)} (スコア:{best_score:.1f}, {best_reason})")
            else:
                self.debug_log(f"マッチなし: {os.path.basename(webodm_filename)} (最高スコア:{best_score:.1f})")
        
        # マッチング結果を適用
        if matched_positions:
            self.image_positions = matched_positions
            self.debug_log(f"マッチング完了: {len(matched_positions)}/{len(original_positions)}件")
            messagebox.showinfo(
                "マッチング結果",
                f"R-JPEG画像のマッチングが完了しました。\n\n"
                f"WebODM画像: {len(original_positions)}件\n"
                f"マッチ成功: {len(matched_positions)}件\n"
                f"マッチ失敗: {len(original_positions) - len(matched_positions)}件"
            )
        else:
            messagebox.showwarning(
                "マッチング失敗",
                "R-JPEG画像とWebODM座標のマッチングに失敗しました。\n\n"
                "以下を確認してください：\n"
                "- R-JPEG画像にGPS情報が含まれているか\n"
                "- ファイル名が類似しているか\n"
                "- 撮影日時が近いか"
            )

    def load_webodm_assets(self):
        """WebODMアセットを読み込み、座標情報を解析する"""
        try:
            # 1. オルソ画像のパスとサイズ取得
            ortho_path = os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto.tif')
            with Image.open(ortho_path) as img:
                self.ortho_image_size = img.size
                try:
                    self.debug_log(f"orthophoto size via Pillow: {self.ortho_image_size}")
                except Exception:
                    pass

            # 2. カバレッジ画像の読み込み
            self.coverage_image_path = os.path.join(self.webodm_path, 'images', 'shot_coverage.png')
            self.coverage_image = Image.open(self.coverage_image_path)

            # 3. 座標情報の読み込みと解析
            geo_ref_path = os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_model_geo.txt')
            self.image_positions = []
            with open(geo_ref_path, 'r', encoding='utf-8') as f:
                for line in f:
                    parts = line.strip().split()
                    if len(parts) >= 3:
                        filename = parts[0]
                        image_path = os.path.join(self.webodm_path, 'images', filename)
                        if os.path.exists(image_path):
                            self.image_positions.append({
                                'filename': filename,
                                'path': image_path,
                                'x': float(parts[1]),
                                'y': float(parts[2])
                            })
            
            # 4. R-JPEG画像とのマッチング（フォルダが指定されている場合）
            if self.rjpeg_folder and os.path.isdir(self.rjpeg_folder):
                self.match_rjpeg_images()
            
            self.display_coverage_image()
            self.update_info()

        except FileNotFoundError as e:
            messagebox.showerror("エラー", f"必要なWebODMファイルが見つかりません: {e.filename}")
            self.window.destroy()
        except Exception as e:
            messagebox.showerror("エラー", f"WebODMアセットの読み込みに失敗しました: {e}")
            self.window.destroy()

    def display_coverage_image(self):
        """カバレッジ画像と各種マーカーを表示"""
        if self.coverage_image:
            display_size = (int(self.coverage_image.width * self.zoom_factor), int(self.coverage_image.height * self.zoom_factor))
            resized_image = self.coverage_image.resize(display_size, Image.Resampling.LANCZOS)
            self.canvas_image = ImageTk.PhotoImage(resized_image)
            
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.canvas_image)
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            
            self.draw_image_markers()
            self.draw_current_annotation()

    def draw_current_annotation(self):
        """アノテーション位置をスケーリングして正確に描画"""
        if not self.coverage_image or not self.ortho_image_size:
            return

        # スケール計算
        ortho_w, ortho_h = self.ortho_image_size
        coverage_w, coverage_h = self.coverage_image.size
        scale_x = coverage_w / ortho_w
        scale_y = coverage_h / ortho_h

        # アノテーション座標をカバレッジ画像上の座標に変換
        ann_x = self.annotation['x'] * scale_x * self.zoom_factor
        ann_y = self.annotation['y'] * scale_y * self.zoom_factor
        
        # ... (以降の描画ロジックは前回とほぼ同じ、座標変数だけ変更) ...
        defect_type = self.annotation.get('defect_type', 'ホットスポット')
        # 色マップはアプリ本体（app_ref）から取得。無ければデフォルト色群にフォールバック。
        default_colors = {
            "ホットスポット": "#FF0000",
            "クラスタ異常": "#FF8C00",
            "破損": "#FFD700",
            "ストリング異常": "#0000FF",
            "系統異常": "#8A2BE2",
            "影": "#008000",
        }
        if getattr(self, 'app_ref', None) and hasattr(self.app_ref, 'defect_types'):
            color = self.app_ref.defect_types.get(defect_type, default_colors.get(defect_type, "#FF0000"))
        else:
            color = default_colors.get(defect_type, "#FF0000")
        shape = self.annotation.get('shape', 'cross')
        
        if shape == "cross": self.draw_cross_annotation(ann_x, ann_y, color)
        # ... (他の形状も同様に呼び出す) ...

    # 注釈描画の簡易ヘルパー（最低限：十字）。他形状は必要に応じて追加可。
    def draw_cross_annotation(self, x, y, color):
        size = 12 * self.zoom_factor
        self.canvas.create_line(x, y - size, x, y + size, fill=color, width=2, tags="ann_preview")
        self.canvas.create_line(x - size, y, x + size, y, fill=color, width=2, tags="ann_preview")

    def draw_image_markers(self):
        """georeferencing情報から写真の位置マーカーを正確に描画"""
        if not self.ortho_image_size or not self.coverage_image:
            return

        ortho_w, ortho_h = self.ortho_image_size
        coverage_w, coverage_h = self.coverage_image.size
        scale_x = coverage_w / ortho_w
        scale_y = coverage_h / ortho_h

        for i, pos_info in enumerate(self.image_positions):
            # オルソ座標をカバレッジ座標に変換
            x = pos_info['x'] * scale_x * self.zoom_factor
            y = pos_info['y'] * scale_y * self.zoom_factor
            
            marker_size = 5 * self.zoom_factor
            self.canvas.create_oval(
                x - marker_size, y - marker_size, x + marker_size, y + marker_size,
                fill="blue", outline="white", width=1, tags=f"marker_{i}"
            )
            # 判別用ラベル（番号＋ファイル名）
            try:
                base = os.path.basename(pos_info.get('path') or pos_info.get('filename') or '')
                label = f"{i+1}: {base}" if base else f"{i+1}"
                self.canvas.create_text(
                    x + 10 * self.zoom_factor, y - 10 * self.zoom_factor,
                    text=label, anchor="w", fill="#003366",
                    font=("Arial", max(8, int(9 * self.zoom_factor))),
                    tags=f"label_{i}"
                )
            except Exception:
                pass

    def on_canvas_click(self, event):
        """キャンバスクリックで最も近い写真を選択"""
        if not self.coverage_image or not self.image_positions:
            return
        
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)

        ortho_w, ortho_h = self.ortho_image_size
        coverage_w, coverage_h = self.coverage_image.size
        scale_x = coverage_w / ortho_w
        scale_y = coverage_h / ortho_h
        
        min_distance = float('inf')
        selected_index = -1
        
        for i, pos_info in enumerate(self.image_positions):
            marker_x = pos_info['x'] * scale_x * self.zoom_factor
            marker_y = pos_info['y'] * scale_y * self.zoom_factor
            
            distance = math.sqrt((canvas_x - marker_x)**2 + (canvas_y - marker_y)**2)
            
            if distance < 10 * self.zoom_factor and distance < min_distance:
                min_distance = distance
                selected_index = i
        
        if selected_index >= 0:
            self.select_image_by_index(selected_index)
    
    def select_image_by_index(self, index):
        """指定されたインデックスの画像を選択"""
        if 0 <= index < len(self.image_positions):
            selected_image = self.image_positions[index]
            self.selected_image_path = selected_image.get('path')
            
            # 選択状態を視覚的に表示
            self.canvas.delete("selected")
            
            ortho_w, ortho_h = self.ortho_image_size
            coverage_w, coverage_h = self.coverage_image.size
            scale_x = coverage_w / ortho_w
            scale_y = coverage_h / ortho_h
            x = selected_image.get('x', 0) * scale_x * self.zoom_factor
            y = selected_image.get('y', 0) * scale_y * self.zoom_factor
            marker_size = 15 * self.zoom_factor
            
            self.canvas.create_oval(
                x - marker_size, y - marker_size,
                x + marker_size, y + marker_size,
                fill="yellow", outline="black", width=3,
                tags="selected"
            )
            
            self.update_info()
    
    def on_mouse_wheel(self, event):
        """マウスホイールでズーム"""
        if self.coverage_image:
            if event.delta > 0:
                self.zoom_factor *= 1.1
            else:
                self.zoom_factor /= 1.1
            
            self.zoom_factor = max(0.1, min(5.0, self.zoom_factor))
            self.display_coverage_image()
    
    def update_info(self):
        """情報表示を更新"""
        info_text = ""
        if self.coverage_image_path:
            info_text += f"カバレッジ画像: {os.path.basename(self.coverage_image_path)}\n"
        
        if self.image_positions:
            info_text += f"画像数: {len(self.image_positions)}枚\n"
        
        if self.selected_image_path:
            info_text += f"選択画像: {os.path.basename(self.selected_image_path)}"
        
        self.info_label.config(text=info_text or "カバレッジ画像と画像フォルダを選択してください")
    
    def confirm_selection(self):
        """選択を確定"""
        if self.selected_image_path:
            self.callback(self.selected_image_path)
            self.window.destroy()
        else:
            messagebox.showwarning("警告", "画像を選択してください")



# --- ODMImageSelector extra methods (monkey patch) ---

def _odmselector_debug_log(self, msg: str):
    try:
        if getattr(self, 'debug', True):
            print(f"[DEBUG][ODM] {msg}")
    except Exception:
        pass

def _odmselector_update_preview(self, image_path):
    try:
        if not image_path or not os.path.exists(image_path):
            if hasattr(self, 'preview_label'):
                self.preview_label.config(text="プレビューなし", image="")
            self._odm_preview_imgtk = None
            return
        img = Image.open(image_path)
        max_w, max_h = 380, 240
        w, h = img.size
        scale = min(1.0, min(max_w / float(w), max_h / float(h)))
        new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
        imgtk = ImageTk.PhotoImage(img)
        if hasattr(self, 'preview_label'):
            self.preview_label.config(image=imgtk, text="")
        self._odm_preview_imgtk = imgtk
        _odmselector_debug_log(self, f"preview updated size={new_size} for {os.path.basename(image_path)}")
    except Exception as e:
        try:
            if hasattr(self, 'preview_label'):
                self.preview_label.config(text=f"プレビュー失敗: {e}", image="")
        except Exception:
            pass
        self._odm_preview_imgtk = None
        _odmselector_debug_log(self, f"preview failed: {e}")

def _odmselector_select_image_by_index(self, index):
    """指定されたインデックスの画像を選択（プレビュー＆ログ付き）"""
    if 0 <= index < len(self.image_positions):
        selected_image = self.image_positions[index]
        self.selected_image_path = selected_image.get('path')
        # 選択状態の視覚表示
        self.canvas.delete("selected")
        ortho_w, ortho_h = self.ortho_image_size
        coverage_w, coverage_h = self.coverage_image.size
        scale_x = coverage_w / ortho_w
        scale_y = coverage_h / ortho_h
        x = selected_image.get('x', 0) * scale_x * self.zoom_factor
        y = selected_image.get('y', 0) * scale_y * self.zoom_factor
        marker_size = 15 * self.zoom_factor
        self.canvas.create_oval(
            x - marker_size, y - marker_size,
            x + marker_size, y + marker_size,
            fill="yellow", outline="black", width=3,
            tags="selected"
        )
        # プレビュー更新
        try:
            if self.selected_image_path:
                _odmselector_update_preview(self, self.selected_image_path)
        except Exception:
            pass
        self.update_info()
        _odmselector_debug_log(self, f"selected index={index}, path={self.selected_image_path}")

# メソッドをクラスへ付与

def _odmselector_create_selector_window(self):
    """ODM画像選択ウィンドウを作成（プレビュー＋画像フォルダ選択ボタン付き）"""
    self.window = tk.Toplevel(self.parent)
    self.window.title(f"ODM画像選択 - {self.image_type}")
    self.window.geometry("1000x800")
    self.window.transient(self.parent)
    self.window.grab_set()

    # デバッグログ有効化フラグ
    self.debug = True

    main_frame = ttk.Frame(self.window)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    info_frame = ttk.LabelFrame(main_frame, text="選択情報", padding=5)
    info_frame.pack(fill=tk.X, pady=(0, 10))
    self.info_label = ttk.Label(info_frame, text="WebODMアセットを読み込み中...")
    self.info_label.pack()

    # サムネイルプレビュー枠
    preview_frame = ttk.LabelFrame(main_frame, text="プレビュー", padding=5)
    preview_frame.pack(fill=tk.X, pady=(0, 10))
    self.preview_label = tk.Label(preview_frame, text="画像を選択するとプレビューを表示します", anchor="center")
    self.preview_label.pack(fill=tk.X)
    self._odm_preview_imgtk = None  # 参照保持

    image_frame = ttk.LabelFrame(main_frame, text="カバレッジ画像", padding=5)
    image_frame.pack(fill=tk.BOTH, expand=True)

    canvas_frame = ttk.Frame(image_frame)
    canvas_frame.pack(fill=tk.BOTH, expand=True)

    self.canvas = tk.Canvas(canvas_frame, bg="white")
    v_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
    h_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
    self.canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

    self.canvas.grid(row=0, column=0, sticky="nsew")
    v_scrollbar.grid(row=0, column=1, sticky="ns")
    h_scrollbar.grid(row=1, column=0, sticky="ew")

    canvas_frame.grid_rowconfigure(0, weight=1)
    canvas_frame.grid_columnconfigure(0, weight=1)

    self.canvas.bind("<Button-1>", self.on_canvas_click)
    self.canvas.bind("<MouseWheel>", self.on_mouse_wheel)

    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill=tk.X, pady=(10, 0))
    ttk.Button(button_frame, text="画像フォルダ選択", command=self.select_images_folder).pack(side=tk.LEFT, padx=(0, 10))
    ttk.Button(button_frame, text="選択", command=self.confirm_selection).pack(side=tk.RIGHT, padx=(10, 0))
    ttk.Button(button_frame, text="キャンセル", command=self.window.destroy).pack(side=tk.RIGHT)


def _odmselector_norm_key(name: str):
    """拡張子を除き、大小文字差を無視した照合キーを生成（DJI_XXXX.*対応）"""
    try:
        if not name:
            return ""
        base = os.path.basename(name)
        stem, _ext = os.path.splitext(base)
        return stem.lower()
    except Exception:
        return str(name).lower()


def _odmselector_select_images_folder(self):
    """画像フォルダを選択し、basename（拡張子除く）とgeoreferencingの照合でマーカーを展開"""
    try:
        folder = filedialog.askdirectory(title="座標付き画像フォルダを選択")
        if not folder:
            return
        target_exts = {'.jpg', '.jpeg', '.png', '.tif', '.tiff'}
        files = []
        for root, dirs, fnames in os.walk(folder):
            for fn in fnames:
                ext = os.path.splitext(fn)[1].lower()
                if ext in target_exts:
                    files.append(os.path.join(root, fn))
        _odmselector_debug_log(self, f"scan folder done: {folder}, candidates={len(files)}")

        # georeferencing由来の basename（stem小文字） -> (px, py) インデックス
        geo_index = {}
        for p in getattr(self, 'image_positions', []):
            base = os.path.basename(p.get('filename') or p.get('path') or '')
            if not base:
                base = os.path.basename(p.get('path') or '')
            key = _odmselector_norm_key(base)
            if key:
                geo_index[key] = (float(p.get('x', 0)), float(p.get('y', 0)))
        self._geo_index = geo_index

        matched = []
        unmatched_samples = []
        for fp in files:
            key = _odmselector_norm_key(os.path.basename(fp))
            if key in geo_index:
                px, py = geo_index[key]
                matched.append({'filename': os.path.basename(fp), 'path': fp, 'x': float(px), 'y': float(py)})
            else:
                if len(unmatched_samples) < 5:
                    unmatched_samples.append(os.path.basename(fp))

        if matched:
            self.image_positions = matched
            self.display_coverage_image()
            self.update_info()
            try:
                if len(self.image_positions) > 0:
                    self.select_image_by_index(0)
            except Exception:
                pass
        else:
            messagebox.showwarning("警告", "このフォルダ内で位置情報に一致する画像は見つかりませんでした。")

        total = len(files)
        um = total - len(matched)
        _odmselector_debug_log(self, f"folder match: total={total}, matched={len(matched)}, unmatched={um}")
        if um > 0:
            sample_text = ", ".join(unmatched_samples)
            messagebox.showinfo("情報", f"未マッチ: {um}件\n例: {sample_text}")
    except Exception as e:
        _odmselector_debug_log(self, f"select_images_folder failed: {e}")
        messagebox.showerror("エラー", f"画像フォルダ選択処理でエラー: {e}")


def _odmselector_update_info(self):
    """情報表示を更新＋必要ならフォルダ選択を促す"""
    # geo index 構築（未構築時）
    try:
        if not hasattr(self, '_geo_index') or not self._geo_index:
            geo_index = {}
            for p in getattr(self, 'image_positions', []) or []:
                base = os.path.basename(p.get('filename') or p.get('path') or '')
                if not base:
                    base = os.path.basename(p.get('path') or '')
                key = _odmselector_norm_key(base)
                if key:
                    geo_index[key] = (float(p.get('x', 0)), float(p.get('y', 0)))
            self._geo_index = geo_index
            _odmselector_debug_log(self, f"geo index built in update_info: keys={len(self._geo_index)}")
    except Exception:
        pass

    info_text = ""
    if getattr(self, 'coverage_image_path', None):
        info_text += f"カバレッジ画像: {os.path.basename(self.coverage_image_path)}\n"
    if getattr(self, 'image_positions', None):
        info_text += f"画像数: {len(self.image_positions)}枚\n"
    if getattr(self, 'selected_image_path', None):
        info_text += f"選択画像: {os.path.basename(self.selected_image_path)}"
    self.info_label.config(text=info_text or "カバレッジ画像と画像フォルダを選択してください")
    # 初期マーカーが少ない場合のフォルダ選択促し（1度だけ）
    try:
        if not getattr(self, '_folder_prompted', False):
            cnt = len(self.image_positions) if isinstance(self.image_positions, list) else 0
            if cnt <= 2:
                self._folder_prompted = True
                if messagebox.askyesno("確認", "マーカー数が少ないため、画像フォルダを選択してマーカーを展開しますか？"):
                    if hasattr(self, 'select_images_folder'):
                        self.select_images_folder()
    except Exception:
        pass

ODMImageSelector.debug_log = _odmselector_debug_log
ODMImageSelector.update_preview = _odmselector_update_preview
ODMImageSelector.select_image_by_index = _odmselector_select_image_by_index
ODMImageSelector.select_images_folder = _odmselector_select_images_folder
ODMImageSelector._norm_key = _odmselector_norm_key
ODMImageSelector.update_info = _odmselector_update_info
ODMImageSelector.create_selector_window = _odmselector_create_selector_window

# --- Robust WebODM asset loader (monkey patch) ---

def load_webodm_assets_robust(self):
    """WebODMアセットを読み込み、座標情報を解析する（堅牢化版）"""
    import re
    try:
        self.debug_log(f"start load_webodm_assets: webodm_path={self.webodm_path}")
    except Exception:
        pass
    try:
        def find_first_existing(base_path, relative_paths):
            for rel in relative_paths:
                p = os.path.join(base_path, rel)
                if os.path.exists(p):
                    return p
            return None

        def read_world_file(raster_path):
            base, _ = os.path.splitext(raster_path)
            candidates = [base + '.tfw', base + '.wld']
            for wf in candidates:
                if os.path.exists(wf):
                    try:
                        with open(wf, 'r') as f:
                            vals = [float(line.strip()) for line in f if line.strip()]
                        if len(vals) >= 6:
                            return {
                                'A': vals[0],  # pixel size x
                                'B': vals[1],  # rotation
                                'D': vals[2],  # rotation
                                'E': vals[3],  # pixel size y (negative)
                                'C': vals[4],  # top-left x
                                'F': vals[5],  # top-left y
                            }
                    except Exception:
                        pass
            return None

        def world_to_pixel(gt, x, y):
            if not gt:
                return None
            A, B, D, E, C, F = gt['A'], gt['B'], gt['D'], gt['E'], gt['C'], gt['F']
            # no rotation
            if abs(B) < 1e-9 and abs(D) < 1e-9:
                px = (x - C) / A
                py = (F - y) / abs(E)
                return px, py
            det = A * E - B * D
            if abs(det) < 1e-12:
                return None
            px = (E * (x - C) - B * (y - F)) / det
            py = (-D * (x - C) + A * (y - F)) / det
            return px, py

        def extract_filename_and_xy(line):
            # filename
            m = re.search(r'([^\s,\"]+\.(?:jpg|jpeg|png|tif|tiff))', line, re.IGNORECASE)
            fname = m.group(1) if m else None
            # numbers
            nums = re.findall(r'[-+]?(?:\d*\.\d+|\d+)', line)
            if len(nums) >= 2:
                try:
                    return fname, float(nums[0]), float(nums[1])
                except Exception:
                    return fname, None, None
            return fname, None, None

        def resolve_image_path(base_path, filename):
            if not filename:
                return None
            candidates = []
            if os.path.isabs(filename):
                candidates.append(filename)
            else:
                candidates.append(os.path.join(base_path, filename))
                candidates.append(os.path.join(base_path, 'images', filename))
                candidates.append(os.path.join(base_path, 'images', os.path.basename(filename)))
            for c in candidates:
                if os.path.exists(c):
                    return c
            return None

        # 1) オルソ画像
        ortho_candidates = [
            os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto.tif'),
            os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto.tiff'),
            os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto.png'),
        ]
        self.ortho_path = next((p for p in ortho_candidates if os.path.exists(p)), None)
        if not self.ortho_path:
            raise FileNotFoundError(os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto.tif'))
        # オルソ画像サイズを堅牢に取得（Pillow失敗時にtifffile/rasterioフォールバック）
        self.ortho_image_size = None
        try:
            with Image.open(self.ortho_path) as img:
                self.ortho_image_size = img.size
                try:
                    self.debug_log(f"orthophoto size via Pillow: {self.ortho_image_size}")
                except Exception:
                    pass
        except Exception:
            try:
                import tifffile as tiff
                with tiff.TiffFile(self.ortho_path) as tf:
                    page = tf.pages[0]
                    w = getattr(page, 'imagewidth', None)
                    h = getattr(page, 'imagelength', None)
                    if w is not None and h is not None:
                        self.ortho_image_size = (int(w), int(h))
                    else:
                        shp = getattr(page, 'shape', None)
                        if shp:
                            if len(shp) == 2:
                                self.ortho_image_size = (int(shp[1]), int(shp[0]))
                            elif len(shp) == 3:
                                self.ortho_image_size = (int(shp[2]), int(shp[1]))
            except Exception:
                try:
                    import rasterio
                    with rasterio.open(self.ortho_path) as ds:
                        self.ortho_image_size = (int(ds.width), int(ds.height))
                except Exception:
                    pass
        if not self.ortho_image_size:
            # 最終手段: 実読込してサイズ取得（重い可能性あり）
            try:
                import tifffile as tiff, numpy as np
                arr = tiff.imread(self.ortho_path)
                if arr.ndim == 2:
                    h, w = arr.shape
                elif arr.ndim == 3:
                    if arr.shape[0] in (1, 3, 4) and arr.shape[-1] not in (3, 4):
                        # (bands, H, W) → (H, W, bands)
                        arr = np.transpose(arr, (1, 2, 0))
                    h, w = arr.shape[0], arr.shape[1]
                else:
                    raise RuntimeError('Unsupported TIFF shape')
                self.ortho_image_size = (int(w), int(h))
            except Exception as e:
                raise RuntimeError(f"orthophoto size detection failed: {e}")

        # 2) ワールドファイル
        geotransform = read_world_file(self.ortho_path)

        # 3) カバレッジ画像（任意）
        cov = find_first_existing(self.webodm_path, [
            'images/shot_coverage.png',
            'odm_report/assets/odm_orthophoto_coverage.png',
            'odm_report/odm_orthophoto_coverage.png',
            'odm_orthophoto/odm_orthophoto.png',
            'odm_orthophoto/odm_orthophoto_preview.png',
        ])
        if cov and os.path.exists(cov):
            try:
                self.coverage_image = Image.open(cov)
                self.coverage_image_path = cov
            except Exception:
                self.coverage_image = None
                self.coverage_image_path = None
        if not getattr(self, 'coverage_image', None):
            # orthophoto から低解像度プレビューを生成（tifffile→rasterio→最後にPillowの順で試行）
            preview_img = None
            # 1) tifffile で読み取り→8bit化→縮小
            try:
                import tifffile as tiff, numpy as np
                arr = tiff.imread(self.ortho_path)
                arr = np.nan_to_num(arr, nan=0.0, posinf=0.0, neginf=0.0)
                if arr.ndim == 3:
                    # (bands, H, W) を (H, W, bands) に補正
                    if arr.shape[0] in (1, 2, 3, 4) and arr.shape[-1] not in (1, 2, 3, 4):
                        arr = np.transpose(arr, (1, 2, 0))
                # min-max で 8bit 化
                if arr.dtype != np.uint8:
                    arrf = arr.astype(np.float32)
                    mn = float(np.nanmin(arrf))
                    mx = float(np.nanmax(arrf))
                    if mx > mn:
                        arr8 = ((arrf - mn) / (mx - mn) * 255.0).clip(0, 255).astype(np.uint8)
                    else:
                        # 同次元のゼロ配列を返す（2D なら(H,W)、3Dなら(H,W,C)）
                        if arr.ndim == 2:
                            arr8 = np.zeros(arr.shape, dtype=np.uint8)
                        else:
                            arr8 = np.zeros(arr.shape, dtype=np.uint8)
                else:
                    arr8 = arr
                # チャンネル処理（Pillowが扱えるRGBへ）
                if arr8.ndim == 2:
                    preview_img = Image.fromarray(arr8).convert('RGB')
                elif arr8.ndim == 3:
                    c = arr8.shape[2]
                    if c == 1:
                        preview_img = Image.fromarray(arr8[..., 0]).convert('RGB')
                    elif c == 2:
                        # 2chは Band1 グレースケールRGBで可視化（緑かぶり防止）
                        c0 = arr8[..., 0]
                        arr3 = np.dstack([c0, c0, c0])
                        preview_img = Image.fromarray(arr3)
                    elif c >= 3:
                        preview_img = Image.fromarray(arr8[..., :3])
                    else:
                        preview_img = Image.fromarray(arr8[..., 0]).convert('RGB')
                else:
                    preview_img = None
            except Exception:
                preview_img = None
            # 2) rasterio で縮小読み取り
            if preview_img is None:
                try:
                    import rasterio, numpy as np
                    from rasterio.enums import Resampling
                    with rasterio.open(self.ortho_path) as ds:
                        w, h = ds.width, ds.height
                        max_side = 2000
                        scale = min(1.0, max_side / float(max(w, h)))
                        out_w = max(1, int(w * scale))
                        out_h = max(1, int(h * scale))
                        arr = ds.read(out_shape=(min(ds.count, 3), out_h, out_w), resampling=Resampling.bilinear)
                        arr = arr.transpose(1, 2, 0)
                        # 正規化
                        arr = arr.astype(np.float32)
                        mn = float(np.nanmin(arr))
                        mx = float(np.nanmax(arr))
                        if mx > mn:
                            arr8 = ((arr - mn) / (mx - mn) * 255.0).clip(0, 255).astype(np.uint8)
                        else:
                            arr8 = np.zeros((out_h, out_w, 3), dtype=np.uint8)
                        if len(arr8.shape) == 3 and arr8.shape[2] == 2:
                            arr8 = np.dstack([arr8[...,0], arr8[...,0], arr8[...,0]])
                        preview_img = Image.fromarray(arr8)
                except Exception:
                    preview_img = None
            # 3) Pillow（PNGや非GeoTIFFの場合はそのまま）
            if preview_img is None:
                try:
                    with Image.open(self.ortho_path) as img:
                        w, h = img.size
                        max_side = 2000
                        scale = min(1.0, max_side / max(w, h))
                        preview_img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
                except Exception:
                    preview_img = None
            if preview_img is not None:
                self.coverage_image = preview_img.convert('RGB')
                self.coverage_image_path = self.ortho_path
            else:
                self.coverage_image = None
                self.coverage_image_path = None

        # 4) 座標ファイル
        geo_candidates = [
            os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_model_geo.txt'),
            os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_model_geo.csv'),
            os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_model.txt'),
            os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_model.csv'),
            os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_utm.txt'),
        ]
        geo_ref_path = next((p for p in geo_candidates if os.path.exists(p)), None)
        if not geo_ref_path:
            raise FileNotFoundError(os.path.join(self.webodm_path, 'odm_georeferencing', 'odm_georeferencing_model_geo.txt'))

        # 5) ファイル読込とXY抽出
        raw_positions = []
        with open(geo_ref_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                fname, x, y = extract_filename_and_xy(line)
                if x is None or y is None:
                    continue
                img_path = resolve_image_path(self.webodm_path, fname) if fname else None
                raw_positions.append({'filename': os.path.basename(fname) if fname else '', 'path': img_path, 'wx': x, 'wy': y})

        # 6) ピクセル座標へ
        self.image_positions = []
        ortho_w, ortho_h = self.ortho_image_size
        if geotransform:
            for p in raw_positions:
                res = world_to_pixel(geotransform, p['wx'], p['wy'])
                if not res:
                    continue
                px, py = res
                self.image_positions.append({'filename': p['filename'], 'path': p['path'], 'x': float(px), 'y': float(py)})
        else:
            xs = [p['wx'] for p in raw_positions]
            ys = [p['wy'] for p in raw_positions]
            if xs and ys:
                minx, maxx = min(xs), max(xs)
                miny, maxy = min(ys), max(ys)
                dx = max(maxx - minx, 1e-6)
                dy = max(maxy - miny, 1e-6)
                for p in raw_positions:
                    px = (p['wx'] - minx) / dx * ortho_w
                    py = (maxy - p['wy']) / dy * ortho_h  # y軸反転
                    self.image_positions.append({'filename': p['filename'], 'path': p['path'], 'x': float(px), 'y': float(py)})

        # 7) 表示
        self.display_coverage_image()
        self.update_info()

    except FileNotFoundError as e:
        messagebox.showerror("エラー", f"必要なWebODMファイルが見つかりません: {e.filename}")
        self.window.destroy()
    except Exception as e:
        messagebox.showerror("エラー", f"WebODMアセットの読み込みに失敗しました: {e}")
        self.window.destroy()

# 既存メソッドを差し替え
ODMImageSelector.load_webodm_assets = load_webodm_assets_robust

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None

# 管理レベルと拡張出力用の定数/ユーティリティ
MANAGEMENT_LEVELS = ['S', 'A', 'B', 'N']

DEFECT_V2_HEADERS = [
    'エリア(工区)','接続箱No.','接続箱','回路No.','アレイ番号','モジュール場所',
    '不良分類','管理レベル','シリアルナンバー(交換前)','報告書番号','離線した日','離線場所',
    'シリアルナンバー(交換後)','モジュール','列1','送電完了日','報告書番号2','備考',
    '報告年月日','採番'
]

def normalize_management_level(value):
    if not value:
        return 'S'
    v = str(value).upper().strip()
    return v if v in MANAGEMENT_LEVELS else 'S'

class ThermalVisibleFileDialog:
    """サーモ画像と可視画像を同時に選択するカスタムダイアログ"""

    IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
    THUMBNAIL_SIZE = (1000, 1000)
    PREVIEW_BASE_SIZE = (1000, 650)
    PREVIEW_MIN_ZOOM = 0.25
    PREVIEW_MAX_ZOOM = 3.0

    def __init__(self, parent, initial_dir=None, title="サーモ画像同時選択"):
        self.parent = parent
        self.result = None
        self.thermal_photo = None
        self.visible_photo = None
        self.display_mode = "list"
        self.files_in_current_dir = []
        self.thumbnail_cache = {}
        self.thumbnail_buttons = {}
        self.thumbnail_columns = 3
        self.selected_file = None
        self._suppress_listbox_event = False
        self.preview_zoom = 1.0
        self._preview_refresh_job = None
        self._updating_zoom_var = False

        self.layout_config_path = Path.home() / ".ortho_annotation_system_v7" / "layout.json"
        self.layout_preferences = {"main_ratio": 0.5, "preview_ratio": 1.0, "last_page": 1}
        self.load_layout_preferences()
        self.zoom_options = [
            ("50%", 0.5),
            ("75%", 0.75),
            ("100%", 1.0),
            ("125%", 1.25),
            ("150%", 1.5),
            ("200%", 2.0),
            ("300%", 3.0),
        ]
        self.page_size_options = [10, 50, 500, 400]
        self.page_size = 100
        # 前回保存されたページ番号で初期化（後でload_layout_preferencesから設定される）
        self.current_page = self.layout_preferences.get("last_page", 1)

        if initial_dir:
            try:
                init_path = Path(initial_dir).expanduser()
            except Exception:
                init_path = Path.home()
        else:
            init_path = Path.home()

        if not init_path.exists() or not init_path.is_dir():
            init_path = Path.home()
        self.current_dir = init_path

        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("1920x1100")
        self.window.transient(parent)
        self.window.grab_set()
        self.window.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.window.bind("<Configure>", self.on_window_configure)

        self.dir_var = tk.StringVar(value=str(self.current_dir))

        self.create_widgets()
        self.populate_file_list()
        self.bind_keyboard_shortcuts()
        self.window.after(200, self.restore_layout_preferences)

    def create_widgets(self):
        dir_frame = ttk.Frame(self.window)
        dir_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(dir_frame, text="フォルダ:").pack(side=tk.LEFT)
        ttk.Entry(dir_frame, textvariable=self.dir_var, state="readonly").pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 10)
        )
        ttk.Button(dir_frame, text="フォルダ選択", command=self.select_directory).pack(side=tk.LEFT)

        mode_frame = ttk.Frame(self.window)
        mode_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        self.list_view_button = ttk.Button(mode_frame, text="ファイル名表示", command=lambda: self.set_display_mode("list"))
        self.thumbnail_view_button = ttk.Button(mode_frame, text="サムネイル表示", command=lambda: self.set_display_mode("thumbnail"))
        self.list_view_button.pack(side=tk.LEFT, padx=(0, 5))
        self.thumbnail_view_button.pack(side=tk.LEFT)

        pagination_frame = ttk.Frame(self.window)
        pagination_frame.pack(fill=tk.X, padx=10, pady=(0, 8))
        ttk.Label(pagination_frame, text="表示件数:").pack(side=tk.LEFT)
        self.page_size_var = tk.StringVar(value=str(self.page_size))
        self.page_size_combo = ttk.Combobox(
            pagination_frame,
            textvariable=self.page_size_var,
            state="readonly",
            width=5,
            values=[str(v) for v in self.page_size_options]
        )
        self.page_size_combo.pack(side=tk.LEFT, padx=(5, 10))
        self.page_size_combo.bind("<<ComboboxSelected>>", self.on_page_size_change)
        ttk.Frame(pagination_frame).pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.prev_page_button = ttk.Button(pagination_frame, text="前へ", width=6, command=lambda: self.change_page(-1))
        self.prev_page_button.pack(side=tk.LEFT, padx=(0, 5))
        # ページ番号指定ボックス
        ttk.Label(pagination_frame, text="ページ:").pack(side=tk.LEFT, padx=(0, 2))
        self.page_jump_var = tk.StringVar(value="1")
        self.page_jump_entry = ttk.Entry(pagination_frame, textvariable=self.page_jump_var, width=6, justify="center")
        self.page_jump_entry.pack(side=tk.LEFT, padx=(0, 2))
        self.page_jump_entry.bind("<Return>", lambda e: self.jump_to_page())
        self.page_jump_button = ttk.Button(pagination_frame, text="移動", width=4, command=self.jump_to_page)
        self.page_jump_button.pack(side=tk.LEFT, padx=(0, 5))
        self.next_page_button = ttk.Button(pagination_frame, text="次へ", width=6, command=lambda: self.change_page(1))
        self.next_page_button.pack(side=tk.LEFT, padx=(0, 10))
        self.page_info_var = tk.StringVar(value="")
        ttk.Label(pagination_frame, textvariable=self.page_info_var).pack(side=tk.LEFT)

        self.main_paned = ttk.Panedwindow(self.window, orient=tk.HORIZONTAL)
        self.main_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.main_paned.bind("<ButtonRelease-1>", lambda _evt: self.schedule_preview_refresh())

        list_frame = ttk.LabelFrame(self.main_paned, text="サーモ画像ファイル")
        self.main_paned.add(list_frame, weight=1)
        self.list_container = ttk.Frame(list_frame)
        self.list_container.pack(fill=tk.BOTH, expand=True)

        self.listbox_frame = ttk.Frame(self.list_container)
        self.listbox_frame.pack(fill=tk.BOTH, expand=True)
        self.file_listbox = tk.Listbox(self.listbox_frame, exportselection=False)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scrollbar = ttk.Scrollbar(self.listbox_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_listbox.configure(yscrollcommand=list_scrollbar.set)
        self.file_listbox.bind("<<ListboxSelect>>", self.on_file_select)

        self.thumbnail_frame = ttk.Frame(self.list_container)
        self.thumbnail_canvas = tk.Canvas(self.thumbnail_frame, highlightthickness=0)
        self.thumbnail_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.thumbnail_scrollbar = ttk.Scrollbar(
            self.thumbnail_frame, orient=tk.VERTICAL, command=self.thumbnail_canvas.yview
        )
        self.thumbnail_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.thumbnail_canvas.configure(yscrollcommand=self.thumbnail_scrollbar.set)
        self.thumbnail_inner = ttk.Frame(self.thumbnail_canvas)
        self.thumbnail_window = self.thumbnail_canvas.create_window((0, 0), window=self.thumbnail_inner, anchor="nw")
        self.thumbnail_inner.bind(
            "<Configure>",
            lambda _: self.thumbnail_canvas.configure(scrollregion=self.thumbnail_canvas.bbox("all"))
        )
        self.thumbnail_canvas.bind("<Configure>", self._on_thumbnail_canvas_configure)
        self.thumbnail_frame.pack_forget()

        preview_wrapper = ttk.Frame(self.main_paned)
        self.main_paned.add(preview_wrapper, weight=1)

        preview_frame = ttk.LabelFrame(preview_wrapper, text="プレビュー")
        preview_frame.pack(fill=tk.BOTH, expand=True)

        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)

        preview_container = ttk.Frame(preview_frame)
        preview_container.grid(row=0, column=0, sticky="nsew", padx=5, pady=(5, 0))
        preview_container.columnconfigure(0, weight=1)
        preview_container.rowconfigure(0, weight=1)

        self.preview_canvas = tk.Canvas(preview_container, highlightthickness=0)
        self.preview_canvas.grid(row=0, column=0, sticky="nsew")
        preview_scrollbar = ttk.Scrollbar(preview_container, orient=tk.VERTICAL, command=self.preview_canvas.yview)
        preview_scrollbar.grid(row=0, column=1, sticky="ns")
        self.preview_canvas.configure(yscrollcommand=preview_scrollbar.set)

        self.preview_inner = ttk.Frame(self.preview_canvas)
        self.preview_window = self.preview_canvas.create_window((0, 0), window=self.preview_inner, anchor="nw")
        self.preview_inner.bind(
            "<Configure>",
            lambda _: self.preview_canvas.configure(scrollregion=self.preview_canvas.bbox("all"))
        )
        self.preview_canvas.bind("<Configure>", self._on_preview_canvas_configure)

        self.preview_vertical_paned = ttk.Panedwindow(self.preview_inner, orient=tk.VERTICAL)
        self.preview_vertical_paned.pack(fill=tk.BOTH, expand=True)
        self.preview_vertical_paned.bind("<ButtonRelease-1>", lambda _evt: self.schedule_preview_refresh())

        thermal_section = ttk.Frame(self.preview_vertical_paned)
        self.preview_vertical_paned.add(thermal_section, weight=1)
        thermal_frame = ttk.LabelFrame(thermal_section, text="サーモ画像")
        thermal_frame.pack(fill=tk.BOTH, expand=True)
        self.thermal_preview = tk.Label(thermal_frame, text="ファイルを選択してください", anchor="center", justify="center")
        self.thermal_preview.pack(fill=tk.BOTH, expand=True)

        visible_section = ttk.Frame(self.preview_vertical_paned)
        self.preview_vertical_paned.add(visible_section, weight=1)
        visible_frame = ttk.LabelFrame(visible_section, text="可視画像")
        visible_frame.pack(fill=tk.BOTH, expand=True)
        self.visible_preview = tk.Label(visible_frame, text="対応する可視画像を表示します", anchor="center", justify="center")
        self.visible_preview.pack(fill=tk.BOTH, expand=True)

        controls_frame = ttk.Frame(preview_frame)
        controls_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=(0, 5))
        self.zoom_out_button = ttk.Button(
            controls_frame,
            text="🔍－",
            width=4,
            command=lambda: self.adjust_preview_zoom(factor=0.8)
        )
        self.zoom_out_button.pack(side=tk.LEFT)
        self.zoom_in_button = ttk.Button(
            controls_frame,
            text="🔍＋",
            width=4,
            command=lambda: self.adjust_preview_zoom(factor=1.25)
        )
        self.zoom_in_button.pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(controls_frame, text="倍率:").pack(side=tk.LEFT, padx=(10, 5))
        self.zoom_var = tk.StringVar()
        self.zoom_combo = ttk.Combobox(
            controls_frame,
            textvariable=self.zoom_var,
            state="readonly",
            width=7,
            values=[label for label, _ in self.zoom_options]
        )
        self.zoom_combo.pack(side=tk.LEFT)
        self.zoom_combo.bind("<<ComboboxSelected>>", self.on_zoom_combo_change)

        self.visible_status_var = tk.StringVar(value="")
        ttk.Label(preview_frame, textvariable=self.visible_status_var, foreground="#666666").grid(
            row=2, column=0, sticky="ew", padx=5, pady=(0, 5)
        )

        button_frame = ttk.Frame(self.window)
        button_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        ttk.Button(button_frame, text="決定", command=self.on_confirm).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(button_frame, text="キャンセル", command=self.on_cancel).pack(side=tk.RIGHT)

        self.update_mode_buttons()
        self.update_zoom_display()

    def populate_file_list(self, rescan=True):
        if rescan or not hasattr(self, "files_in_current_dir"):
            if not self.current_dir.exists():
                self.files_in_current_dir = []
            else:
                self.files_in_current_dir = [
                    p for p in sorted(self.current_dir.iterdir())
                    if p.is_file() and p.suffix.lower() in self.IMAGE_EXTS and p.stem[-1:].upper() == "T"
                ]
        self.ensure_current_page_valid()
        self.refresh_views()
        self.ensure_selection_in_current_page()

    def select_directory(self):
        selected_dir = filedialog.askdirectory(parent=self.window, initialdir=str(self.current_dir))
        if selected_dir:
            path = Path(selected_dir)
            if path.exists():
                self.current_dir = path
                self.dir_var.set(str(self.current_dir))
                self.selected_file = None
                self.current_page = 1
                self.thumbnail_cache.clear()
                self.populate_file_list()

    def clear_previews(self):
        self.thermal_photo = None
        self.visible_photo = None
        self.thermal_preview.config(image="", text="ファイルを選択してください")
        self.visible_preview.config(image="", text="対応する可視画像を表示します")
        self.visible_status_var.set("")
        if hasattr(self, "preview_canvas"):
            self.preview_canvas.yview_moveto(0)

    def set_display_mode(self, mode: str):
        if mode not in ("list", "thumbnail"):
            return
        if mode == self.display_mode:
            return
        self.display_mode = mode
        if mode == "list":
            self.thumbnail_frame.pack_forget()
            self.listbox_frame.pack(fill=tk.BOTH, expand=True)
        else:
            self.listbox_frame.pack_forget()
            self.thumbnail_frame.pack(fill=tk.BOTH, expand=True)
        self.update_mode_buttons()
        self.refresh_views()
        self.ensure_selection_in_current_page()

    def update_mode_buttons(self):
        if self.display_mode == "list":
            self.list_view_button.state(["disabled"])
            self.thumbnail_view_button.state(["!disabled"])
        else:
            self.thumbnail_view_button.state(["disabled"])
            self.list_view_button.state(["!disabled"])

    def bind_keyboard_shortcuts(self):
        if not hasattr(self, "window"):
            return
        self.window.bind("<Left>", self.on_key_left)
        self.window.bind("<Right>", self.on_key_right)
        self.window.bind("<Up>", self.on_key_up)
        self.window.bind("<Down>", self.on_key_down)

    def on_key_left(self, event):
        if self.display_mode != "thumbnail":
            return
        self.move_thumbnail_selection(delta_row=0, delta_col=-1)
        return "break"

    def on_key_right(self, event):
        if self.display_mode != "thumbnail":
            return
        self.move_thumbnail_selection(delta_row=0, delta_col=1)
        return "break"

    def on_key_up(self, event):
        if self.display_mode != "thumbnail":
            return
        self.move_thumbnail_selection(delta_row=-1, delta_col=0)
        return "break"

    def on_key_down(self, event):
        if self.display_mode != "thumbnail":
            return
        self.move_thumbnail_selection(delta_row=1, delta_col=0)
        return "break"

    def move_thumbnail_selection(self, delta_row=0, delta_col=0):
        page_files = self.get_page_files()
        if not page_files:
            return
        columns = max(1, self.thumbnail_columns)
        if self.selected_file in page_files:
            current_index = page_files.index(self.selected_file)
        else:
            current_index = 0
        new_index = current_index + delta_row * columns + delta_col
        new_index = max(0, min(new_index, len(page_files) - 1))
        target = page_files[new_index]
        self.set_selected_file(target)

    def load_layout_preferences(self):
        try:
            if self.layout_config_path.exists():
                with self.layout_config_path.open("r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    main_ratio = data.get("main_ratio")
                    preview_ratio = data.get("preview_ratio")
                    last_page = data.get("last_page")
                    if isinstance(main_ratio, (int, float)):
                        self.layout_preferences["main_ratio"] = float(main_ratio)
                    if isinstance(preview_ratio, (int, float)):
                        self.layout_preferences["preview_ratio"] = float(preview_ratio)
                    if isinstance(last_page, int) and last_page >= 1:
                        self.layout_preferences["last_page"] = int(last_page)
        except Exception:
            pass

    def restore_layout_preferences(self, attempt=0):
        if not hasattr(self, "main_paned") or not hasattr(self, "preview_vertical_paned"):
            return
        main_ratio = float(self.layout_preferences.get("main_ratio", 0.5))
        preview_ratio = float(self.layout_preferences.get("preview_ratio", 0.5))
        main_ratio = max(0.1, min(0.9, main_ratio))
        preview_ratio = max(0.1, min(0.9, preview_ratio))
        # 保存済みの preview_ratio をそのまま使う（固定上書きを廃止）
        main_width = self.main_paned.winfo_width()
        preview_height = self.preview_vertical_paned.winfo_height()
        if (main_width <= 1 or preview_height <= 1) and attempt < 10:
            self.window.after(200, lambda: self.restore_layout_preferences(attempt + 1))
            return
        try:
            if main_width > 1:
                self.main_paned.sashpos(0, int(main_width * main_ratio))
        except tk.TclError:
            pass
        try:
            if preview_height > 1:
                self.preview_vertical_paned.sashpos(0, int(preview_height * preview_ratio))
                # 固定位置の上書きを撤廃（比率ベースの配置に一本化）
        except tk.TclError:
            pass

    def save_layout_preferences(self):
        if not hasattr(self, "main_paned") or not hasattr(self, "preview_vertical_paned"):
            return
        try:
            main_width = self.main_paned.winfo_width()
            if main_width > 0:
                pos = self.main_paned.sashpos(0)
                if isinstance(pos, int):
                    self.layout_preferences["main_ratio"] = max(0.1, min(0.9, pos / main_width))
            preview_height = self.preview_vertical_paned.winfo_height()
            if preview_height > 0:
                pos_v = self.preview_vertical_paned.sashpos(0)
                if isinstance(pos_v, int):
                    self.layout_preferences["preview_ratio"] = max(0.1, min(0.9, pos_v / preview_height))
        except tk.TclError:
            pass
        # 現在のページ番号を保存
        try:
            if hasattr(self, "current_page") and isinstance(self.current_page, int):
                self.layout_preferences["last_page"] = self.current_page
        except Exception:
            pass
        try:
            self.layout_config_path.parent.mkdir(parents=True, exist_ok=True)
            with self.layout_config_path.open("w", encoding="utf-8") as f:
                json.dump(self.layout_preferences, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def refresh_views(self):
        self.refresh_listbox()
        self.refresh_thumbnail_view()
        self.update_pagination_controls()

    def get_total_pages(self):
        if not hasattr(self, "files_in_current_dir") or not self.files_in_current_dir:
            return 0
        if self.page_size <= 0:
            return 1
        return math.ceil(len(self.files_in_current_dir) / self.page_size)

    def get_page_files(self):
        if not hasattr(self, "files_in_current_dir") or not self.files_in_current_dir:
            return []
        if self.page_size <= 0:
            return list(self.files_in_current_dir)
        total_pages = self.get_total_pages()
        if total_pages == 0:
            return []
        page = max(1, min(self.current_page, total_pages))
        start = (page - 1) * self.page_size
        end = start + self.page_size
        return self.files_in_current_dir[start:end]

    def ensure_current_page_valid(self):
        total_pages = self.get_total_pages()
        if total_pages == 0:
            self.current_page = 1
            if getattr(self, "selected_file", None) is not None:
                self.selected_file = None
            return
        self.current_page = max(1, min(self.current_page, total_pages))
        if self.page_size > 0 and self.selected_file and self.selected_file in self.files_in_current_dir:
            page_index = self.files_in_current_dir.index(self.selected_file) // self.page_size + 1
            if page_index != self.current_page:
                self.current_page = page_index
        elif self.selected_file and self.selected_file not in self.files_in_current_dir:
            self.selected_file = None

    def ensure_selection_in_current_page(self):
        page_files = self.get_page_files()
        if not page_files:
            self.set_selected_file(None)
            return
        if self.selected_file in page_files:
            self.set_selected_file(self.selected_file)
        else:
            self.set_selected_file(page_files[0])

    def update_pagination_controls(self):
        total_files = len(self.files_in_current_dir) if hasattr(self, "files_in_current_dir") else 0
        total_pages = self.get_total_pages()
        if hasattr(self, "page_size_var") and self.page_size_var.get() != str(self.page_size):
            self.page_size_var.set(str(self.page_size))
        # ページ番号入力ボックスの値を現在のページ番号に更新
        if hasattr(self, "page_jump_var") and self.page_jump_var.get() != str(self.current_page):
            self.page_jump_var.set(str(self.current_page))
        if total_pages == 0:
            page_text = "ページ 0 / 0 (0件)"
        else:
            page_text = f"ページ {self.current_page} / {total_pages} ({total_files}件)"
        if hasattr(self, "page_info_var"):
            self.page_info_var.set(page_text)
        if hasattr(self, "prev_page_button"):
            if total_pages <= 1 or self.current_page <= 1:
                self.prev_page_button.state(["disabled"])
            else:
                self.prev_page_button.state(["!disabled"])
        if hasattr(self, "next_page_button"):
            if total_pages <= 1 or self.current_page >= total_pages:
                self.next_page_button.state(["disabled"])
            else:
                self.next_page_button.state(["!disabled"])

    def change_page(self, delta):
        total_pages = self.get_total_pages()
        if total_pages <= 1:
            return
        new_page = max(1, min(self.current_page + delta, total_pages))
        if new_page == self.current_page:
            return
        self.current_page = new_page
        self.selected_file = None
        self.populate_file_list(rescan=False)

    def jump_to_page(self):
        """ページ番号指定ボックスからページにジャンプ"""
        try:
            target_page = int(self.page_jump_var.get())
        except (ValueError, TypeError):
            messagebox.showwarning("入力エラー", "ページ番号は数値で入力してください。", parent=self.window)
            self.page_jump_var.set(str(self.current_page))
            return
        
        total_pages = self.get_total_pages()
        if total_pages == 0:
            messagebox.showwarning("エラー", "表示可能なページがありません。", parent=self.window)
            self.page_jump_var.set("1")
            return
        
        if target_page < 1 or target_page > total_pages:
            messagebox.showwarning(
                "範囲外エラー", 
                f"ページ番号は 1 ～ {total_pages} の範囲で指定してください。", 
                parent=self.window
            )
            self.page_jump_var.set(str(self.current_page))
            return
        
        if target_page == self.current_page:
            return
        
        self.current_page = target_page
        self.selected_file = None
        self.populate_file_list(rescan=False)

    def on_page_size_change(self, event=None):
        try:
            new_size = int(self.page_size_var.get())
        except (ValueError, TypeError):
            self.page_size_var.set(str(self.page_size))
            return
        if new_size <= 0:
            new_size = self.page_size_options[0]
        if new_size == self.page_size:
            return
        self.page_size = new_size
        self.current_page = 1
        self.populate_file_list(rescan=False)

    def refresh_listbox(self):
        self._suppress_listbox_event = True
        self.file_listbox.delete(0, tk.END)
        page_files = self.get_page_files()
        for file_path in page_files:
            self.file_listbox.insert(tk.END, file_path.name)
        if self.selected_file and self.selected_file in page_files:
            index = page_files.index(self.selected_file)
            self.file_listbox.selection_set(index)
            self.file_listbox.see(index)
        self._suppress_listbox_event = False

    def refresh_thumbnail_view(self):
        for child in self.thumbnail_inner.winfo_children():
            child.destroy()
        self.thumbnail_buttons.clear()
        page_files = self.get_page_files()
        if not page_files:
            ttk.Label(self.thumbnail_inner, text="サーモ画像が見つかりません").grid(row=0, column=0, padx=10, pady=10)
            bbox = self.thumbnail_canvas.bbox("all")
            self.thumbnail_canvas.configure(scrollregion=bbox if bbox else (0, 0, 0, 0))
            return
        for idx, file_path in enumerate(page_files):
            photo = self.get_thumbnail_image(file_path)
            btn = tk.Button(
                self.thumbnail_inner,
                image=photo,
                text=file_path.name,
                compound="top",
                wraplength=self.THUMBNAIL_SIZE[0] + 10,
                relief=tk.RAISED,
                bd=2,
                justify="center",
                command=lambda p=file_path: self.on_thumbnail_click(p)
            )
            btn.grid(row=idx // self.thumbnail_columns, column=idx % self.thumbnail_columns, padx=5, pady=5, sticky="nsew")
            self.thumbnail_buttons[file_path] = btn
        for col in range(self.thumbnail_columns):
            self.thumbnail_inner.grid_columnconfigure(col, weight=1)
        if self.selected_file:
            self.update_thumbnail_selection(self.selected_file)
        self.thumbnail_canvas.yview_moveto(0)
        bbox = self.thumbnail_canvas.bbox("all")
        self.thumbnail_canvas.configure(scrollregion=bbox if bbox else (0, 0, 0, 0))

    def _on_thumbnail_canvas_configure(self, event):
        self.thumbnail_canvas.itemconfigure(self.thumbnail_window, width=event.width)

    def _on_preview_canvas_configure(self, event):
        if hasattr(self, "preview_window"):
            self.preview_canvas.itemconfigure(self.preview_window, width=event.width)
            # 高さもキャンバスに追従
            try:
                self.preview_canvas.itemconfigure(self.preview_window, height=event.height)
            except tk.TclError:
                pass

    def on_file_select(self, event=None):
        if self._suppress_listbox_event:
            return
        selection = self.file_listbox.curselection()
        if not selection:
            self.set_selected_file(None)
            return
        index = selection[0]
        page_files = self.get_page_files()
        if index >= len(page_files):
            return
        self.set_selected_file(page_files[index])

    def on_thumbnail_click(self, file_path: Path):
        self.set_selected_file(file_path)

    def set_selected_file(self, thermal_path, **_):
        thermal_path = Path(thermal_path) if thermal_path is not None else None
        self.selected_file = thermal_path
        self.update_listbox_selection(thermal_path)
        self.update_thumbnail_selection(thermal_path)
        if thermal_path:
            self.update_previews(thermal_path)
        else:
            self.clear_previews()

    def update_listbox_selection(self, thermal_path):
        self._suppress_listbox_event = True
        self.file_listbox.selection_clear(0, tk.END)
        if thermal_path:
            page_files = self.get_page_files()
            if thermal_path in page_files:
                index = page_files.index(thermal_path)
                self.file_listbox.selection_set(index)
                self.file_listbox.see(index)
        self._suppress_listbox_event = False

    def update_thumbnail_selection(self, thermal_path):
        for path, button in self.thumbnail_buttons.items():
            if thermal_path and path == thermal_path:
                button.configure(relief=tk.SUNKEN, bd=3)
            else:
                button.configure(relief=tk.RAISED, bd=2)

    def update_previews(self, thermal_path: Path):
        thermal_path = Path(thermal_path)
        if not thermal_path.exists():
            self.thermal_photo = None
            self.visible_photo = None
            self.thermal_preview.config(image="", text="ファイルが存在しません")
            self.visible_preview.config(image="", text="対応する可視画像を表示します")
            self.visible_status_var.set("")
            return
        self.visible_status_var.set("")
        thermal_size = self.get_preview_size(self.thermal_preview)
        self.thermal_photo = self.load_preview_image(thermal_path, size=thermal_size, allow_upscale=True)
        if self.thermal_photo:
            self.thermal_preview.config(image=self.thermal_photo, text="")
        else:
            self.thermal_preview.config(image="", text="プレビューできません")

        visible_path = self.find_visible_counterpart(thermal_path)
        if visible_path and visible_path.exists():
            visible_size = self.get_preview_size(self.visible_preview)
            self.visible_photo = self.load_preview_image(visible_path, size=visible_size, allow_upscale=True)
            if self.visible_photo:
                self.visible_preview.config(image=self.visible_photo, text="")
            else:
                self.visible_preview.config(image="", text="プレビューできません")
            self.visible_status_var.set(f"可視画像: {visible_path.name}")
        else:
            self.visible_photo = None
            self.visible_preview.config(image="", text="同名ファイルが見つかりません")
            self.visible_status_var.set("")

        if hasattr(self, "preview_canvas"):
            self.preview_canvas.yview_moveto(0)

    def get_preview_size(self, target_widget):
        base_width, base_height = self.PREVIEW_BASE_SIZE
        if target_widget is not None:
            try:
                target_widget.update_idletasks()
            except Exception:
                pass
            current_width = target_widget.winfo_width()
            current_height = target_widget.winfo_height()
            if isinstance(current_width, int) and current_width > 1:
                base_width = max(base_width, current_width)
            if isinstance(current_height, int) and current_height > 1:
                base_height = max(base_height, current_height)
        return (
            max(1, int(base_width * self.preview_zoom)),
            max(1, int(base_height * self.preview_zoom)),
        )

    def load_preview_image(self, path: Path, size=None, allow_upscale=True):
        try:
            with Image.open(path) as img:
                img = img.copy()
                img = img.convert("RGBA")
            target_size = size or self.PREVIEW_BASE_SIZE
            max_w, max_h = target_size
            if max_w <= 0:
                max_w = 1
            if max_h <= 0:
                max_h = 1
            orig_w, orig_h = img.size
            if orig_w <= 0 or orig_h <= 0:
                return None
            scale_w = max_w / orig_w
            scale_h = max_h / orig_h
            scale = min(scale_w, scale_h)
            if not allow_upscale:
                scale = min(scale, 1.0)
            if scale <= 0:
                scale = 1.0
            new_width = max(1, int(orig_w * scale))
            new_height = max(1, int(orig_h * scale))
            resample = Image.Resampling.LANCZOS if scale <= 1.0 else Image.Resampling.BICUBIC
            img = img.resize((new_width, new_height), resample=resample)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    def get_thumbnail_image(self, path: Path):
        key = (str(path), self.THUMBNAIL_SIZE)
        if key in self.thumbnail_cache:
            return self.thumbnail_cache[key]
        try:
            with Image.open(path) as img:
                img = img.copy()
                img = img.convert("RGBA")
            img.thumbnail(self.THUMBNAIL_SIZE, Image.Resampling.LANCZOS)
        except Exception:
            img = Image.new("RGB", self.THUMBNAIL_SIZE, color="#cccccc")
        photo = ImageTk.PhotoImage(img)
        self.thumbnail_cache[key] = photo
        return photo

    def adjust_preview_zoom(self, factor=None, absolute=None):
        if absolute is not None:
            new_zoom = absolute
        elif factor is not None:
            new_zoom = self.preview_zoom * factor
        else:
            return
        new_zoom = max(self.PREVIEW_MIN_ZOOM, min(self.PREVIEW_MAX_ZOOM, new_zoom))
        if abs(new_zoom - self.preview_zoom) < 1e-3:
            return
        self.preview_zoom = new_zoom
        self.update_zoom_display()
        self.refresh_current_previews()

    def update_zoom_display(self):
        if not hasattr(self, "zoom_var"):
            return
        label = None
        for option_label, value in self.zoom_options:
            if abs(value - self.preview_zoom) < 1e-3:
                label = option_label
                break
        if label is None:
            label = f"{int(self.preview_zoom * 100)}%"
        self._updating_zoom_var = True
        self.zoom_var.set(label)
        self._updating_zoom_var = False

        if hasattr(self, "zoom_out_button"):
            if self.preview_zoom <= self.PREVIEW_MIN_ZOOM + 1e-3:
                self.zoom_out_button.state(["disabled"])
            else:
                self.zoom_out_button.state(["!disabled"])
        if hasattr(self, "zoom_in_button"):
            if self.preview_zoom >= self.PREVIEW_MAX_ZOOM - 1e-3:
                self.zoom_in_button.state(["disabled"])
            else:
                self.zoom_in_button.state(["!disabled"])

    def on_zoom_combo_change(self, event=None):
        if self._updating_zoom_var:
            return
        value_str = self.zoom_var.get().replace("%", "").strip()
        try:
            value = float(value_str) / 100.0
        except ValueError:
            self.update_zoom_display()
            return
        self.adjust_preview_zoom(absolute=value)

    def refresh_current_previews(self):
        self._preview_refresh_job = None
        if self.selected_file:
            self.update_previews(self.selected_file)

    def schedule_preview_refresh(self, delay=120):
        if not hasattr(self, "window"):
            return
        if getattr(self, "_preview_refresh_job", None) is not None:
            try:
                self.window.after_cancel(self._preview_refresh_job)
            except Exception:
                pass
        try:
            self._preview_refresh_job = self.window.after(delay, self.refresh_current_previews)
        except Exception:
            self._preview_refresh_job = None

    def on_window_configure(self, event=None):
        if event is not None and event.widget is not self.window:
            return
        self.schedule_preview_refresh()

    @staticmethod
    def find_visible_counterpart(thermal_path: Path):
        try:
            stem = thermal_path.stem
            if not stem:
                return None
            if stem[-1].upper() != "T":
                return None
            counterpart_stem = stem[:-1] + "V"
            visible_dir = thermal_path.parent.parent / "可視画像"
            visible_path = visible_dir / f"{counterpart_stem}{thermal_path.suffix}"
            if visible_path.exists():
                return visible_path
            return None
        except Exception:
            return None

    def on_confirm(self):
        if not self.selected_file:
            messagebox.showwarning("警告", "サーモ画像を選択してください。", parent=self.window)
            return
        thermal_path = Path(self.selected_file)
        visible_path = self.find_visible_counterpart(thermal_path)
        if not visible_path or not visible_path.exists():
            messagebox.showwarning("警告", "同名ファイルが見つかりません。", parent=self.window)
            return
        self.result = (str(thermal_path), str(visible_path))
        self.save_layout_preferences()
        self.window.destroy()

    def on_cancel(self):
        self.result = None
        self.save_layout_preferences()
        self.window.destroy()

    def show(self):
        self.window.wait_window()
        return self.result


class OrthoImageAnnotationSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("太陽光発電所 オルソ画像アノテーションシステム v2.2")
        self.root.geometry("1200x900")


        # 変数の初期化
        self.current_image = None
        self.image_path = None
        self.annotations = []
        self.next_id = 1
        self.project_name = ""
        self.project_path = ""
        self.canvas_image = None
        self.zoom_factor = 1.0
        self.pan_start_x = 0
        self.pan_start_y = 0
        self.webodm_path = None  # WebODMフォルダのパスを保持する変数を追加
        self.last_thermal_visible_dir = None  # サーモ・可視画像同時選択用の直近フォルダ

        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.annotation_icon_dir = os.path.join(base_dir, "アノテーション画像フォルダ")
        self.annotation_icon_cache = {}
        self.annotation_icon_tk_cache = {}
        self.canvas_icon_refs = []  # キャンバス上でTkイメージを保持
        self.annotation_default_icon_size = 64
        self._icon_warning_shown = False
        self._icon_dir_missing_warned = False
        self.missing_icon_types = set()

        # 不具合分類と色の設定
        self.defect_types = {
            "ホットスポット": "#FF0000",  # 赤
            "クラスタ異常": "#FF8C00",    # オレンジ
            "破損": "#FFD700",           # 黄
            "ストリング異常": "#0000FF",  # 青
            "系統異常": "#8A2BE2",       # 紫
            "影": "#008000"              # 緑
        }

        # アノテーション形状の設定
        self.annotation_shapes = {
            "十字": "cross",
            "矢印": "arrow",
            "円形": "circle",
            "四角": "rectangle"
        }

        self.annotation_scale_vars = {
            "overall": tk.StringVar(value="1.0"),
            "thermal": tk.StringVar(value="1.0"),
            "visible": tk.StringVar(value="1.0"),
        }
        self.scale_entries = {}
        
        # 画像拡張設定（個別全体図保存時に使用）
        self.image_extension_ratio = 1/3  # デフォルト: 元画像の高さの1/3を上下に追加
        
        # ズームコントロール用の変数
        self.zoom_options = [
            ("25%", 0.25),
            ("50%", 0.5),
            ("75%", 0.75),
            ("100%", 1.0),
            ("125%", 1.25),
            ("150%", 1.5),
            ("200%", 2.0),
            ("300%", 3.0),
            ("400%", 4.0),
            ("500%", 5.0),
        ]
        self.zoom_var = tk.StringVar(value="100%")
        self._updating_zoom_var = False

        # アノテーション移動機能用の状態変数
        self.move_mode = False              # 移動モードのON/OFF
        self.moving_annotation = None       # 現在移動中のアノテーション
        self.move_start_pos = None          # ドラッグ開始位置（Ctrl+ドラッグ用）
        self.move_original_pos = None       # 元の位置（キャンセル用）

        # アノテーション位置オフセット設定（サーモ画像・可視画像・オルソ画像用）
        self.thermal_offset_x = 0           # サーモ画像のX軸オフセット（ピクセル）
        self.thermal_offset_y = 0           # サーモ画像のY軸オフセット（ピクセル）
        self.visible_offset_x = 0           # 可視画像のX軸オフセット（ピクセル）
        self.visible_offset_y = 0           # 可視画像のY軸オフセット（ピクセル）
        self.ortho_offset_x = 0             # オルソ画像（全体図）のX軸オフセット（ピクセル）
        self.ortho_offset_y = 0             # オルソ画像（全体図）のY軸オフセット（ピクセル）

        # 差分保存機能（Requirement 5）
        self.last_saved_snapshot = None     # 前回保存時のスナップショット
        self.has_unsaved_changes = False    # 未保存の変更フラグ
        self.changed_annotation_ids = set() # 変更されたアノテーションID

        self.initialize_annotation_icons()

        self.setup_ui()
        self.create_or_load_project()

    def setup_ui(self):
        # メインフレーム
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 画像表示エリア
        image_frame = ttk.LabelFrame(main_frame, text="オルソ画像表示", padding=10)
        image_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # キャンバスとスクロールバー
        canvas_frame = ttk.Frame(image_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(canvas_frame, bg="white", width=800, height=600)
        v_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        h_scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)

        self.canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        self.canvas.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        # キャンバスイベント
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
        self.canvas.bind("<Button-3>", self.on_canvas_right_click)
        self.canvas.bind("<MouseWheel>", self.on_mouse_wheel)
        self.canvas.bind("<Button-2>", self.start_pan)
        self.canvas.bind("<B2-Motion>", self.do_pan)
        # Ctrl+ドラッグでアノテーション移動
        self.canvas.bind("<Control-Button-1>", self.on_ctrl_click)
        self.canvas.bind("<Control-B1-Motion>", self.on_ctrl_drag)
        self.canvas.bind("<Control-ButtonRelease-1>", self.on_ctrl_release)

        # コントロールパネル
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))

        # ボタン群
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ttk.Button(button_frame, text="画像選択", command=self.select_image).pack(side=tk.LEFT, padx=(0, 5))
        # WebODMフォルダ選択ボタンを追加
        self.webodm_button = ttk.Button(button_frame, text="WebODMフォルダ選択", command=self.select_webodm_folder)
        self.webodm_button.pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="画像リセット", command=self.reset_image).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="保存", command=self.save_project).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="中止", command=self.quit_application).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="色設定", command=self.customize_settings).pack(side=tk.LEFT, padx=(0, 5))

        scale_frame = ttk.Frame(button_frame)
        scale_frame.pack(side=tk.LEFT, padx=(5, 0))

        scale_config = [
            ("全体", "overall"),
            ("サーモ", "thermal"),
            ("可視", "visible"),
        ]
        for label, key in scale_config:
            ttk.Label(scale_frame, text=f"{label}×").pack(side=tk.LEFT)
            entry = ttk.Entry(scale_frame, width=5, textvariable=self.annotation_scale_vars[key], justify="right")
            entry.pack(side=tk.LEFT, padx=(0, 5))
            self.scale_entries[key] = entry
        
        # ズームコントロール
        zoom_frame = ttk.Frame(button_frame)
        zoom_frame.pack(side=tk.LEFT, padx=(10, 0))
        
        ttk.Label(zoom_frame, text="表示倍率:").pack(side=tk.LEFT)
        self.zoom_out_button = ttk.Button(zoom_frame, text="🔍-", width=3, command=self.zoom_out)
        self.zoom_out_button.pack(side=tk.LEFT, padx=(5, 0))
        
        self.zoom_combo = ttk.Combobox(
            zoom_frame,
            textvariable=self.zoom_var,
            values=[label for label, _ in self.zoom_options],
            state="readonly",
            width=7
        )
        self.zoom_combo.pack(side=tk.LEFT, padx=(5, 0))
        self.zoom_combo.bind("<<ComboboxSelected>>", self.on_zoom_combo_change)
        
        self.zoom_in_button = ttk.Button(zoom_frame, text="🔍+", width=3, command=self.zoom_in)
        self.zoom_in_button.pack(side=tk.LEFT, padx=(5, 0))

        # 設定選択エリア
        settings_frame = ttk.Frame(control_frame)
        settings_frame.pack(side=tk.RIGHT)

        # 不良分類選択
        ttk.Label(settings_frame, text="不良分類:").pack(side=tk.LEFT, padx=(0, 5))
        self.defect_var = tk.StringVar(value=list(self.defect_types.keys())[0])
        self.defect_combo = ttk.Combobox(settings_frame, textvariable=self.defect_var, 
                                        values=list(self.defect_types.keys()), 
                                        state="readonly", width=15)
        self.defect_combo.pack(side=tk.LEFT, padx=(0, 10))

        # 形状選択
        ttk.Label(settings_frame, text="形状:").pack(side=tk.LEFT, padx=(0, 5))
        self.shape_var = tk.StringVar(value=list(self.annotation_shapes.keys())[0])
        self.shape_combo = ttk.Combobox(settings_frame, textvariable=self.shape_var, 
                                       values=list(self.annotation_shapes.keys()), 
                                       state="readonly", width=10)
        self.shape_combo.pack(side=tk.LEFT)

        # 不具合テーブル
        table_frame = ttk.LabelFrame(main_frame, text="不具合テーブル", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True)

        # Treeview - 列を追加
        columns = ("ID", "エリア番号", "PCS番号", "接続箱番号", "回路番号", "アレイ№", "モジュール位置", "シリアル№", "不良分類", "形状", "サーモ画像", "可視画像", "備考", "報告年月日")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=8)

        # 列の設定
        for col in columns:
            self.tree.heading(col, text=col)
            if col == "ID":
                self.tree.column(col, width=50, anchor="center")
            elif col in ["エリア番号", "PCS番号", "接続箱番号", "回路番号", "アレイ№"]:
                self.tree.column(col, width=80, anchor="center")
            elif col in ["モジュール位置", "シリアル№"]:
                self.tree.column(col, width=100, anchor="center")
            elif col in ["不良分類", "形状"]:
                self.tree.column(col, width=100, anchor="center")
            elif col in ["サーモ画像", "可視画像"]:
                self.tree.column(col, width=120, anchor="center")
            elif col == "報告年月日":
                self.tree.column(col, width=100, anchor="center")
            else:
                self.tree.column(col, width=150, anchor="w")

    # （後半部分は変更なし）

        # スクロールバー
        tree_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # テーブルイベント
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Delete>", self.delete_annotation)

        # キーボードショートカット（ズーム操作）
        self.root.bind("<Control-plus>", lambda e: self.zoom_in())
        self.root.bind("<Control-equal>", lambda e: self.zoom_in())  # Shift無し+でも対応
        self.root.bind("<Control-minus>", lambda e: self.zoom_out())
        self.root.bind("<Control-Key-0>", lambda e: self.set_zoom_factor(1.0))  # 100%にリセット
        # Escキーでアノテーション移動をキャンセル
        self.root.bind("<Escape>", self.cancel_move)

    def select_webodm_folder(self):
        """WebODMのアセットフォルダを選択し、オルソ画像を自動で読み込む"""
        folder_path = filedialog.askdirectory(title="WebODMアセットフォルダを選択")
        if not folder_path:
            return

        # WebODMの主要なファイルの存在をチェック
        geo_ref_path = os.path.join(folder_path, 'odm_georeferencing', 'odm_georeferencing_model_geo.txt')
        coverage_path = os.path.join(folder_path, 'images', 'shot_coverage.png')
        ortho_path = os.path.join(folder_path, 'odm_orthophoto', 'odm_orthophoto.tif')

        if os.path.exists(geo_ref_path) and os.path.exists(ortho_path):
            self.webodm_path = folder_path
            messagebox.showinfo("成功", f"WebODMフォルダを設定しました:\n{folder_path}\nオルソ画像を自動で読み込みます。")
            # メインのオルソ画像を自動で読み込む
            self.load_image(ortho_path)
        else:
            messagebox.showerror("エラー", "選択されたフォルダに必要なWebODMアセットが見つかりません。\n- odm_georeferencing/odm_georeferencing_model_geo.(txt|csv)\n- odm_orthophoto/odm_orthophoto.(tif|tiff|png)\nを確認してください。")
            self.webodm_path = None


    def create_or_load_project(self):
        """新しいプロジェクトを作成または既存プロジェクトを読み込み"""
        dialog = tk.Toplevel(self.root)
        dialog.title("プロジェクト選択")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()

        # プロジェクト選択フレーム
        project_frame = ttk.LabelFrame(dialog, text="プロジェクト", padding=10)
        project_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 新規プロジェクト
        new_project_frame = ttk.Frame(project_frame)
        new_project_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(new_project_frame, text="新規プロジェクト名:").pack(side=tk.LEFT)
        self.project_name_entry = ttk.Entry(new_project_frame, width=20)
        self.project_name_entry.pack(side=tk.LEFT, padx=(5, 0))

        # ボタンフレーム
        button_frame = ttk.Frame(project_frame)
        button_frame.pack(fill=tk.X, pady=10)

        def create_new_project():
            project_name = self.project_name_entry.get().strip()
            if not project_name:
                project_name = f"Project_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            self.project_name = project_name
            self.project_path = os.path.join(os.getcwd(), self.project_name)
            
            # プロジェクトフォルダ作成
            os.makedirs(self.project_path, exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "アノテーション設定フォルダ"), exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "サーモ画像フォルダ"), exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "可視画像フォルダ"), exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "不具合一覧表フォルダ"), exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "アノテーション入り画像フォルダ"), exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "全体図位置フォルダ"), exist_ok=True)  # 追加

            self.root.title(f"太陽光発電所 オルソ画像アノテーションシステム v2.1.2 - {self.project_name}")
            dialog.destroy()

        def load_existing_project():
            folder_path = filedialog.askdirectory(title="既存プロジェクトフォルダを選択")
            if folder_path:
                self.project_path = folder_path
                self.project_name = os.path.basename(folder_path)
                
                # アノテーション設定フォルダが存在するかチェック
                annotation_folder = os.path.join(self.project_path, "アノテーション設定フォルダ")
                if os.path.exists(annotation_folder):
                    # 既存のアノテーションを読み込み
                    self.load_annotations()
                    
                    # 画像パスが保存されていれば自動で読み込み
                    if self.image_path and os.path.exists(self.image_path):
                        self.load_image(self.image_path)
                    
                    self.root.title(f"太陽光発電所 オルソ画像アノテーションシステム v2.1 - {self.project_name}")
                    dialog.destroy()
                else:
                    messagebox.showerror("エラー", "有効なプロジェクトフォルダではありません。")

        ttk.Button(button_frame, text="新規作成", command=create_new_project).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="既存プロジェクト読み込み", command=load_existing_project).pack(side=tk.LEFT)

    def select_image(self):
        """画像を選択して表示"""
        file_path = filedialog.askopenfilename(
            title="オルソ画像を選択",
            filetypes=[("画像ファイル", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif"), ("すべてのファイル", "*.*")]
        )

        if file_path:
            self.load_image(file_path)

    def load_image(self, file_path):
        """画像を読み込んで表示"""
        try:
            self.image_path = file_path
            self.current_image = Image.open(file_path)
            self.zoom_factor = 1.0
            self.display_image()

        except Exception as e:
            # Pillowで読めないGeoTIFFなどに対するフォールバック
            try:
                ext = os.path.splitext(file_path)[1].lower()
                if ext in ['.tif', '.tiff']:
                    # まずtifffileで試す（float/BIGTIFF/圧縮に強い）
                    pil_img = self._read_tiff_with_tifffile(file_path)
                    if pil_img is None:
                        # 次にOpenCV（日本語パス対策: np.fromfile + imdecode）
                        try:
                            data = np.fromfile(file_path, dtype=np.uint8)
                            img_cv = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
                        except Exception:
                            img_cv = cv2.imread(file_path, cv2.IMREAD_UNCHANGED)
                        if img_cv is not None:
                            # 16/32bit想定 → 8bit正規化
                            if img_cv.dtype != np.uint8:
                                img_cv = cv2.normalize(img_cv, None, 0, 255, cv2.NORM_MINMAX)
                                img_cv = img_cv.astype(np.uint8)
                            if len(img_cv.shape) == 2:
                                pil_img = Image.fromarray(img_cv).convert('RGB')
                            elif len(img_cv.shape) == 3:
                                c = img_cv.shape[2]
                                if c == 3:
                                    pil_img = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB))
                                elif c == 4:
                                    pil_img = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGRA2RGBA))
                                elif c > 4:
                                    pil_img = Image.fromarray(cv2.cvtColor(img_cv[:, :, :3], cv2.COLOR_BGR2RGB))
                    if pil_img is not None:
                        self.current_image = pil_img
                        self.zoom_factor = 1.0
                        self.display_image()
                        return
                # WebODMのプレビューへフォールバック
                if getattr(self, 'webodm_path', None):
                    for p in [
                        os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto.png'),
                        os.path.join(self.webodm_path, 'odm_orthophoto', 'odm_orthophoto_preview.png'),
                    ]:
                        if os.path.exists(p):
                            self.image_path = p
                            self.current_image = Image.open(p)
                            self.zoom_factor = 1.0
                            self.display_image()
                            return
            except Exception:
                pass
            messagebox.showerror("エラー", f"画像の読み込みに失敗しました: {str(e)}")

    def _read_tiff_with_tifffile(self, file_path):
        """tifffileでTIFFを読む（float/BIGTIFF/各種圧縮に強い）→ PIL.Imageに変換"""
        try:
            import tifffile as tiff
            arr = tiff.imread(file_path)
            # 正規化して8bitへ（NaN/Inf対応）
            if arr.dtype != np.uint8:
                arrf = arr.astype(np.float32)
                finite = np.isfinite(arrf)
                if not finite.any():
                    arr8 = np.zeros(arr.shape[:2], dtype=np.uint8)
                else:
                    vmin = float(arrf[finite].min())
                    vmax = float(arrf[finite].max())
                    if vmax - vmin < 1e-12:
                        arr8 = np.zeros(arr.shape[:2], dtype=np.uint8)
                    else:
                        arr8 = np.clip((arrf - vmin) / (vmax - vmin) * 255.0, 0, 255).astype(np.uint8)
            else:
                arr8 = arr
            if arr8.ndim == 2:
                return Image.fromarray(arr8).convert('RGB')
            elif arr8.ndim == 3:
                c = arr8.shape[2]
                if c == 3:
                    return Image.fromarray(arr8)
                elif c >= 4:
                    return Image.fromarray(arr8[:, :, :3])
                else:
                    # 1ch/2chなどは先頭chをRGB化
                    return Image.fromarray(arr8[:, :, 0]).convert('RGB')
        except Exception:
            return None

    def _read_tiff_with_opencv_unicode(self, file_path):
        """np.fromfile + cv2.imdecodeで日本語パス/float系も正規化してPIL化"""
        try:
            data = np.fromfile(file_path, dtype=np.uint8)
            if data is None or data.size == 0:
                return None
            img = cv2.imdecode(data, cv2.IMREAD_UNCHANGED)
            if img is None:
                img = cv2.imread(file_path, cv2.IMREAD_UNCHANGED)
            if img is None:
                return None
            if img.dtype != np.uint8:
                img = cv2.normalize(img, None, 0, 255, cv2.NORM_MINMAX)
                img = img.astype(np.uint8)
            if len(img.shape) == 2:
                return Image.fromarray(img).convert('RGB')
            elif len(img.shape) == 3:
                c = img.shape[2]
                if c == 3:
                    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    return Image.fromarray(img)
                elif c == 4:
                    img = cv2.cvtColor(img, cv2.COLOR_BGRA2RGBA)
                    return Image.fromarray(img)
                elif c > 4:
                    img = img[:, :, :3]
                    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    return Image.fromarray(img)
            return None
        except Exception:
            return None

    def display_image(self):
        """画像をキャンバスに表示"""
        if self.current_image:
            # ズーム適用
            display_size = (
                int(self.current_image.width * self.zoom_factor),
                int(self.current_image.height * self.zoom_factor)
            )
            resized_image = self.current_image.resize(display_size, Image.Resampling.LANCZOS)
            self.canvas_image = ImageTk.PhotoImage(resized_image)

            # キャンバスクリア
            self.canvas.delete("all")

            # 画像表示
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.canvas_image)

            # スクロール領域設定
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

            # アノテーション描画
            self.draw_annotations()
            
            # 移動モードボタンを再配置（最前面に表示）
            self.create_move_mode_button()

    def create_move_mode_button(self):
        """移動モードボタンを作成・更新"""
        if not hasattr(self, 'move_mode_button'):
            # 初回作成
            self.move_mode_button = tk.Button(
                self.canvas,
                text="↗",
                command=self.toggle_move_mode,
                width=3,
                height=1,
                relief=tk.RAISED,
                font=("Arial", 14, "bold"),
                bg="SystemButtonFace"
            )
            self.move_mode_button_window = self.canvas.create_window(
                10, 10,
                anchor=tk.NW,
                window=self.move_mode_button,
                tags="move_button"
            )
        else:
            # ボタンを最前面に
            self.canvas.tag_raise("move_button")

    def toggle_move_mode(self):
        """移動モードのON/OFFを切り替え"""
        self.move_mode = not self.move_mode
        
        if self.move_mode:
            # 移動モードON
            self.move_mode_button.config(relief=tk.SUNKEN, bg="#90EE90")
            self.canvas.config(cursor="hand2")
            print("[移動モード] ON - アノテーションをクリックして移動できます")
        else:
            # 移動モードOFF
            self.move_mode_button.config(relief=tk.RAISED, bg="SystemButtonFace")
            self.canvas.config(cursor="")
            print("[移動モード] OFF")
            # 選択中のアノテーションをクリア
            if self.moving_annotation:
                self.moving_annotation = None
                self.move_original_pos = None
                self.draw_annotations()

    def find_nearest_annotation(self, image_x, image_y, threshold=30):
        """指定座標から最も近いアノテーションを検索
        
        Args:
            image_x: 画像座標X
            image_y: 画像座標Y
            threshold: 検索範囲の閾値（ピクセル）
        
        Returns:
            最も近いアノテーション、または None
        """
        min_distance = float('inf')
        nearest_annotation = None
        
        for annotation in self.annotations:
            ann_x = annotation["x"]
            ann_y = annotation["y"]
            
            # ユークリッド距離を計算
            distance = math.sqrt((image_x - ann_x)**2 + (image_y - ann_y)**2)
            
            # 閾値内かつ最小距離を更新
            if distance < threshold / self.zoom_factor and distance < min_distance:
                min_distance = distance
                nearest_annotation = annotation
        
        return nearest_annotation

    def initialize_annotation_icons(self, reset_warning=True):
        """アノテーションアイコンを初期化し、フォルダから読み込む"""
        if reset_warning:
            self._icon_warning_shown = False
        self.missing_icon_types.clear()
        self.annotation_icon_cache.clear()
        self.annotation_icon_tk_cache.clear()

        icon_dir = self.annotation_icon_dir
        if not os.path.isdir(icon_dir):
            if not self._icon_dir_missing_warned:
                messagebox.showwarning(
                    "アノテーションアイコン",
                    "プログラムと同じフォルダにある『アノテーション画像フォルダ』からアイコンを読み込めませんでした。\n"
                    "フォルダ配置とアクセス権限を確認してください。\n\n"
                    "PNG形式のアイコンファイルが必要です。"
                )
                self._icon_dir_missing_warned = True
            return

        self._icon_dir_missing_warned = False

        for defect_type in self.defect_types.keys():
            self.load_annotation_icon(defect_type)

    def get_annotation_icon_path(self, defect_type):
        sanitized = str(defect_type or "").strip()
        if not sanitized:
            return None
        sanitized = sanitized.replace("/", "_").replace("\\", "_")
        # PNG優先、次にJPG/JPEGをフォールバック
        png_path = os.path.join(self.annotation_icon_dir, f"{sanitized}.png")
        if os.path.exists(png_path):
            return png_path
        jpg_path = os.path.join(self.annotation_icon_dir, f"{sanitized}.jpg")
        if os.path.exists(jpg_path):
            return jpg_path
        jpeg_path = os.path.join(self.annotation_icon_dir, f"{sanitized}.jpeg")
        if os.path.exists(jpeg_path):
            return jpeg_path
        # 見つからない場合はPNGパスを返す（エラーメッセージ用）
        return png_path

    def load_annotation_icon(self, defect_type):
        if defect_type in self.annotation_icon_cache:
            return self.annotation_icon_cache[defect_type]

        path = self.get_annotation_icon_path(defect_type)
        if not path or not os.path.exists(path):
            self.annotation_icon_cache[defect_type] = None
            self._register_missing_icon(defect_type, f"ファイルが見つかりません: {path}")
            return None

        try:
            # Pillowで直接読み込み（PNG/JPG対応）
            image = Image.open(path).convert("RGBA")
            
            # 標準サイズ（256x256）にリサイズ（必要な場合）
            if image.size != (256, 256):
                image = image.resize((256, 256), Image.Resampling.LANCZOS)
            
            self.annotation_icon_cache[defect_type] = image
            return image
        except Exception as e:
            self.annotation_icon_cache[defect_type] = None
            self._register_missing_icon(defect_type, f"画像読み込みに失敗しました: {e}")
            return None

    def _register_missing_icon(self, defect_type, message):
        if defect_type in self.missing_icon_types:
            return
        self.missing_icon_types.add(defect_type)
        print(f"[WARN] annotation icon missing for '{defect_type}': {message}")
        if not self._icon_warning_shown:
            messagebox.showwarning(
                "アノテーションアイコン",
                "アノテーション用アイコンを読み込めませんでした。\n"
                "フォルダ『アノテーション画像フォルダ』とPNG/JPG形式のアイコンファイルを確認してください。\n\n"
                "ヒント: convert_svg_to_png.py スクリプトでSVGをPNGに変換できます。"
            )
            self._icon_warning_shown = True

    def get_tk_icon(self, defect_type, size, alpha=1.0):
        """透明度を指定してTkinterアイコンを取得
        
        Args:
            defect_type: 不良分類
            size: アイコンサイズ
            alpha: 透明度（0.0～1.0）
        """
        size = max(16, int(size))
        key = (defect_type, size, alpha)
        if key in self.annotation_icon_tk_cache:
            return self.annotation_icon_tk_cache[key]

        base_icon = self.load_annotation_icon(defect_type)
        if base_icon is None:
            return None

        if base_icon.size != (size, size):
            icon_image = base_icon.resize((size, size), Image.Resampling.LANCZOS)
        else:
            icon_image = base_icon.copy()

        # 透明度適用
        if alpha < 1.0:
            # RGBAモードに変換して透明度を適用
            if icon_image.mode != 'RGBA':
                icon_image = icon_image.convert('RGBA')
            # アルファチャンネルを調整
            data = icon_image.getdata()
            new_data = []
            for item in data:
                # RGBAの各ピクセル
                new_data.append((item[0], item[1], item[2], int(item[3] * alpha)))
            icon_image.putdata(new_data)

        tk_icon = ImageTk.PhotoImage(icon_image)
        self.annotation_icon_tk_cache[key] = tk_icon
        return tk_icon

    def _get_annotation_scale(self, key):
        var = self.annotation_scale_vars.get(key)
        if var is None:
            var = self.annotation_scale_vars.get("overall")
        if var is None:
            return 1.0
        value = var.get()
        try:
            scale = float(value)
            if not math.isfinite(scale) or scale <= 0:
                raise ValueError
            return scale
        except (TypeError, ValueError):
            scale = 1.0
            try:
                var.set("1.0")
            except Exception:
                pass
            return scale

    def _draw_id_label_on_image(self, draw, x, y, annotation_id, color, image_size, scale_multiplier=1.0, icon_height=None, image_type=None):
        """
        画像上にID番号ラベルを描画
        
        Args:
            draw: ImageDrawオブジェクト
            x, y: 描画座標
            annotation_id: アノテーションID
            color: 色
            image_size: 画像サイズ (width, height)
            scale_multiplier: スケール倍率
            icon_height: アイコンの高さ
            image_type: 画像タイプ ('thermal'=サーモ画像, 'visible'=可視画像, 'ortho'=オルソ画像全体図, None=オフセットなし)
        """
        # オフセットの適用
        offset_x = 0
        offset_y = 0
        if image_type == 'thermal':
            offset_x = self.thermal_offset_x
            offset_y = self.thermal_offset_y
        elif image_type == 'visible':
            offset_x = self.visible_offset_x
            offset_y = self.visible_offset_y
        elif image_type == 'ortho':
            offset_x = self.ortho_offset_x
            offset_y = self.ortho_offset_y
        
        # オフセット適用後の座標
        adjusted_x = x + offset_x
        adjusted_y = y + offset_y
        
        img_w, img_h = image_size
        base_font_size = max(12, min(24, int(min(img_w, img_h) / 50))) if img_w and img_h else 12
        font_size = max(8, int(round(base_font_size * scale_multiplier)))
        font_size = min(font_size, 128)
        try:
            font = ImageFont.truetype("DejaVuSans-Bold.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.truetype("Arial.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()
        stroke_width = max(1, int(round(2 * scale_multiplier)))
        offset = 25 * scale_multiplier
        if icon_height:
            offset = max(offset, (icon_height / 2) + 6 * scale_multiplier)
        text_x = adjusted_x + offset
        text_y = adjusted_y - offset
        if img_w:
            text_x = max(0, min(img_w - 1, text_x))
        if img_h:
            text_y = max(0, min(img_h - 1, text_y))
        draw.text(
            (text_x, text_y),
            f"ID{annotation_id}",
            fill=color,
            stroke_width=stroke_width,
            stroke_fill="white",
            font=font,
        )

    def draw_annotations(self):
        """アノテーションを描画"""
        self.canvas_icon_refs.clear()
        for annotation in self.annotations:
            self.canvas.delete(f"annotation_{annotation['id']}")
            x = annotation["x"] * self.zoom_factor
            y = annotation["y"] * self.zoom_factor
            defect_type = annotation.get("defect_type")
            color = self.defect_types.get(defect_type, "#FF0000")

            # 移動中のアノテーションは半透明化
            is_moving = (self.moving_annotation and 
                        annotation['id'] == self.moving_annotation['id'])
            alpha = 0.5 if is_moving else 1.0

            icon_size = max(24, int(self.annotation_default_icon_size * self.zoom_factor))
            tk_icon = self.get_tk_icon(defect_type, icon_size, alpha=alpha)

            if tk_icon is None:
                icon_path = self.get_annotation_icon_path(defect_type)
                if icon_path and os.path.isfile(icon_path):
                    self.initialize_annotation_icons(reset_warning=False)
                    tk_icon = self.get_tk_icon(defect_type, icon_size, alpha=alpha)

            if tk_icon:
                self.canvas.create_image(
                    x, y,
                    image=tk_icon,
                    anchor=tk.CENTER,
                    tags=f"annotation_{annotation['id']}"
                )
                self.canvas_icon_refs.append(tk_icon)

                label_offset = (tk_icon.height() / 2) + 12 * self.zoom_factor
                # 移動中はIDテキストも半透明化（stippleで表現）
                if is_moving:
                    self.draw_id_text(x, y - label_offset, color, annotation['id'], stipple='gray50')
                else:
                    self.draw_id_text(x, y - label_offset, color, annotation['id'])
                continue

            shape = annotation.get("shape", "cross")

            # 形状に応じて描画（移動中は半透明効果を追加）
            if shape == "cross":
                self.draw_cross(x, y, color, annotation['id'], stipple='gray50' if is_moving else None)
            elif shape == "arrow":
                self.draw_arrow(x, y, color, annotation['id'], stipple='gray50' if is_moving else None)
            elif shape == "circle":
                self.draw_circle(x, y, color, annotation['id'], stipple='gray50' if is_moving else None)
            elif shape == "rectangle":
                self.draw_rectangle(x, y, color, annotation['id'], stipple='gray50' if is_moving else None)

    def draw_annotation_icon_on_image(self, image, draw, x, y, defect_type, color, fallback_shape, scale_multiplier=1.0, image_type=None):
        """
        画像上にアノテーションアイコンを描画
        
        Args:
            image: 描画対象の画像
            draw: ImageDrawオブジェクト
            x, y: 描画座標
            defect_type: 不具合タイプ
            color: 色
            fallback_shape: フォールバック形状
            scale_multiplier: スケール倍率
            image_type: 画像タイプ ('thermal'=サーモ画像, 'visible'=可視画像, 'ortho'=オルソ画像全体図, None=オフセットなし)
        
        Returns:
            アイコンの高さ
        """
        try:
            scale_multiplier = float(scale_multiplier)
        except (TypeError, ValueError):
            scale_multiplier = 1.0
        if not math.isfinite(scale_multiplier) or scale_multiplier <= 0:
            scale_multiplier = 1.0

        # オフセットの適用
        offset_x = 0
        offset_y = 0
        if image_type == 'thermal':
            offset_x = self.thermal_offset_x
            offset_y = self.thermal_offset_y
        elif image_type == 'visible':
            offset_x = self.visible_offset_x
            offset_y = self.visible_offset_y
        elif image_type == 'ortho':
            offset_x = self.ortho_offset_x
            offset_y = self.ortho_offset_y
        
        # オフセット適用後の座標
        adjusted_x = x + offset_x
        adjusted_y = y + offset_y

        base_icon = self.load_annotation_icon(defect_type)
        if base_icon is not None:
            icon_w, icon_h = base_icon.size
            base_length = max(image.width, image.height)
            min_dim = max(1, min(image.width, image.height))
            base_target = min(max(24, int(base_length * 0.05)), max(16, min_dim), 128)
            target_edge = max(4, int(round(base_target * scale_multiplier)))
            target_edge = min(target_edge, min_dim)
            scale = target_edge / max(icon_w, icon_h) if max(icon_w, icon_h) else 1.0
            resized_size = (
                max(1, int(round(icon_w * scale))),
                max(1, int(round(icon_h * scale)))
            )
            icon_image = base_icon.resize(resized_size, Image.Resampling.LANCZOS) if resized_size != base_icon.size else base_icon

            paste_x = int(round(adjusted_x - icon_image.width / 2))
            paste_y = int(round(adjusted_y - icon_image.height / 2))
            paste_x = max(0, min(image.width - icon_image.width, paste_x))
            paste_y = max(0, min(image.height - icon_image.height, paste_y))

            image.paste(icon_image, (paste_x, paste_y), icon_image)
            return icon_image.height

        if fallback_shape == "cross":
            return self.draw_cross_on_image(draw, adjusted_x, adjusted_y, color, scale_multiplier)
        elif fallback_shape == "arrow":
            return self.draw_arrow_on_image(draw, adjusted_x, adjusted_y, color, scale_multiplier)
        elif fallback_shape == "circle":
            return self.draw_circle_on_image(draw, adjusted_x, adjusted_y, color, scale_multiplier)
        elif fallback_shape == "rectangle":
            return self.draw_rectangle_on_image(draw, adjusted_x, adjusted_y, color, scale_multiplier)
        else:
            return self.draw_cross_on_image(draw, adjusted_x, adjusted_y, color, scale_multiplier)

    def draw_cross(self, x, y, color, annotation_id, stipple=None):
        """十字形状を描画"""
        size = 20 * self.zoom_factor
        line_kwargs = {"fill": color, "width": 3, "tags": f"annotation_{annotation_id}"}
        if stipple:
            line_kwargs["stipple"] = stipple
        self.canvas.create_line(x, y - size, x, y + size, **line_kwargs)
        self.canvas.create_line(x - size, y, x + size, y, **line_kwargs)
        self.draw_id_text(x, y - size - 15 * self.zoom_factor, color, annotation_id, stipple)

    def draw_arrow(self, x, y, color, annotation_id, stipple=None):
        """矢印形状を描画"""
        size = 20 * self.zoom_factor
        # 矢印の軸
        line_kwargs = {"fill": color, "width": 3, "tags": f"annotation_{annotation_id}"}
        if stipple:
            line_kwargs["stipple"] = stipple
        self.canvas.create_line(x, y - size, x, y + size, **line_kwargs)
        # 矢印の先端
        poly_kwargs = {"fill": color, "tags": f"annotation_{annotation_id}"}
        if stipple:
            poly_kwargs["stipple"] = stipple
        self.canvas.create_polygon(
            x, y - size,
            x - 8 * self.zoom_factor, y - size + 15 * self.zoom_factor,
            x + 8 * self.zoom_factor, y - size + 15 * self.zoom_factor,
            **poly_kwargs
        )
        self.draw_id_text(x, y - size - 15 * self.zoom_factor, color, annotation_id, stipple)

    def draw_circle(self, x, y, color, annotation_id, stipple=None):
        """円形状を描画"""
        radius = 25 * self.zoom_factor
        oval_kwargs = {"outline": color, "width": 3, "tags": f"annotation_{annotation_id}"}
        if stipple:
            oval_kwargs["stipple"] = stipple
        self.canvas.create_oval(x - radius, y - radius, x + radius, y + radius, **oval_kwargs)
        self.draw_id_text(x, y - radius - 15 * self.zoom_factor, color, annotation_id, stipple)

    def draw_rectangle(self, x, y, color, annotation_id, stipple=None):
        """四角形状を描画"""
        size = 20 * self.zoom_factor
        rect_kwargs = {"outline": color, "width": 3, "tags": f"annotation_{annotation_id}"}
        if stipple:
            rect_kwargs["stipple"] = stipple
        self.canvas.create_rectangle(x - size, y - size, x + size, y + size, **rect_kwargs)
        self.draw_id_text(x, y - size - 15 * self.zoom_factor, color, annotation_id, stipple)

    def draw_id_text(self, x, y, color, annotation_id, stipple=None):
        """ID番号を描画"""
        text_size = max(10, int(12 * self.zoom_factor))
        
        # 背景付きテキスト表示
        text_kwargs = {
            "text": f"ID{annotation_id}",
            "fill": "white",
            "font": ("Arial", text_size, "bold"),
            "anchor": "center",
            "tags": f"annotation_{annotation_id}"
        }
        if stipple:
            text_kwargs["stipple"] = stipple
        
        text_id = self.canvas.create_text(x, y, **text_kwargs)

        # テキストの背景
        bbox = self.canvas.bbox(text_id)
        if bbox:
            rect_kwargs = {
                "fill": color,
                "outline": color,
                "tags": f"annotation_{annotation_id}"
            }
            if stipple:
                rect_kwargs["stipple"] = stipple
            
            self.canvas.create_rectangle(
                bbox[0] - 2, bbox[1] - 1,
                bbox[2] + 2, bbox[3] + 1,
                **rect_kwargs
            )
            # テキストを前面に移動
            self.canvas.tag_raise(text_id)

    def _confirm_id_change(self, old_id, new_id, parent=None):
        if old_id == new_id:
            return True
        try:
            old_id_str = f"ID{int(old_id)}"
        except (TypeError, ValueError):
            old_id_str = f"ID{old_id}"
        confirm_message = (
            f"{old_id_str} を ID{new_id} に変更します。\n"
            "重複しているIDがある場合は、後続のID番号を自動で繰り上げます。\n"
            "よろしいですか？"
        )
        return messagebox.askokcancel("ID番号の変更", confirm_message, parent=parent)

    def apply_annotation_id_change(self, annotation, new_id, parent=None, notify_conflicts=True):
        if annotation not in self.annotations:
            return
        try:
            new_id_int = int(new_id)
        except (TypeError, ValueError):
            raise ValueError("ID番号は整数で指定してください。")

        if new_id_int <= 0:
            raise ValueError("ID番号は 1 以上である必要があります。")

        previous_ids = {id(ann): ann.get('id') for ann in self.annotations}
        annotation['id'] = new_id_int
        self._resolve_id_conflicts(annotation)

        changed_annotations = []
        for ann in self.annotations:
            previous_id = previous_ids.get(id(ann))
            if previous_id is not None and previous_id != ann.get('id'):
                changed_annotations.append((ann, previous_id, ann.get('id')))

        if changed_annotations:
            for target_annotation, old_id, updated_id in changed_annotations:
                self._update_annotation_files_for_id_change(target_annotation, old_id, updated_id)

        order_map = {id(ann): idx for idx, ann in enumerate(self.annotations)}
        self.annotations.sort(key=lambda ann: (ann.get('id', 0), order_map.get(id(ann), 0)))

        if self.annotations:
            self.next_id = max(ann.get('id', 0) for ann in self.annotations) + 1
        else:
            self.next_id = 1

        self.update_table()
        self.draw_annotations()
        self._select_tree_item_by_id(annotation.get('id'))

        if notify_conflicts and len(changed_annotations) > 1:
            info_lines = [
                f"ID{old_id} → ID{updated_id}"
                for ann_obj, old_id, updated_id in changed_annotations
                if ann_obj is not annotation
            ]
            if info_lines:
                messagebox.showinfo(
                    "ID番号の更新",
                    "重複していたIDを次番号へ繰り上げました:\n" + "\n".join(info_lines),
                    parent=parent
                )

    def _resolve_id_conflicts(self, primary_annotation):
        order_map = {id(ann): idx for idx, ann in enumerate(self.annotations)}
        sorted_candidates = sorted(
            self.annotations,
            key=lambda ann: (
                ann.get('id', 0),
                0 if ann is primary_annotation else 1,
                order_map.get(id(ann), 0)
            )
        )

        seen_ids = set()
        for ann in sorted_candidates:
            current_id = int(ann.get('id', 0) or 0)
            if current_id <= 0:
                current_id = 1
            while current_id in seen_ids:
                current_id += 1
            if current_id != ann.get('id'):
                ann['id'] = current_id
            seen_ids.add(current_id)

    def _is_path_inside_project(self, path):
        if not path or not self.project_path:
            return False
        try:
            project_root = os.path.abspath(self.project_path)
            target_path = os.path.abspath(path)
            return os.path.commonpath([project_root, target_path]) == project_root
        except Exception:
            return False

    def _rename_file_with_new_id(self, path, old_id, new_id):
        if not path:
            return path
        try:
            resolved_path = os.path.abspath(path)
        except Exception:
            resolved_path = path
        if not os.path.exists(resolved_path):
            return path
        if not self._is_path_inside_project(resolved_path):
            return path

        directory, filename = os.path.split(resolved_path)
        match = re.search(r"ID(\d+)", filename)
        if not match:
            return path

        original_digits = match.group(1)
        new_digits = str(new_id).zfill(len(original_digits))
        new_filename = f"{filename[:match.start(1)]}{new_digits}{filename[match.end(1):]}"
        new_path = os.path.join(directory, new_filename)

        if new_path == resolved_path:
            return path

        try:
            os.makedirs(directory, exist_ok=True)
            os.replace(resolved_path, new_path)
            return new_path
        except Exception as e:
            print(f"[WARN] ファイル名を変更できませんでした: {resolved_path} -> {new_path} ({e})")
            return path

    def _update_annotation_files_for_id_change(self, annotation, old_id, new_id):
        if old_id == new_id:
            return
        for key in [
            "thermal_image",
            "visible_image",
            "thermal_annotated_image",
            "visible_annotated_image"
        ]:
            path = annotation.get(key)
            new_path = self._rename_file_with_new_id(path, old_id, new_id)
            if new_path and new_path != path:
                annotation[key] = new_path

    def find_annotation_by_id(self, annotation_id):
        for annotation in self.annotations:
            try:
                if int(annotation.get('id')) == int(annotation_id):
                    return annotation
            except (TypeError, ValueError):
                continue
        return None

    def _select_tree_item_by_id(self, annotation_id):
        if annotation_id is None:
            return
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            if values and str(values[0]) == str(annotation_id):
                self.tree.selection_set(item)
                self.tree.focus(item)
                self.tree.see(item)
                break

    def prompt_id_change_from_table(self, annotation):
        if not annotation:
            return
        old_id = annotation.get('id')
        new_id = simpledialog.askinteger(
            "ID番号の編集",
            f"ID{old_id} を新しい番号に変更します。\n1以上の整数を入力してください。",
            initialvalue=old_id,
            minvalue=1,
            parent=self.root
        )
        if new_id is None or new_id == old_id:
            return
        if not self._confirm_id_change(old_id, new_id, parent=self.root):
            return
        try:
            self.apply_annotation_id_change(annotation, new_id, parent=self.root)
        except ValueError as e:
            messagebox.showwarning("警告", str(e), parent=self.root)

    def on_tree_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        item = self.tree.item(row_id)
        values = item.get("values")
        if not values:
            return
        try:
            annotation_id = int(values[0])
        except (TypeError, ValueError):
            return

        annotation = self.find_annotation_by_id(annotation_id)
        if not annotation:
            return

        column_id = self.tree.identify_column(event.x)
        if column_id == "#1":
            self.prompt_id_change_from_table(annotation)
        else:
            self.edit_annotation_dialog(annotation)

    def on_canvas_click(self, event):
        """キャンバスクリック時の処理"""
        if self.current_image:
            # キャンバス座標を画像座標に変換
            canvas_x = self.canvas.canvasx(event.x)
            canvas_y = self.canvas.canvasy(event.y)

            image_x = canvas_x / self.zoom_factor
            image_y = canvas_y / self.zoom_factor

            # 画像範囲内かチェック
            if 0 <= image_x <= self.current_image.width and 0 <= image_y <= self.current_image.height:
                
                # 移動モード中の処理
                if self.move_mode:
                    self.handle_move_mode_click(image_x, image_y)
                    return
                
                # 通常モード: アノテーション追加
                self.add_annotation(image_x, image_y)

    def add_annotation(self, x, y):
        """アノテーションを追加"""
        defect_type = self.defect_var.get()
        shape = self.annotation_shapes[self.shape_var.get()]

        annotation = {
            "id": self.next_id,
            "x": x,
            "y": y,
            "defect_type": defect_type,
            "shape": shape,
            "area_no": "",
            "pcs_no": "",
            "junction_box_no": "",
            "circuit_no": "",
            "array_no": "",           # 追加
            "module_position": "",    # 追加
            "serial_no": "",          # 追加
            "thermal_image": "",
            "visible_image": "",
            "remarks": "",
            "found_by": None,
            "management_level": 'S',
            "report_date": datetime.now().strftime('%Y-%m-%d')  # 追加（現在日付を自動設定）
        }

        self.annotations.append(annotation)
        self.next_id += 1
        
        # 変更追跡
        self.has_unsaved_changes = True
        self.changed_annotation_ids.add(annotation['id'])

        self.update_table()
        self.draw_annotations()

    def handle_move_mode_click(self, image_x, image_y):
        """移動モード中のクリック処理"""
        
        if self.moving_annotation is None:
            # 1回目のクリック: アノテーションを選択
            annotation = self.find_nearest_annotation(image_x, image_y)
            
            if annotation:
                self.moving_annotation = annotation
                self.move_original_pos = (annotation["x"], annotation["y"])
                # 半透明化して再描画
                self.draw_annotations()
                print(f"[移動モード] アノテーションID {annotation['id']} を選択しました")
            else:
                messagebox.showinfo("情報", "アノテーションが見つかりません。", parent=self.root)
        
        else:
            # 2回目のクリック: 選択中のアノテーションを移動
            old_x, old_y = self.move_original_pos
            self.moving_annotation["x"] = image_x
            self.moving_annotation["y"] = image_y
            
            # 画像範囲内にクランプ
            self.moving_annotation["x"] = max(0, min(self.current_image.width, self.moving_annotation["x"]))
            self.moving_annotation["y"] = max(0, min(self.current_image.height, self.moving_annotation["y"]))
            
            # ログ出力
            print(f"[移動モード] アノテーションID {self.moving_annotation['id']} を移動: "
                  f"({old_x:.1f}, {old_y:.1f}) → ({self.moving_annotation['x']:.1f}, {self.moving_annotation['y']:.1f})")
            
            # テーブル更新
            self.update_table()
            
            # 変更追跡
            self.has_unsaved_changes = True
            self.changed_annotation_ids.add(self.moving_annotation['id'])
            
            # 選択状態をクリア
            self.moving_annotation = None
            self.move_original_pos = None
            
            # 再描画（通常状態に戻す）
            self.draw_annotations()

    def on_ctrl_click(self, event):
        """Ctrl + クリック時の処理（ドラッグ開始）"""
        if not self.current_image:
            return
        
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)
        
        image_x = canvas_x / self.zoom_factor
        image_y = canvas_y / self.zoom_factor
        
        # 最も近いアノテーションを検索
        annotation = self.find_nearest_annotation(image_x, image_y)
        
        if annotation:
            self.moving_annotation = annotation
            self.move_start_pos = (image_x, image_y)
            self.move_original_pos = (annotation["x"], annotation["y"])
            self.canvas.config(cursor="fleur")  # 移動カーソル
            print(f"[Ctrl+ドラッグ] アノテーションID {annotation['id']} を選択")

    def on_ctrl_drag(self, event):
        """Ctrl + ドラッグ中の処理（リアルタイム更新）"""
        if not self.moving_annotation or not self.move_start_pos:
            return
        
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)
        
        image_x = canvas_x / self.zoom_factor
        image_y = canvas_y / self.zoom_factor
        
        # 移動量を計算
        dx = image_x - self.move_start_pos[0]
        dy = image_y - self.move_start_pos[1]
        
        # アノテーション位置を更新
        self.moving_annotation["x"] = self.move_original_pos[0] + dx
        self.moving_annotation["y"] = self.move_original_pos[1] + dy
        
        # 画像範囲内にクランプ
        self.moving_annotation["x"] = max(0, min(self.current_image.width, self.moving_annotation["x"]))
        self.moving_annotation["y"] = max(0, min(self.current_image.height, self.moving_annotation["y"]))
        
        # 再描画
        self.draw_annotations()

    def on_ctrl_release(self, event):
        """Ctrl + ドラッグ終了時の処理（確定）"""
        if self.moving_annotation:
            # ログ出力
            print(f"[Ctrl+ドラッグ] アノテーションID {self.moving_annotation['id']} を移動完了: "
                  f"({self.move_original_pos[0]:.1f}, {self.move_original_pos[1]:.1f}) → "
                  f"({self.moving_annotation['x']:.1f}, {self.moving_annotation['y']:.1f})")
            
            # テーブルを更新
            self.update_table()
            
            # 変更追跡
            self.has_unsaved_changes = True
            self.changed_annotation_ids.add(self.moving_annotation['id'])
            
            # 状態をリセット
            self.moving_annotation = None
            self.move_start_pos = None
            self.move_original_pos = None
            self.canvas.config(cursor="hand2" if self.move_mode else "")

    def cancel_move(self, event=None):
        """移動操作をキャンセル"""
        if self.moving_annotation and self.move_original_pos:
            # 元の位置に戻す
            self.moving_annotation["x"] = self.move_original_pos[0]
            self.moving_annotation["y"] = self.move_original_pos[1]
            
            print(f"[キャンセル] アノテーションID {self.moving_annotation['id']} の移動をキャンセル")
            
            # 状態をリセット
            self.moving_annotation = None
            self.move_original_pos = None
            self.move_start_pos = None
            
            # カーソルを元に戻す
            self.canvas.config(cursor="hand2" if self.move_mode else "")
            
            # 再描画
            self.draw_annotations()

    def on_canvas_double_click(self, event):
        """ダブルクリックでアノテーション編集"""
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)

        # 最も近いアノテーションを検索
        min_distance = float('inf')
        selected_annotation = None

        for annotation in self.annotations:
            ann_x = annotation["x"] * self.zoom_factor
            ann_y = annotation["y"] * self.zoom_factor
            distance = ((canvas_x - ann_x) ** 2 + (canvas_y - ann_y) ** 2) ** 0.5

            if distance < 30 and distance < min_distance:
                min_distance = distance
                selected_annotation = annotation

        if selected_annotation:
            self.edit_annotation_dialog(selected_annotation)

    def on_canvas_right_click(self, event):
        """右クリックでアノテーション削除"""
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)

        # 最も近いアノテーションを検索
        min_distance = float('inf')
        selected_annotation = None

        for annotation in self.annotations:
            ann_x = annotation["x"] * self.zoom_factor
            ann_y = annotation["y"] * self.zoom_factor
            distance = ((canvas_x - ann_x) ** 2 + (canvas_y - ann_y) ** 2) ** 0.5

            if distance < 30 and distance < min_distance:
                min_distance = distance
                selected_annotation = annotation

        if selected_annotation:
            if messagebox.askyesno("確認", f"ID{selected_annotation['id']}のアノテーションを削除しますか？"):
                self.delete_annotation_by_id(selected_annotation['id'])

    def delete_annotation_by_id(self, annotation_id):
        """指定されたIDのアノテーションを削除し、ID番号を振り直し"""
        # キャンバスからアノテーション画像を削除
        self.canvas.delete(f"annotation_{annotation_id}")
        
        # 変更追跡（削除前に記録）
        self.has_unsaved_changes = True
        self.changed_annotation_ids.add(annotation_id)
        
        # アノテーションを削除
        self.annotations = [ann for ann in self.annotations if ann['id'] != annotation_id]

        # ID番号を振り直し
        self.reassign_ids()

        # 表示を更新
        self.update_table()
        self.draw_annotations()

    def reassign_ids(self):
        """ID番号を1から連番で振り直し"""
        if not self.annotations:
            self.next_id = 1
            return

        order_map = {id(ann): idx for idx, ann in enumerate(self.annotations)}
        self.annotations.sort(key=lambda ann: (ann.get('id', 0), order_map.get(id(ann), 0)))

        changed = []
        for index, annotation in enumerate(self.annotations, 1):
            current_id = annotation.get('id')
            if current_id != index:
                changed.append((annotation, current_id, index))
                annotation['id'] = index

        if changed:
            for ann, old_id, new_id in changed:
                self._update_annotation_files_for_id_change(ann, old_id, new_id)

        self.next_id = len(self.annotations) + 1

    def delete_annotation(self, event):
        """テーブルからアノテーション削除"""
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item[0])
            annotation_id = int(item['values'][0])

            if messagebox.askyesno("確認", f"ID{annotation_id}のアノテーションを削除しますか？"):
                self.delete_annotation_by_id(annotation_id)

    def edit_annotation(self, event):
        """テーブルからアノテーション編集"""
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item[0])
            annotation_id = int(item['values'][0])

            annotation = next((ann for ann in self.annotations if ann['id'] == annotation_id), None)
            if annotation:
                self.edit_annotation_dialog(annotation)

    def edit_annotation_dialog_legacy(self, annotation):
        """アノテーション編集ダイアログ"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"ID{annotation['id']} 編集")
        dialog.geometry("400x750")  # 高さを増加
        dialog.transient(self.root)
        dialog.grab_set()

        # 入力フィールド - 項目を追加
        fields = [
            ("エリア番号", "area_no"),
            ("PCS番号", "pcs_no"),
            ("接続箱番号", "junction_box_no"),
            ("回路番号", "circuit_no"),
            ("アレイ№", "array_no"),           # 追加
            ("モジュール位置", "module_position"), # 追加
            ("シリアル№", "serial_no"),         # 追加
            ("備考", "remarks"),
            ("報告年月日", "report_date")        # 追加
        ]

        entries = {}

        for i, (label, key) in enumerate(fields):
            ttk.Label(dialog, text=f"{label}:").grid(row=i, column=0, sticky="w", padx=10, pady=5)
            
            if key == "report_date":
                # 日付入力用のエントリ（カレンダー形式にしたい場合は別途実装）
                entry = ttk.Entry(dialog, width=30)
                # 既存の日付があればそれを、なければ現在日付を設定
                current_date = annotation.get(key, datetime.now().strftime('%Y-%m-%d'))
                entry.insert(0, current_date)
            else:
                entry = ttk.Entry(dialog, width=30)
                entry.insert(0, annotation.get(key, ""))
            
            entry.grid(row=i, column=1, padx=10, pady=5)
            entries[key] = entry

        # 不良分類選択
        # 管理レベル選択
        ttk.Label(dialog, text="管理レベル:").grid(row=len(fields), column=0, sticky="w", padx=10, pady=5)
        mgmt_var = tk.StringVar(value=(annotation.get("management_level") or 'S'))
        mgmt_combo = ttk.Combobox(dialog, values=['S','A','B','N'], state="readonly", width=27, textvariable=mgmt_var)
        mgmt_combo.grid(row=len(fields), column=1, padx=10, pady=5)

        # 不良分類選択
        ttk.Label(dialog, text="不良分類:").grid(row=len(fields)+1, column=0, sticky="w", padx=10, pady=5)
        defect_combo = ttk.Combobox(dialog, values=list(self.defect_types.keys()), 
                                state="readonly", width=27)
        defect_combo.set(annotation.get("defect_type", ""))
        defect_combo.grid(row=len(fields)+1, column=1, padx=10, pady=5)

        # 形状選択
        ttk.Label(dialog, text="形状:").grid(row=len(fields)+2, column=0, sticky="w", padx=10, pady=5)
        shape_combo = ttk.Combobox(dialog, values=list(self.annotation_shapes.keys()), 
                                state="readonly", width=27)
        current_shape = annotation.get("shape", "cross")
        shape_name = next((k for k, v in self.annotation_shapes.items() if v == current_shape), "十字")
        shape_combo.set(shape_name)
        shape_combo.grid(row=len(fields)+2, column=1, padx=10, pady=5)

        # 画像選択ボタン（ODM機能付き）
        def select_thermal_visible_pair():
            candidates = [
                annotation.get("thermal_image"),
                annotation.get("visible_image"),
                self.last_thermal_visible_dir
            ]
            initial_dir = None
            for candidate in candidates:
                if not candidate:
                    continue
                candidate_dir = candidate if os.path.isdir(candidate) else os.path.dirname(candidate)
                if candidate_dir and os.path.isdir(candidate_dir):
                    initial_dir = candidate_dir
                    break
            selector = ThermalVisibleFileDialog(dialog, initial_dir=initial_dir)
            result = selector.show()
            if result:
                thermal_path, visible_path = result
                annotation["thermal_image"] = thermal_path
                annotation["visible_image"] = visible_path
                self.last_thermal_visible_dir = os.path.dirname(thermal_path)
                thermal_label.config(text=f"サーモ画像: {os.path.basename(thermal_path)}")
                visible_label.config(text=f"可視画像: {os.path.basename(visible_path)}")

        def select_thermal_image():
            file_path = filedialog.askopenfilename(
                title="サーモ画像を選択",
                filetypes=[("画像ファイル", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif")]
            )
            if file_path:
                annotation["thermal_image"] = file_path
                self.last_thermal_visible_dir = os.path.dirname(file_path)
                thermal_label.config(text=f"サーモ画像: {os.path.basename(file_path)}")


        def select_thermal_image_odm():
            # WebODMフォルダが選択されているかチェック
            if not self.webodm_path:
                messagebox.showwarning("警告", "メイン画面で「WebODMフォルダ選択」ボタンからアセットフォルダを先に選択してください。")
                return
            
            def callback(selected_path):
                annotation["thermal_image"] = selected_path
                self.last_thermal_visible_dir = os.path.dirname(selected_path)
                thermal_label.config(text=f"サーモ画像: {os.path.basename(selected_path)}")
            
            # ODMImageSelectorにwebodm_pathを渡す
            ODMImageSelector(dialog, annotation, "thermal", self.webodm_path, callback, app_ref=self)



        def select_visible_image():
            file_path = filedialog.askopenfilename(
                title="可視画像を選択",
                filetypes=[("画像ファイル", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif")]
            )
            if file_path:
                annotation["visible_image"] = file_path
                self.last_thermal_visible_dir = os.path.dirname(file_path)
                visible_label.config(text=f"可視画像: {os.path.basename(file_path)}")

        def select_visible_image_odm():
            # WebODMフォルダが選択されているかチェック
            if not self.webodm_path:
                messagebox.showwarning("警告", "メイン画面で「WebODMフォルダ選択」ボタンからアセットフォルダを先に選択してください。")
                return

            def callback(selected_path):
                annotation["visible_image"] = selected_path
                self.last_thermal_visible_dir = os.path.dirname(selected_path)
                visible_label.config(text=f"可視画像: {os.path.basename(selected_path)}")
            
            # ODMImageSelectorにwebodm_pathを渡す
            ODMImageSelector(dialog, annotation, "visible", self.webodm_path, callback, app_ref=self)


        # サーモ画像選択
        thermal_frame = ttk.Frame(dialog)
        thermal_frame.grid(row=len(fields)+3, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

        ttk.Button(thermal_frame, text="サーモ画像選択", command=select_thermal_image).pack(side=tk.LEFT, padx=(0, 10))
        odm_button_group = ttk.Frame(thermal_frame)
        odm_button_group.pack(side=tk.LEFT)
        ttk.Button(odm_button_group, text="ODM選択", command=select_thermal_image_odm).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(odm_button_group, text="サーモ画像同時選択", command=select_thermal_visible_pair).pack(side=tk.LEFT)

        thermal_label = ttk.Label(dialog, text=f"サーモ画像: {os.path.basename(annotation.get('thermal_image', ''))}")
        thermal_label.grid(row=len(fields)+4, column=0, columnspan=2, padx=10, pady=5, sticky="w")

        # 可視画像選択
        visible_frame = ttk.Frame(dialog)
        visible_frame.grid(row=len(fields)+5, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        
        ttk.Button(visible_frame, text="可視画像選択", command=select_visible_image).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(visible_frame, text="ODM選択", command=select_visible_image_odm).pack(side=tk.LEFT)
        
        visible_label = ttk.Label(dialog, text=f"可視画像: {os.path.basename(annotation.get('visible_image', ''))}")
        visible_label.grid(row=len(fields)+6, column=0, columnspan=2, padx=10, pady=5, sticky="w")

        # ボタン
        def save_changes():
            for key, entry in entries.items():
                annotation[key] = entry.get()
            annotation["defect_type"] = defect_combo.get() or annotation.get("defect_type", "")
            annotation["shape"] = self.annotation_shapes.get(shape_combo.get(), annotation.get("shape", "cross"))
            annotation["management_level"] = (mgmt_var.get() or 'S')

            new_id_str = id_var.get().strip()
            if not new_id_str:
                messagebox.showwarning("警告", "ID番号を入力してください。", parent=dialog)
                return
            if not new_id_str.isdigit():
                messagebox.showwarning("警告", "ID番号は1以上の整数で入力してください。", parent=dialog)
                return

            new_id = int(new_id_str)
            if new_id <= 0:
                messagebox.showwarning("警告", "ID番号は1以上の整数で入力してください。", parent=dialog)
                return

            if new_id != annotation.get("id"):
                if not self._confirm_id_change(annotation.get("id"), new_id, parent=dialog):
                    return
                try:
                    self.apply_annotation_id_change(annotation, new_id, parent=dialog)
                except ValueError as e:
                    messagebox.showwarning("警告", str(e), parent=dialog)
                    return
            else:
                self.update_table()
                self.draw_annotations()

            dialog.destroy()


        def cancel_changes():
            dialog.destroy()

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=len(fields)+7, column=0, columnspan=2, pady=20)

        ttk.Button(button_frame, text="保存", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="キャンセル", command=cancel_changes).pack(side=tk.LEFT, padx=5)

    def edit_annotation_dialog(self, annotation):
        """アノテーション編集ダイアログ"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"ID{annotation['id']} 編集")
        dialog.geometry("980x780")
        dialog.transient(self.root)
        dialog.grab_set()

        annotation.setdefault("thermal_overlays", [])
        annotation.setdefault("visible_overlays", [])
        
        # 編集前のスナップショットを保存（変更追跡用）
        original_annotation = copy.deepcopy(annotation)

        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        main_frame.columnconfigure(0, weight=0)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)

        form_frame = ttk.Frame(main_frame)
        form_frame.grid(row=0, column=0, sticky="ns")
        form_frame.columnconfigure(1, weight=1)

        fields = [
            ("エリア番号", "area_no"),
            ("PCS番号", "pcs_no"),
            ("接続箱番号", "junction_box_no"),
            ("回路番号", "circuit_no"),
            ("アレイ№", "array_no"),
            ("モジュール位置", "module_position"),
            ("シリアル№", "serial_no"),
            ("備考", "remarks"),
            ("報告年月日", "report_date")
        ]

        entries = {}
        current_row = 0

        ttk.Label(form_frame, text="ID番号:").grid(row=current_row, column=0, sticky="w", padx=10, pady=5)
        id_var = tk.StringVar(value=str(annotation.get("id", "")))

        def _validate_id_input(new_value):
            return new_value.isdigit() or new_value == ""

        id_entry = ttk.Entry(
            form_frame,
            width=30,
            textvariable=id_var,
            validate="key",
            validatecommand=(form_frame.register(_validate_id_input), "%P")
        )
        id_entry.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")
        current_row += 1

        for label, key in fields:
            ttk.Label(form_frame, text=f"{label}:").grid(row=current_row, column=0, sticky="w", padx=10, pady=5)
            if key == "report_date":
                entry = ttk.Entry(form_frame, width=30)
                current_date = annotation.get(key, datetime.now().strftime('%Y-%m-%d'))
                entry.insert(0, current_date)
            else:
                entry = ttk.Entry(form_frame, width=30)
                entry.insert(0, annotation.get(key, ""))
            entry.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")
            entries[key] = entry
            current_row += 1

        ttk.Label(form_frame, text="管理レベル:").grid(row=current_row, column=0, sticky="w", padx=10, pady=5)
        mgmt_var = tk.StringVar(value=(annotation.get("management_level") or 'S'))
        mgmt_combo = ttk.Combobox(form_frame, values=['S', 'A', 'B', 'N'], state="readonly", width=27, textvariable=mgmt_var)
        mgmt_combo.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")
        current_row += 1

        ttk.Label(form_frame, text="不良分類:").grid(row=current_row, column=0, sticky="w", padx=10, pady=5)
        defect_combo = ttk.Combobox(form_frame, values=list(self.defect_types.keys()), state="readonly", width=27)
        defect_combo.set(annotation.get("defect_type", ""))
        defect_combo.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")
        current_row += 1

        ttk.Label(form_frame, text="形状:").grid(row=current_row, column=0, sticky="w", padx=10, pady=5)
        shape_combo = ttk.Combobox(form_frame, values=list(self.annotation_shapes.keys()), state="readonly", width=27)
        current_shape = annotation.get("shape", "cross")
        shape_name = next((k for k, v in self.annotation_shapes.items() if v == current_shape), "十字")
        shape_combo.set(shape_name)
        shape_combo.grid(row=current_row, column=1, padx=10, pady=5, sticky="ew")
        current_row += 1

        preview_container = ttk.Frame(main_frame)
        preview_container.grid(row=0, column=1, sticky="nsew", padx=(15, 0))
        preview_container.columnconfigure(0, weight=1)
        preview_container.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(preview_container)
        notebook.grid(row=0, column=0, sticky="nsew")

        outer = self

        class ImageAnnotationPreview:
            def __init__(self, parent, image_type, annotation_dict, defect_combo_ref, shape_combo_ref, initial_path, overlay_points):
                self.outer = outer
                self.annotation = annotation_dict
                self.image_type = image_type
                self.defect_combo = defect_combo_ref
                self.shape_combo = shape_combo_ref
                self.annotation_key = f"{image_type}_overlays"
                self.annotated_path_key = f"{image_type}_annotated_image"
                self.canvas_width = 520
                self.canvas_height = 520
                self.frame = ttk.Frame(parent)

                self.status_var = tk.StringVar(value="画像を選択してください。")
                ttk.Label(self.frame, textvariable=self.status_var, justify="left", wraplength=self.canvas_width).pack(fill=tk.X, pady=(0, 6))

                self.canvas = tk.Canvas(self.frame, width=self.canvas_width, height=self.canvas_height, bg="#1f1f1f", highlightthickness=0)
                self.canvas.pack(fill=tk.BOTH, expand=True)

                ttk.Label(self.frame, text="左クリック: アノテーション追加 / 右クリック: 削除", foreground="#666666").pack(fill=tk.X, pady=(6, 8))

                control_frame = ttk.Frame(self.frame)
                control_frame.pack(fill=tk.X)
                ttk.Button(control_frame, text="決定", command=self.confirm).pack(side=tk.RIGHT, padx=(8, 0))
                ttk.Button(control_frame, text="キャンセル", command=self.cancel).pack(side=tk.RIGHT)

                self.canvas.bind("<Button-1>", self.on_left_click)
                self.canvas.bind("<Button-3>", self.on_right_click)

                self.image_path = None
                self.tk_image = None
                self.display_scale = 1.0
                self.offset_x = 0
                self.offset_y = 0
                self.display_width = 0
                self.display_height = 0
                self.original_points = []
                self.working_points = []
                self.icon_refs = []

                self.update_source(initial_path, overlay_points or [])

            def update_source(self, path, overlay_points=None):
                if overlay_points is None:
                    overlay_points = []
                self.image_path = path if path else None
                self.original_points = copy.deepcopy(overlay_points)
                self.working_points = copy.deepcopy(overlay_points)
                self._render_image()

            def _render_image(self):
                self.canvas.delete("all")
                self.tk_image = None
                self.display_scale = 1.0
                self.offset_x = 0
                self.offset_y = 0
                self.display_width = 0
                self.display_height = 0

                if not self.image_path or not os.path.exists(self.image_path):
                    self.status_var.set("画像未選択です。")
                    self.canvas.create_text(self.canvas_width / 2, self.canvas_height / 2, text="画像が選択されていません", fill="#aaaaaa")
                    return

                try:
                    with Image.open(self.image_path) as img:
                        img = img.convert("RGBA")
                        width, height = img.size
                        scale = min(self.canvas_width / width, self.canvas_height / height, 1.0)
                        display_size = (max(1, int(width * scale)), max(1, int(height * scale)))
                        display_image = img.resize(display_size, Image.Resampling.LANCZOS) if scale != 1.0 else img.copy()
                except Exception as e:
                    self.status_var.set(f"画像を読み込めません: {e}")
                    self.canvas.create_text(self.canvas_width / 2, self.canvas_height / 2, text="読み込み失敗", fill="#ff6666")
                    return

                self.display_width, self.display_height = display_size
                self.display_scale = width and (display_size[0] / width) or 1.0
                self.offset_x = (self.canvas_width - self.display_width) / 2
                self.offset_y = (self.canvas_height - self.display_height) / 2

                self.canvas.create_rectangle(0, 0, self.canvas_width, self.canvas_height, fill="#1f1f1f", outline="")
                self.tk_image = ImageTk.PhotoImage(display_image)
                self.canvas.create_image(self.offset_x, self.offset_y, anchor=tk.NW, image=self.tk_image, tags="base_image")
                self.status_var.set(f"{os.path.basename(self.image_path)} / {width}x{height}px")
                self.redraw_overlays()

            def _current_color(self):
                defect_name = self.defect_combo.get() or self.annotation.get("defect_type")
                return self.outer.defect_types.get(defect_name, "#FF0000")

            def _current_shape_code(self):
                shape_label = self.shape_combo.get()
                if shape_label in self.outer.annotation_shapes:
                    return self.outer.annotation_shapes[shape_label]
                return self.annotation.get("shape", "cross")

            def redraw_overlays(self):
                self.canvas.delete("overlay")
                if not self.tk_image:
                    return

                self.icon_refs = []
                color = self._current_color()
                shape_code = self._current_shape_code()
                defect_name = self.defect_combo.get() or self.annotation.get("defect_type")
                icon_size = max(24, int(self.outer.annotation_default_icon_size * 0.9))
                tk_icon = self.outer.get_tk_icon(defect_name, icon_size)

                for point in self.working_points:
                    disp_x = point["x"] * self.display_scale + self.offset_x
                    disp_y = point["y"] * self.display_scale + self.offset_y

                    if tk_icon:
                        self.canvas.create_image(
                            disp_x,
                            disp_y,
                            image=tk_icon,
                            anchor=tk.CENTER,
                            tags="overlay"
                        )
                        self.icon_refs.append(tk_icon)
                        label_y = disp_y - (tk_icon.height() / 2) - 6
                        self.canvas.create_text(
                            disp_x,
                            label_y,
                            text=f"ID{self.annotation['id']}",
                            fill="white",
                            font=("Arial", 10, "bold"),
                            tags="overlay"
                        )
                    else:
                        self._draw_shape_on_canvas(disp_x, disp_y, color, shape_code)

            def _draw_shape_on_canvas(self, x, y, color, shape_code):
                size = 14
                if shape_code == "cross":
                    self.canvas.create_line(x, y - size, x, y + size, fill=color, width=2, tags="overlay")
                    self.canvas.create_line(x - size, y, x + size, y, fill=color, width=2, tags="overlay")
                elif shape_code == "arrow":
                    self.canvas.create_line(x, y - size, x, y + size, fill=color, width=2, tags="overlay")
                    self.canvas.create_polygon(x, y - size, x - 6, y - size + 12, x + 6, y - size + 12, fill=color, outline=color, tags="overlay")
                elif shape_code == "circle":
                    radius = size
                    self.canvas.create_oval(x - radius, y - radius, x + radius, y + radius, outline=color, width=2, tags="overlay")
                elif shape_code == "rectangle":
                    self.canvas.create_rectangle(x - size, y - size, x + size, y + size, outline=color, width=2, tags="overlay")

                self.canvas.create_text(x, y - size - 6, text=f"ID{self.annotation['id']}", fill="white", font=("Arial", 10, "bold"), tags="overlay")

            def on_left_click(self, event):
                if not self.tk_image:
                    return
                canvas_x = self.canvas.canvasx(event.x)
                canvas_y = self.canvas.canvasy(event.y)
                norm_x = canvas_x - self.offset_x
                norm_y = canvas_y - self.offset_y
                if norm_x < 0 or norm_y < 0 or norm_x > self.display_width or norm_y > self.display_height:
                    return
                original_x = norm_x / self.display_scale
                original_y = norm_y / self.display_scale
                self.working_points.append({"x": original_x, "y": original_y})
                self.redraw_overlays()
                self.status_var.set(f"アノテーションを追加しました ({len(self.working_points)}件)")

            def on_right_click(self, event):
                if not self.tk_image or not self.working_points:
                    return
                canvas_x = self.canvas.canvasx(event.x)
                canvas_y = self.canvas.canvasy(event.y)
                norm_x = canvas_x - self.offset_x
                norm_y = canvas_y - self.offset_y
                if norm_x < 0 or norm_y < 0 or norm_x > self.display_width or norm_y > self.display_height:
                    return

                min_index = None
                min_distance = float('inf')
                for idx, point in enumerate(self.working_points):
                    disp_x = point["x"] * self.display_scale
                    disp_y = point["y"] * self.display_scale
                    distance = math.hypot(norm_x - disp_x, norm_y - disp_y)
                    if distance < 18 and distance < min_distance:
                        min_distance = distance
                        min_index = idx

                if min_index is not None:
                    self.working_points.pop(min_index)
                    self.redraw_overlays()
                    self.status_var.set("アノテーションを削除しました。")

            def confirm(self):
                if not self.image_path or not os.path.exists(self.image_path):
                    messagebox.showwarning("警告", "画像が選択されていません。")
                    return

                try:
                    with Image.open(self.image_path) as src:
                        working_image = src.convert("RGBA")
                except Exception as e:
                    messagebox.showerror("エラー", f"画像の読み込みに失敗しました: {e}")
                    return

                draw = ImageDraw.Draw(working_image)
                color = self._current_color()
                shape_code = self._current_shape_code()
                defect_name = self.defect_combo.get() or self.annotation.get("defect_type")
                scale_key = self.image_type if self.image_type in ("thermal", "visible") else "overall"
                annotation_scale = self.outer._get_annotation_scale(scale_key)

                for point in self.working_points:
                    x = point["x"]
                    y = point["y"]
                    icon_height = self.outer.draw_annotation_icon_on_image(
                        working_image,
                        draw,
                        x,
                        y,
                        defect_name,
                        color,
                        shape_code,
                        annotation_scale
                    )

                    self.outer._draw_id_label_on_image(
                        draw,
                        x,
                        y,
                        self.annotation['id'],
                        color,
                        working_image.size,
                        annotation_scale,
                        icon_height,
                    )

                if getattr(self.outer, "project_path", None):
                    base_dir = os.path.join(self.outer.project_path, "アノテーション入り画像フォルダ")
                else:
                    base_dir = os.path.join(os.path.dirname(self.image_path), "annotated")
                
                # 画像タイプに応じてサブフォルダを作成
                if self.image_type == "thermal":
                    type_dir = os.path.join(base_dir, "サーモ画像フォルダ")
                elif self.image_type == "visible":
                    type_dir = os.path.join(base_dir, "可視画像フォルダ")
                else:
                    type_dir = base_dir
                
                os.makedirs(type_dir, exist_ok=True)

                name, ext = os.path.splitext(os.path.basename(self.image_path))
                if not ext:
                    ext = ".png"
                output_path = os.path.join(type_dir, f"ID{self.annotation['id']}_{name}_{self.image_type}_annotated{ext}")

                try:
                    file_ext = os.path.splitext(output_path)[1].lower()
                    save_image = working_image
                    if file_ext in (".jpg", ".jpeg") and save_image.mode not in ("RGB", "L"):
                        save_image = save_image.convert("RGB")
                    save_image.save(output_path)
                except Exception as e:
                    messagebox.showerror("エラー", f"注釈付き画像の保存に失敗しました: {e}")
                    return

                self.annotation[self.annotated_path_key] = output_path
                self.annotation[self.annotation_key] = copy.deepcopy(self.working_points)
                self.original_points = copy.deepcopy(self.working_points)
                self.status_var.set(f"保存完了: {os.path.basename(output_path)}")
                messagebox.showinfo("保存完了", f"アノテーション入り画像を保存しました。\n{output_path}")

            def cancel(self):
                self.working_points = copy.deepcopy(self.original_points)
                self.redraw_overlays()
                self.status_var.set("編集内容を破棄しました。")
            
            def has_unsaved_changes(self):
                """未保存の変更があるかチェック"""
                if not self.image_path:
                    return False
                # working_pointsとoriginal_pointsを比較
                if len(self.working_points) != len(self.original_points):
                    return True
                # 各ポイントを比較（順序も含めて）
                for wp, op in zip(self.working_points, self.original_points):
                    if wp.get("x") != op.get("x") or wp.get("y") != op.get("y"):
                        return True
                return False

        thermal_preview = ImageAnnotationPreview(
            notebook,
            "thermal",
            annotation,
            defect_combo,
            shape_combo,
            annotation.get("thermal_image"),
            annotation.get("thermal_overlays", [])
        )
        notebook.add(thermal_preview.frame, text="サーモ画像")

        visible_preview = ImageAnnotationPreview(
            notebook,
            "visible",
            annotation,
            defect_combo,
            shape_combo,
            annotation.get("visible_image"),
            annotation.get("visible_overlays", [])
        )
        notebook.add(visible_preview.frame, text="可視画像")

        def refresh_preview_style(event=None):
            thermal_preview.redraw_overlays()
            visible_preview.redraw_overlays()

        defect_combo.bind("<<ComboboxSelected>>", refresh_preview_style)
        shape_combo.bind("<<ComboboxSelected>>", refresh_preview_style)

        def format_image_label(prefix, path):
            return f"{prefix}: {os.path.basename(path)}" if path else f"{prefix}: 未選択"

        thermal_label_var = tk.StringVar(value=format_image_label("サーモ画像", annotation.get("thermal_image")))
        visible_label_var = tk.StringVar(value=format_image_label("可視画像", annotation.get("visible_image")))

        def select_thermal_visible_pair():
            candidates = [
                annotation.get("thermal_image"),
                annotation.get("visible_image"),
                self.last_thermal_visible_dir
            ]
            initial_dir = None
            for candidate in candidates:
                if not candidate:
                    continue
                candidate_dir = candidate if os.path.isdir(candidate) else os.path.dirname(candidate)
                if candidate_dir and os.path.isdir(candidate_dir):
                    initial_dir = candidate_dir
                    break
            selector = ThermalVisibleFileDialog(dialog, initial_dir=initial_dir)
            result = selector.show()
            if result:
                thermal_path, visible_path = result

                prev_thermal = annotation.get("thermal_image")
                annotation["thermal_image"] = thermal_path
                if prev_thermal != thermal_path:
                    annotation["thermal_overlays"] = []
                thermal_preview.update_source(thermal_path, annotation.get("thermal_overlays", []))
                thermal_label_var.set(format_image_label("サーモ画像", thermal_path))

                prev_visible = annotation.get("visible_image")
                annotation["visible_image"] = visible_path
                if prev_visible != visible_path:
                    annotation["visible_overlays"] = []
                visible_preview.update_source(visible_path, annotation.get("visible_overlays", []))
                visible_label_var.set(format_image_label("可視画像", visible_path))

                self.last_thermal_visible_dir = os.path.dirname(thermal_path)

        def select_thermal_image():
            file_path = filedialog.askopenfilename(
                title="サーモ画像を選択",
                filetypes=[("画像ファイル", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif")]
            )
            if file_path:
                prev_thermal = annotation.get("thermal_image")
                annotation["thermal_image"] = file_path
                if prev_thermal != file_path:
                    annotation["thermal_overlays"] = []
                thermal_preview.update_source(file_path, annotation.get("thermal_overlays", []))
                thermal_label_var.set(format_image_label("サーモ画像", file_path))
                self.last_thermal_visible_dir = os.path.dirname(file_path)

        def select_thermal_image_odm():
            if not self.webodm_path:
                messagebox.showwarning("警告", "メイン画面で「WebODMフォルダ選択」ボタンからアセットフォルダを先に選択してください。")
                return

            def callback(selected_path):
                prev_thermal = annotation.get("thermal_image")
                annotation["thermal_image"] = selected_path
                if prev_thermal != selected_path:
                    annotation["thermal_overlays"] = []
                thermal_preview.update_source(selected_path, annotation.get("thermal_overlays", []))
                thermal_label_var.set(format_image_label("サーモ画像", selected_path))
                self.last_thermal_visible_dir = os.path.dirname(selected_path)

            ODMImageSelector(dialog, annotation, "thermal", self.webodm_path, callback, app_ref=self)

        def select_visible_image():
            file_path = filedialog.askopenfilename(
                title="可視画像を選択",
                filetypes=[("画像ファイル", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif")]
            )
            if file_path:
                prev_visible = annotation.get("visible_image")
                annotation["visible_image"] = file_path
                if prev_visible != file_path:
                    annotation["visible_overlays"] = []
                visible_preview.update_source(file_path, annotation.get("visible_overlays", []))
                visible_label_var.set(format_image_label("可視画像", file_path))
                self.last_thermal_visible_dir = os.path.dirname(file_path)

        def select_visible_image_odm():
            if not self.webodm_path:
                messagebox.showwarning("警告", "メイン画面で「WebODMフォルダ選択」ボタンからアセットフォルダを先に選択してください。")
                return

            def callback(selected_path):
                prev_visible = annotation.get("visible_image")
                annotation["visible_image"] = selected_path
                if prev_visible != selected_path:
                    annotation["visible_overlays"] = []
                visible_preview.update_source(selected_path, annotation.get("visible_overlays", []))
                visible_label_var.set(format_image_label("可視画像", selected_path))
                self.last_thermal_visible_dir = os.path.dirname(selected_path)

            ODMImageSelector(dialog, annotation, "visible", self.webodm_path, callback, app_ref=self)

        thermal_frame = ttk.Frame(form_frame)
        thermal_frame.grid(row=current_row, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        ttk.Button(thermal_frame, text="サーモ画像選択", command=select_thermal_image).pack(side=tk.LEFT, padx=(0, 10))
        odm_button_group = ttk.Frame(thermal_frame)
        odm_button_group.pack(side=tk.LEFT)
        ttk.Button(odm_button_group, text="ODM選択", command=select_thermal_image_odm).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(odm_button_group, text="サーモ画像同時選択", command=select_thermal_visible_pair).pack(side=tk.LEFT)
        current_row += 1

        ttk.Label(form_frame, textvariable=thermal_label_var).grid(row=current_row, column=0, columnspan=2, padx=10, pady=5, sticky="w")
        current_row += 1

        visible_frame = ttk.Frame(form_frame)
        visible_frame.grid(row=current_row, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        ttk.Button(visible_frame, text="可視画像選択", command=select_visible_image).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(visible_frame, text="ODM選択", command=select_visible_image_odm).pack(side=tk.LEFT)
        current_row += 1

        ttk.Label(form_frame, textvariable=visible_label_var).grid(row=current_row, column=0, columnspan=2, padx=10, pady=5, sticky="w")
        current_row += 1

        def save_changes():
            """フォーム項目を保存し、未保存の画像アノテーションがあれば確認する"""
            # サーモ画像と可視画像の未保存変更をチェック
            thermal_has_changes = thermal_preview.has_unsaved_changes()
            visible_has_changes = visible_preview.has_unsaved_changes()
            
            # 未保存の画像アノテーションがある場合、確認ダイアログを表示
            if thermal_has_changes or visible_has_changes:
                unsaved_types = []
                if thermal_has_changes:
                    unsaved_types.append("サーモ画像")
                if visible_has_changes:
                    unsaved_types.append("可視画像")
                
                message = f"{' と '.join(unsaved_types)}のタブに未保存のアノテーションがあります。\n\n保存しますか？"
                
                result = messagebox.askyesnocancel(
                    "未保存の変更",
                    message,
                    parent=dialog
                )
                
                if result is None:  # キャンセル - ダイアログを閉じない
                    return
                elif result:  # はい - 未保存の画像アノテーションを保存
                    if thermal_has_changes:
                        thermal_preview.confirm()
                    if visible_has_changes:
                        visible_preview.confirm()
                # いいえの場合は保存せずに続行
            
            # フォーム項目を保存
            for key, entry in entries.items():
                annotation[key] = entry.get()
            annotation["defect_type"] = defect_combo.get() or annotation.get("defect_type", "")
            annotation["shape"] = self.annotation_shapes.get(shape_combo.get(), annotation.get("shape", "cross"))
            annotation["management_level"] = (mgmt_var.get() or 'S')
            
            # 変更追跡：編集前と比較
            change_type = self._annotation_changed(original_annotation, annotation)
            if change_type:
                self.has_unsaved_changes = True
                self.changed_annotation_ids.add(annotation['id'])

            self.update_table()
            self.draw_annotations()
            dialog.destroy()

        def check_unsaved_and_close():
            """未保存の変更をチェックしてダイアログを閉じる"""
            # サーモ画像と可視画像の両方で未保存変更をチェック
            thermal_has_changes = thermal_preview.has_unsaved_changes()
            visible_has_changes = visible_preview.has_unsaved_changes()
            
            if thermal_has_changes or visible_has_changes:
                # 未保存の変更がある場合、確認ダイアログを表示
                unsaved_types = []
                if thermal_has_changes:
                    unsaved_types.append("サーモ画像")
                if visible_has_changes:
                    unsaved_types.append("可視画像")
                
                message = f"{' と '.join(unsaved_types)}のタブに未保存のアノテーションがあります。\n\n保存しますか？"
                
                result = messagebox.askyesnocancel(
                    "未保存の変更",
                    message,
                    parent=dialog
                )
                
                if result is None:  # キャンセル
                    return
                elif result:  # はい（保存する）
                    # 未保存の変更があるタブの保存を実行
                    if thermal_has_changes:
                        thermal_preview.confirm()
                    if visible_has_changes:
                        visible_preview.confirm()
                # いいえの場合は何もせずにダイアログを閉じる
            
            dialog.destroy()
        
        def cancel_changes():
            check_unsaved_and_close()

        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=current_row, column=0, columnspan=2, pady=20)
        ttk.Button(button_frame, text="保存", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="キャンセル", command=cancel_changes).pack(side=tk.LEFT, padx=5)

        dialog.protocol("WM_DELETE_WINDOW", check_unsaved_and_close)

    def update_table(self):
        """テーブルを更新"""
        # 既存のアイテムをクリア
        for item in self.tree.get_children():
            self.tree.delete(item)

        # アノテーションをIDでソート
        sorted_annotations = sorted(self.annotations, key=lambda x: x['id'])

        # テーブルに追加
        for annotation in sorted_annotations:
            shape_name = next((k for k, v in self.annotation_shapes.items() 
                            if v == annotation.get("shape", "cross")), "十字")
            
            # サーモ画像のファイル名を命名規則に従って生成
            thermal_filename = ""
            if annotation.get('thermal_image'):
                src_path = annotation['thermal_image']
                if os.path.exists(src_path):
                    thermal_filename = f"ID{annotation['id']}_{annotation['defect_type']}_サーモ異常{os.path.splitext(src_path)[1]}"
            
            # 可視画像のファイル名を命名規則に従って生成
            visible_filename = ""
            if annotation.get('visible_image'):
                src_path = annotation['visible_image']
                if os.path.exists(src_path):
                    visible_filename = f"ID{annotation['id']}_{annotation['defect_type']}_可視異常{os.path.splitext(src_path)[1]}"
            
            values = (
                annotation['id'],
                annotation.get('area_no', ''),
                annotation.get('pcs_no', ''),
                annotation.get('junction_box_no', ''),
                annotation.get('circuit_no', ''),
                annotation.get('array_no', ''),           # 追加
                annotation.get('module_position', ''),    # 追加
                annotation.get('serial_no', ''),          # 追加
                annotation.get('defect_type', ''),
                shape_name,
                thermal_filename,
                visible_filename,
                annotation.get('remarks', ''),
                annotation.get('report_date', '')         # 追加
            )
            self.tree.insert("", tk.END, values=values)

    def on_mouse_wheel(self, event):
        """マウスホイールでスクロール（縦横対応）"""
        # Shiftキー押下で横スクロール、通常は縦スクロール
        if event.state & 0x1:  # Shift押下
            # 横スクロール
            if event.delta > 0:
                self.canvas.xview_scroll(-1, "units")
            else:
                self.canvas.xview_scroll(1, "units")
        else:
            # 縦スクロール
            if event.delta > 0:
                self.canvas.yview_scroll(-1, "units")
            else:
                self.canvas.yview_scroll(1, "units")
    
    def zoom_in(self):
        """ズームイン（1.25倍ずつ拡大）"""
        if not self.current_image:
            return
        new_zoom = self.zoom_factor * 1.25
        new_zoom = min(new_zoom, 5.0)  # 最大500%
        self.set_zoom_factor(new_zoom)
    
    def zoom_out(self):
        """ズームアウト（0.8倍ずつ縮小）"""
        if not self.current_image:
            return
        new_zoom = self.zoom_factor * 0.8
        new_zoom = max(new_zoom, 0.1)  # 最小10%
        self.set_zoom_factor(new_zoom)
    
    def set_zoom_factor(self, new_zoom, keep_center=True):
        """ズーム倍率を設定して画像を再描画（中心位置維持オプション）"""
        if not self.current_image:
            return
        
        old_zoom = self.zoom_factor
        
        # 中心位置を計算（keep_center=Trueの場合）
        if keep_center and self.canvas_image:
            try:
                canvas_width = self.canvas.winfo_width()
                canvas_height = self.canvas.winfo_height()
                
                if canvas_width > 1 and canvas_height > 1:
                    # 現在のスクロール位置を取得
                    x_scroll_range = self.canvas.xview()
                    y_scroll_range = self.canvas.yview()
                    
                    # 表示領域の中心座標（画像座標系）
                    center_x = (x_scroll_range[0] + (x_scroll_range[1] - x_scroll_range[0]) / 2) * self.current_image.width
                    center_y = (y_scroll_range[0] + (y_scroll_range[1] - y_scroll_range[0]) / 2) * self.current_image.height
                else:
                    center_x = center_y = None
            except:
                center_x = center_y = None
        else:
            center_x = center_y = None
        
        # ズーム倍率を更新
        self.zoom_factor = new_zoom
        
        # 画像を再描画
        self.display_image()
        
        # 中心位置を維持
        if center_x is not None and center_y is not None and self.canvas_image:
            try:
                canvas_width = self.canvas.winfo_width()
                canvas_height = self.canvas.winfo_height()
                
                if canvas_width > 1 and canvas_height > 1:
                    # 新しいズーム倍率での画像サイズ
                    new_img_width = self.current_image.width * self.zoom_factor
                    new_img_height = self.current_image.height * self.zoom_factor
                    
                    # 中心座標から新しいスクロール位置を計算
                    new_x_scroll = (center_x * self.zoom_factor - canvas_width / 2) / new_img_width
                    new_y_scroll = (center_y * self.zoom_factor - canvas_height / 2) / new_img_height
                    
                    # スクロール位置を調整
                    self.canvas.xview_moveto(max(0, min(1, new_x_scroll)))
                    self.canvas.yview_moveto(max(0, min(1, new_y_scroll)))
            except:
                pass
        
        # 倍率表示を更新
        self.update_zoom_display()
    
    def on_zoom_combo_change(self, event=None):
        """倍率コンボボックス変更時の処理"""
        if self._updating_zoom_var:
            return
        
        selected = self.zoom_combo.get()
        
        # "100%" → 1.0 に変換
        for label, value in self.zoom_options:
            if label == selected:
                self.set_zoom_factor(value)
                break
    
    def update_zoom_display(self):
        """現在の倍率をコンボボックスに反映"""
        if not hasattr(self, 'zoom_combo'):
            return
        
        current_percent = f"{int(self.zoom_factor * 100)}%"
        
        # リストにある値なら選択、なければカスタム表示
        found = False
        for label, value in self.zoom_options:
            if abs(value - self.zoom_factor) < 0.01:
                self._updating_zoom_var = True
                self.zoom_combo.set(label)
                self._updating_zoom_var = False
                found = True
                break
        
        if not found:
            self._updating_zoom_var = True
            self.zoom_combo.set(current_percent)
            self._updating_zoom_var = False

    def start_pan(self, event):
        """パン開始"""
        self.pan_start_x = event.x
        self.pan_start_y = event.y

    def do_pan(self, event):
        """パン実行"""
        dx = event.x - self.pan_start_x
        dy = event.y - self.pan_start_y

        self.canvas.scan_dragto(dx, dy, gain=1)

    def reset_image(self):
        """画像をリセット"""
        if messagebox.askyesno("確認", "画像をリセットしますか？アノテーションもすべて削除されます。"):
            self.current_image = None
            self.image_path = None
            self.annotations = []
            self.next_id = 1
            self.canvas.delete("all")
            self.update_table()

    def customize_settings(self):
        """色とアノテーション位置オフセットを設定"""
        dialog = tk.Toplevel(self.root)
        dialog.title("色・アノテーション位置設定")
        dialog.geometry("500x650")
        dialog.transient(self.root)
        dialog.grab_set()

        # スクロール可能なフレームを作成
        canvas = tk.Canvas(dialog)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # === 色設定セクション ===
        color_frame = ttk.LabelFrame(scrollable_frame, text="色設定", padding=10)
        color_frame.pack(fill=tk.X, padx=10, pady=10)

        color_vars = {}

        for i, (defect_type, color) in enumerate(self.defect_types.items()):
            ttk.Label(color_frame, text=f"{defect_type}:").grid(row=i, column=0, sticky="w", padx=10, pady=5)

            color_var = tk.StringVar(value=color)
            color_vars[defect_type] = color_var

            def choose_color(dt=defect_type, cv=color_var):
                color = colorchooser.askcolor(initialcolor=cv.get())[1]
                if color:
                    cv.set(color)

            ttk.Button(color_frame, text="色選択", command=choose_color).grid(row=i, column=1, padx=10, pady=5)

        # === アノテーション位置調整セクション ===
        offset_frame = ttk.LabelFrame(scrollable_frame, text="アノテーション位置調整 (ピクセル単位)", padding=10)
        offset_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(offset_frame, text="※ 保存時にサーモ画像・可視画像に描画するアノテーションの位置を調整できます", 
                  font=("", 9)).grid(row=0, column=0, columnspan=4, pady=(0, 10))

        # サーモ画像オフセット
        ttk.Label(offset_frame, text="サーモ画像", font=("", 10, "bold")).grid(row=1, column=0, columnspan=4, sticky="w", pady=(5, 5))
        
        ttk.Label(offset_frame, text="X軸オフセット:").grid(row=2, column=0, sticky="e", padx=(10, 5), pady=5)
        thermal_x_var = tk.IntVar(value=self.thermal_offset_x)
        thermal_x_spinbox = ttk.Spinbox(offset_frame, from_=-1000, to=1000, textvariable=thermal_x_var, width=10)
        thermal_x_spinbox.grid(row=2, column=1, sticky="w", padx=(0, 10), pady=5)
        ttk.Label(offset_frame, text="px").grid(row=2, column=2, sticky="w", pady=5)

        ttk.Label(offset_frame, text="Y軸オフセット:").grid(row=3, column=0, sticky="e", padx=(10, 5), pady=5)
        thermal_y_var = tk.IntVar(value=self.thermal_offset_y)
        thermal_y_spinbox = ttk.Spinbox(offset_frame, from_=-1000, to=1000, textvariable=thermal_y_var, width=10)
        thermal_y_spinbox.grid(row=3, column=1, sticky="w", padx=(0, 10), pady=5)
        ttk.Label(offset_frame, text="px").grid(row=3, column=2, sticky="w", pady=5)

        # 可視画像オフセット
        ttk.Separator(offset_frame, orient="horizontal").grid(row=4, column=0, columnspan=4, sticky="ew", pady=10)
        ttk.Label(offset_frame, text="可視画像", font=("", 10, "bold")).grid(row=5, column=0, columnspan=4, sticky="w", pady=(5, 5))
        
        ttk.Label(offset_frame, text="X軸オフセット:").grid(row=6, column=0, sticky="e", padx=(10, 5), pady=5)
        visible_x_var = tk.IntVar(value=self.visible_offset_x)
        visible_x_spinbox = ttk.Spinbox(offset_frame, from_=-1000, to=1000, textvariable=visible_x_var, width=10)
        visible_x_spinbox.grid(row=6, column=1, sticky="w", padx=(0, 10), pady=5)
        ttk.Label(offset_frame, text="px").grid(row=6, column=2, sticky="w", pady=5)

        ttk.Label(offset_frame, text="Y軸オフセット:").grid(row=7, column=0, sticky="e", padx=(10, 5), pady=5)
        visible_y_var = tk.IntVar(value=self.visible_offset_y)
        visible_y_spinbox = ttk.Spinbox(offset_frame, from_=-1000, to=1000, textvariable=visible_y_var, width=10)
        visible_y_spinbox.grid(row=7, column=1, sticky="w", padx=(0, 10), pady=5)
        ttk.Label(offset_frame, text="px").grid(row=7, column=2, sticky="w", pady=5)

        # オルソ画像オフセット
        ttk.Separator(offset_frame, orient="horizontal").grid(row=8, column=0, columnspan=4, sticky="ew", pady=10)
        ttk.Label(offset_frame, text="オルソ画像（全体図）", font=("", 10, "bold")).grid(row=9, column=0, columnspan=4, sticky="w", pady=(5, 5))
        
        ttk.Label(offset_frame, text="X軸オフセット:").grid(row=10, column=0, sticky="e", padx=(10, 5), pady=5)
        ortho_x_var = tk.IntVar(value=self.ortho_offset_x)
        ortho_x_spinbox = ttk.Spinbox(offset_frame, from_=-1000, to=1000, textvariable=ortho_x_var, width=10)
        ortho_x_spinbox.grid(row=10, column=1, sticky="w", padx=(0, 10), pady=5)
        ttk.Label(offset_frame, text="px").grid(row=10, column=2, sticky="w", pady=5)

        ttk.Label(offset_frame, text="Y軸オフセット:").grid(row=11, column=0, sticky="e", padx=(10, 5), pady=5)
        ortho_y_var = tk.IntVar(value=self.ortho_offset_y)
        ortho_y_spinbox = ttk.Spinbox(offset_frame, from_=-1000, to=1000, textvariable=ortho_y_var, width=10)
        ortho_y_spinbox.grid(row=11, column=1, sticky="w", padx=(0, 10), pady=5)
        ttk.Label(offset_frame, text="px").grid(row=11, column=2, sticky="w", pady=5)

        # 適用ボタン
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.pack(fill=tk.X, padx=10, pady=20)

        def apply_settings():
            # 色設定を適用
            for defect_type, color_var in color_vars.items():
                self.defect_types[defect_type] = color_var.get()
            
            # オフセット変更チェック（変更前の値を保存）
            offset_changed = (
                self.thermal_offset_x != thermal_x_var.get() or
                self.thermal_offset_y != thermal_y_var.get() or
                self.visible_offset_x != visible_x_var.get() or
                self.visible_offset_y != visible_y_var.get() or
                self.ortho_offset_x != ortho_x_var.get() or
                self.ortho_offset_y != ortho_y_var.get()
            )
            
            # オフセット設定を適用
            self.thermal_offset_x = thermal_x_var.get()
            self.thermal_offset_y = thermal_y_var.get()
            self.visible_offset_x = visible_x_var.get()
            self.visible_offset_y = visible_y_var.get()
            self.ortho_offset_x = ortho_x_var.get()
            self.ortho_offset_y = ortho_y_var.get()
            
            # オフセット設定を保存
            self.save_offset_settings()
            
            # 変更追跡：オフセット変更時は全アノテーションが変更対象
            if offset_changed:
                self.has_unsaved_changes = True
                self.changed_annotation_ids = set(a['id'] for a in self.annotations)
            
            # 表示を更新
            self.draw_annotations()
            dialog.destroy()
            messagebox.showinfo("設定完了", "色とアノテーション位置の設定を保存しました。")

        ttk.Button(button_frame, text="適用して閉じる", command=apply_settings).pack(pady=10)

    def save_project(self):
        """プロジェクトを差分保存"""
        if not self.annotations:
            messagebox.showwarning("警告", "保存するアノテーションがありません。")
            return

        try:
            # 変更を検出
            changes = self.detect_changes()
            
            # テキストのみ変更の場合、画像生成の確認
            generate_images = True
            if changes['has_text_only_changes'] and not changes['has_visual_changes']:
                result = messagebox.askyesnocancel(
                    "画像生成の確認",
                    "画像に影響しない変更（テキスト情報のみ）が検出されました。\n\n"
                    "画像を再生成しますか？\n\n"
                    "・はい: 画像を再生成する\n"
                    "・いいえ: 一覧表のみ更新する（推奨）\n"
                    "・キャンセル: 保存を中止する",
                    default=messagebox.NO
                )
                if result is None:  # キャンセル
                    return
                generate_images = result
            
            # アノテーション設定を保存（常に実行）
            annotation_data = {
                "project_name": self.project_name,
                "image_path": self.image_path,
                "annotations": self.annotations,
                "defect_types": self.defect_types,
                "annotation_shapes": self.annotation_shapes,
                "created_date": datetime.now().isoformat()
            }

            annotation_file = os.path.join(self.project_path, "アノテーション設定フォルダ", "annotations.json")
            with open(annotation_file, 'w', encoding='utf-8') as f:
                json.dump(annotation_data, f, ensure_ascii=False, indent=2)

            # CSV/Excel出力（常に実行）
            self.export_to_csv()
            self.export_to_excel()
            try:
                self.export_v2_csv()
                self.export_v2_xlsx()
            except Exception as e:
                print(f"[WARN] v2出力でエラー: {e}")
            
            # 画像生成処理（変更がある場合のみ、またはユーザーが選択した場合）
            if generate_images:
                # 削除されたアノテーションの画像を削除
                if changes['deleted_ids']:
                    self.delete_related_images(changes['deleted_ids'])
                
                # 変更・追加されたアノテーションの関連画像を生成
                if changes['changed_ids'] or changes['added_ids']:
                    target_ids = changes['changed_ids'] | changes['added_ids']
                    self.copy_related_images_incremental(target_ids)
                
                # 全体図の再生成（必要な場合のみ）
                if changes['need_full_redraw']:
                    self.save_annotated_image()
                
                # 個別全体図の再生成（変更されたIDのみ）
                if changes['need_individual_redraw']:
                    self.save_individual_annotated_images_incremental(
                        changes['need_individual_redraw']
                    )
            
            # スナップショットを更新
            self.update_snapshot()
            self.has_unsaved_changes = False
            self.changed_annotation_ids.clear()
            
            # 成功メッセージ
            message = self._create_save_message(changes, generate_images)
            messagebox.showinfo("成功", message)

        except Exception as e:
            messagebox.showerror("エラー", f"保存に失敗しました: {str(e)}")

    def save_annotated_image(self):
        """アノテーション入り画像を保存"""
        if not self.current_image:
            return

        # 元画像をコピー
        annotated_image = self.current_image.copy()
        draw = ImageDraw.Draw(annotated_image)
        overall_scale = self._get_annotation_scale("overall")

        # アノテーションを描画
        for annotation in self.annotations:
            x = annotation["x"]
            y = annotation["y"]
            defect_type = annotation.get("defect_type")
            color = self.defect_types.get(defect_type, "#FF0000")
            shape = annotation.get("shape", "cross")

            icon_height = self.draw_annotation_icon_on_image(
                annotated_image,
                draw,
                x,
                y,
                defect_type,
                color,
                shape,
                overall_scale,
                image_type='ortho'  # オルソ画像全体図用のオフセットを適用
            )

            # ID番号を描画
            self._draw_id_label_on_image(
                draw,
                x,
                y,
                annotation['id'],
                color,
                annotated_image.size,
                overall_scale,
                icon_height,
                image_type='ortho'  # オルソ画像全体図用のオフセットを適用
            )

        # 保存
        output_path = os.path.join(self.project_path, "アノテーション入り画像フォルダ", 
                                  f"{self.project_name}_annotated.png")
        annotated_image.save(output_path)

    def _create_extended_image(self, original_image):
        """
        画像の上下に白色背景を追加して拡張する
        
        Args:
            original_image: 元画像（PIL Image）
        
        Returns:
            tuple: (拡張画像, 上部オフセット)
        """
        width, height = original_image.size
        
        # 拡張サイズを計算
        total_extension = int(height * self.image_extension_ratio)
        top_extension = total_extension // 2
        bottom_extension = total_extension - top_extension
        
        # 拡張後の画像サイズ
        new_height = height + top_extension + bottom_extension
        
        # 元画像のモードに合わせて白色背景の新画像を作成
        if original_image.mode == 'RGBA':
            extended_image = Image.new('RGBA', (width, new_height), (255, 255, 255, 255))
        else:
            extended_image = Image.new(original_image.mode, (width, new_height), (255, 255, 255))
        
        # 元画像を上部オフセット位置にペースト
        extended_image.paste(original_image, (0, top_extension))
        
        return extended_image, top_extension

    def save_individual_annotated_images(self):
        """各不具合IDごとに、そのIDのアノテーションのみを配置した全体図を個別に保存"""
        if not self.current_image:
            return
        
        if not self.annotations:
            return
        
        # 全体図位置フォルダのパスを取得
        output_folder = os.path.join(self.project_path, "全体図位置フォルダ")
        os.makedirs(output_folder, exist_ok=True)
        
        overall_scale = self._get_annotation_scale("overall")
        
        # 各不具合IDごとにループ処理
        for annotation in self.annotations:
            try:
                # 画像を拡張（上下に白色背景を追加）
                extended_image, top_offset = self._create_extended_image(self.current_image)
                annotated_image = extended_image.copy()
                draw = ImageDraw.Draw(annotated_image)
                
                # 現在のIDのアノテーションのみを描画（座標を調整）
                x = annotation["x"]
                defect_y = annotation["y"] + top_offset  # 不具合の位置（上部オフセット分だけ調整）
                defect_type = annotation.get("defect_type", "不具合")
                color = self.defect_types.get(defect_type, "#FF0000")
                shape = annotation.get("shape", "cross")
                
                # アノテーションの高さを事前計算（アノテーション下部を不具合位置に合わせるため）
                # アイコンがある場合はアイコンの高さを推定、ない場合は形状の高さを推定
                base_icon = self.load_annotation_icon(defect_type)
                if base_icon is not None:
                    # アイコンの高さを推定（draw_annotation_icon_on_imageと同じロジック）
                    icon_w, icon_h = base_icon.size
                    base_length = max(annotated_image.width, annotated_image.height)
                    min_dim = max(1, min(annotated_image.width, annotated_image.height))
                    base_target = min(max(24, int(base_length * 0.05)), max(16, min_dim), 128)
                    target_edge = max(4, int(round(base_target * overall_scale)))
                    target_edge = min(target_edge, min_dim)
                    scale = target_edge / max(icon_w, icon_h) if max(icon_w, icon_h) else 1.0
                    estimated_height = max(1, int(round(icon_h * scale)))
                else:
                    # アイコンがない場合は形状の高さを推定
                    if shape in ["cross", "arrow", "rectangle"]:
                        estimated_height = max(6, int(round(20 * overall_scale))) * 2
                    elif shape == "circle":
                        estimated_height = max(6, int(round(25 * overall_scale))) * 2
                    else:
                        estimated_height = max(6, int(round(20 * overall_scale))) * 2
                
                # アノテーション描画位置を計算（下部が不具合位置に来るように上方向に調整）
                annotation_y = defect_y - estimated_height / 2
                
                # アノテーションアイコン/シェイプを描画
                icon_height = self.draw_annotation_icon_on_image(
                    annotated_image,
                    draw,
                    x,
                    annotation_y,
                    defect_type,
                    color,
                    shape,
                    overall_scale,
                    image_type='ortho'  # オルソ画像全体図用のオフセットを適用
                )
                
                # ID番号を描画
                self._draw_id_label_on_image(
                    draw,
                    x,
                    annotation_y,
                    annotation['id'],
                    color,
                    annotated_image.size,
                    overall_scale,
                    icon_height,
                    image_type='ortho'  # オルソ画像全体図用のオフセットを適用
                )
                
                # ファイル名を生成: ID{id:03d}_全体図_{不具合名}.jpg
                annotation_id = annotation.get('id', 0)
                # ファイル名として不適切な文字を置換
                safe_defect_type = defect_type.replace("/", "_").replace("\\", "_").replace(":", "_")
                filename = f"ID{annotation_id:03d}_全体図_{safe_defect_type}.jpg"
                output_path = os.path.join(output_folder, filename)
                
                # JPEG保存前にRGBモードに変換（RGBAの場合）
                if annotated_image.mode == 'RGBA':
                    # 白背景に合成してRGBに変換
                    rgb_image = Image.new('RGB', annotated_image.size, (255, 255, 255))
                    rgb_image.paste(annotated_image, mask=annotated_image.split()[3])
                    annotated_image = rgb_image
                elif annotated_image.mode != 'RGB':
                    # その他のモードもRGBに変換
                    annotated_image = annotated_image.convert('RGB')
                
                # JPEG形式で保存（品質95%で高品質を保持）
                annotated_image.save(output_path, 'JPEG', quality=95)
                
            except Exception as e:
                # エラーが発生しても他のIDの処理は続行
                print(f"[WARN] ID{annotation.get('id', '?')}の全体図生成に失敗しました: {e}")
                continue

    def draw_cross_on_image(self, draw, x, y, color, scale_multiplier=1.0):
        """画像上に十字を描画"""
        size = max(6, int(round(20 * scale_multiplier)))
        line_width = max(1, int(round(3 * scale_multiplier)))
        draw.line([(x, y - size), (x, y + size)], fill=color, width=line_width)
        draw.line([(x - size, y), (x + size, y)], fill=color, width=line_width)
        return size * 2

    def draw_arrow_on_image(self, draw, x, y, color, scale_multiplier=1.0):
        """画像上に矢印を描画"""
        size = max(6, int(round(20 * scale_multiplier)))
        line_width = max(1, int(round(3 * scale_multiplier)))
        head_offset = max(6, int(round(15 * scale_multiplier)))
        head_width = max(4, int(round(8 * scale_multiplier)))
        draw.line([(x, y - size), (x, y + size)], fill=color, width=line_width)
        draw.polygon([
            (x, y - size),
            (x - head_width, y - size + head_offset),
            (x + head_width, y - size + head_offset)
        ], fill=color)
        return size * 2

    def draw_circle_on_image(self, draw, x, y, color, scale_multiplier=1.0):
        """画像上に円を描画"""
        radius = max(6, int(round(25 * scale_multiplier)))
        line_width = max(1, int(round(3 * scale_multiplier)))
        draw.ellipse([(x - radius, y - radius), (x + radius, y + radius)], outline=color, width=line_width)
        return radius * 2

    def draw_rectangle_on_image(self, draw, x, y, color, scale_multiplier=1.0):
        """画像上に四角を描画"""
        size = max(6, int(round(20 * scale_multiplier)))
        line_width = max(1, int(round(3 * scale_multiplier)))
        draw.rectangle([(x - size, y - size), (x + size, y + size)], outline=color, width=line_width)
        return size * 2

    def export_to_csv(self):
        """CSV形式で不具合一覧表を出力"""
        csv_file = os.path.join(self.project_path, "不具合一覧表フォルダ", "defect_list.csv")

        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # ヘッダー - 項目を追加
            headers = ["ID", "エリア番号", "PCS番号", "接続箱番号", "回路番号", "アレイ№", "モジュール位置", "シリアル№",
                    "不良分類", "形状", "サーモ画像", "可視画像", "備考", "報告年月日", "X座標", "Y座標"]
            writer.writerow(headers)

            # データ
            for annotation in sorted(self.annotations, key=lambda x: x['id']):
                shape_name = next((k for k, v in self.annotation_shapes.items() 
                                if v == annotation.get("shape", "cross")), "十字")
                
                # サーモ画像のファイル名を命名規則に従って生成
                thermal_filename = ""
                if annotation.get('thermal_image'):
                    src_path = annotation['thermal_image']
                    if os.path.exists(src_path):
                        thermal_filename = f"ID{annotation['id']}_{annotation['defect_type']}_サーモ異常{os.path.splitext(src_path)[1]}"
                
                # 可視画像のファイル名を命名規則に従って生成
                visible_filename = ""
                if annotation.get('visible_image'):
                    src_path = annotation['visible_image']
                    if os.path.exists(src_path):
                        visible_filename = f"ID{annotation['id']}_{annotation['defect_type']}_可視異常{os.path.splitext(src_path)[1]}"
                
                row = [
                    annotation['id'],
                    annotation.get('area_no', ''),
                    annotation.get('pcs_no', ''),
                    annotation.get('junction_box_no', ''),
                    annotation.get('circuit_no', ''),
                    annotation.get('array_no', ''),           # 追加
                    annotation.get('module_position', ''),    # 追加
                    annotation.get('serial_no', ''),          # 追加
                    annotation.get('defect_type', ''),
                    shape_name,
                    thermal_filename,
                    visible_filename,
                    annotation.get('remarks', ''),
                    annotation.get('report_date', ''),        # 追加
                    annotation['x'],
                    annotation['y']
                ]
                writer.writerow(row)

    def export_to_excel(self):
        """Excel形式で不具合一覧表を出力"""
        excel_file = os.path.join(self.project_path, "不具合一覧表フォルダ", "defect_list.xlsx")

        # データフレーム作成
        data = []
        for annotation in sorted(self.annotations, key=lambda x: x['id']):
            shape_name = next((k for k, v in self.annotation_shapes.items() 
                            if v == annotation.get("shape", "cross")), "十字")
            
            # サーモ画像のファイル名を命名規則に従って生成
            thermal_filename = ""
            if annotation.get('thermal_image'):
                src_path = annotation['thermal_image']
                if os.path.exists(src_path):
                    thermal_filename = f"ID{annotation['id']}_{annotation['defect_type']}_サーモ異常{os.path.splitext(src_path)[1]}"
            
            # 可視画像のファイル名を命名規則に従って生成
            visible_filename = ""
            if annotation.get('visible_image'):
                src_path = annotation['visible_image']
                if os.path.exists(src_path):
                    visible_filename = f"ID{annotation['id']}_{annotation['defect_type']}_可視異常{os.path.splitext(src_path)[1]}"
            
            data.append({
                "ID": annotation['id'],
                "エリア番号": annotation.get('area_no', ''),
                "PCS番号": annotation.get('pcs_no', ''),
                "接続箱番号": annotation.get('junction_box_no', ''),
                "回路番号": annotation.get('circuit_no', ''),
                "アレイ№": annotation.get('array_no', ''),           # 追加
                "モジュール位置": annotation.get('module_position', ''), # 追加
                "シリアル№": annotation.get('serial_no', ''),          # 追加
                "不良分類": annotation.get('defect_type', ''),
                "形状": shape_name,
                "サーモ画像": thermal_filename,
                "可視画像": visible_filename,
                "備考": annotation.get('remarks', ''),
                "報告年月日": annotation.get('report_date', ''),        # 追加
                "X座標": annotation['x'],
                "Y座標": annotation['y']
            })

        df = pd.DataFrame(data)
        df.to_excel(excel_file, index=False, engine='openpyxl')

    # ------------- 拡張フォーマット（v2）出力 -------------
    def to_v2_row(self, annotation: dict):
        def g(key):
            v = annotation.get(key)
            if v is None:
                return None
            if isinstance(v, str) and v.strip() == '':
                return None
            return v
        return [
            g('area_no'),                            # エリア(工区)
            g('junction_box_no'),                    # 接続箱No.
            None,                                    # 接続箱（該当なし）
            g('circuit_no'),                         # 回路No.
            g('array_no'),                           # アレイ番号
            g('module_position'),                    # モジュール場所
            g('defect_type'),                        # 不良分類
            normalize_management_level(annotation.get('management_level')),  # 管理レベル
            g('serial_no'),                          # シリアルナンバー(交換前)
            g('report_no'),                          # 報告書番号
            g('disconnect_date'),                    # 離線した日
            g('disconnect_location'),                # 離線場所
            g('serial_after'),                       # シリアルナンバー(交換後)
            None,                                    # モジュール
            None,                                    # 列1
            g('reenergize_date'),                    # 送電完了日
            g('report_no2'),                         # 報告書番号2
            g('remarks'),                            # 備考
            g('report_date'),                        # 報告年月日
            g('id'),                                 # 採番
        ]

    def export_v2_csv(self):
        csv_file = os.path.join(self.project_path, "不具合一覧表フォルダ", "defect_list_v2.csv")
        os.makedirs(os.path.dirname(csv_file), exist_ok=True)
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f, lineterminator='\r\n')
            w.writerow(DEFECT_V2_HEADERS)
            for annotation in sorted(self.annotations, key=lambda x: x['id']):
                row = self.to_v2_row(annotation)
                w.writerow(['' if v is None else v for v in row])

    def export_v2_xlsx(self):
        excel_file = os.path.join(self.project_path, "不具合一覧表フォルダ", "defect_list.xlsx")
        os.makedirs(os.path.dirname(excel_file), exist_ok=True)
        try:
            wb = load_workbook(excel_file)
        except Exception:
            wb = Workbook()
        if 'defect_list_v2' in wb.sheetnames:
            wb.remove(wb['defect_list_v2'])
        ws = wb.create_sheet('defect_list_v2')
        ws.append(DEFECT_V2_HEADERS)
        for annotation in sorted(self.annotations, key=lambda x: x['id']):
            ws.append(self.to_v2_row(annotation))
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            try:
                wb.remove(wb['Sheet'])
            except Exception:
                pass
        wb.save(excel_file)

    def copy_related_images(self):
        """
        関連画像（サーモ・可視）にアノテーションを描画して保存
        
        Requirement 1実装:
        - アノテーション入り画像のみを保存
        - オフセット設定を適用
        - 統一された命名規則: ID{id:03d}_{defect_type}_サーモ異常/可視異常{ext}
        
        修正内容（Requirement 4対応）:
        - thermal_overlays/visible_overlaysに保存された座標を使用
        - 編集ダイアログで設定した位置を基準にオフセット適用
        """
        for annotation in self.annotations:
            annotation_id = annotation['id']
            defect_type = annotation.get('defect_type', '不具合')
            shape = annotation.get('shape', 'cross')
            color = self.defect_types.get(defect_type, '#FF0000')
            
            # サーモ画像の処理
            if annotation.get('thermal_image'):
                src_path = annotation['thermal_image']
                # thermal_overlaysから座標を取得（編集ダイアログで設定した位置）
                thermal_overlays = annotation.get('thermal_overlays', [])
                
                if os.path.exists(src_path) and thermal_overlays:
                    try:
                        # 画像を読み込み
                        thermal_image = Image.open(src_path)
                        
                        # thermal_overlaysの各座標にアノテーションを描画
                        for overlay_point in thermal_overlays:
                            x = overlay_point['x']
                            y = overlay_point['y']
                            
                            # アノテーションを描画
                            thermal_image = self._draw_annotation_on_related_image(
                                thermal_image,
                                x, y,
                                annotation_id,
                                defect_type,
                                color,
                                shape,
                                image_type='thermal'
                            )
                        
                        annotated_thermal = thermal_image
                        
                        # ファイル名を生成（命名規則に従う）
                        ext = os.path.splitext(src_path)[1]
                        filename = f"ID{annotation_id:03d}_{defect_type}_サーモ異常{ext}"
                        dst_path = os.path.join(self.project_path, "サーモ画像フォルダ", filename)
                        
                        # 保存（JPEG形式の場合はRGB変換）
                        if ext.lower() in ['.jpg', '.jpeg']:
                            if annotated_thermal.mode == 'RGBA':
                                rgb_image = Image.new('RGB', annotated_thermal.size, (255, 255, 255))
                                rgb_image.paste(annotated_thermal, mask=annotated_thermal.split()[3])
                                annotated_thermal = rgb_image
                            elif annotated_thermal.mode != 'RGB':
                                annotated_thermal = annotated_thermal.convert('RGB')
                            annotated_thermal.save(dst_path, 'JPEG', quality=95)
                        else:
                            annotated_thermal.save(dst_path)
                            
                    except Exception as e:
                        print(f"[WARN] サーモ画像の処理に失敗 (ID{annotation_id}): {e}")

            # 可視画像の処理
            if annotation.get('visible_image'):
                src_path = annotation['visible_image']
                # visible_overlaysから座標を取得（編集ダイアログで設定した位置）
                visible_overlays = annotation.get('visible_overlays', [])
                
                if os.path.exists(src_path) and visible_overlays:
                    try:
                        # 画像を読み込み
                        visible_image = Image.open(src_path)
                        
                        # visible_overlaysの各座標にアノテーションを描画
                        for overlay_point in visible_overlays:
                            x = overlay_point['x']
                            y = overlay_point['y']
                            
                            # アノテーションを描画
                            visible_image = self._draw_annotation_on_related_image(
                                visible_image,
                                x, y,
                                annotation_id,
                                defect_type,
                                color,
                                shape,
                                image_type='visible'
                            )
                        
                        annotated_visible = visible_image
                        
                        # ファイル名を生成（命名規則に従う）
                        ext = os.path.splitext(src_path)[1]
                        filename = f"ID{annotation_id:03d}_{defect_type}_可視異常{ext}"
                        dst_path = os.path.join(self.project_path, "可視画像フォルダ", filename)
                        
                        # 保存（JPEG形式の場合はRGB変換）
                        if ext.lower() in ['.jpg', '.jpeg']:
                            if annotated_visible.mode == 'RGBA':
                                rgb_image = Image.new('RGB', annotated_visible.size, (255, 255, 255))
                                rgb_image.paste(annotated_visible, mask=annotated_visible.split()[3])
                                annotated_visible = rgb_image
                            elif annotated_visible.mode != 'RGB':
                                annotated_visible = annotated_visible.convert('RGB')
                            annotated_visible.save(dst_path, 'JPEG', quality=95)
                        else:
                            annotated_visible.save(dst_path)
                            
                    except Exception as e:
                        print(f"[WARN] 可視画像の処理に失敗 (ID{annotation_id}): {e}")
    
    def _draw_annotation_on_related_image(self, image, x, y, annotation_id, defect_type, color, shape, image_type):
        """
        サーモ画像・可視画像にアノテーションを描画
        
        Args:
            image: 元画像 (PIL Image)
            x, y: アノテーション座標
            annotation_id: アノテーションID
            defect_type: 不具合タイプ
            color: 色
            shape: 形状
            image_type: 画像タイプ ('thermal' または 'visible')
        
        Returns:
            アノテーション描画済み画像
        """
        # 画像をコピー
        annotated_image = image.copy()
        draw = ImageDraw.Draw(annotated_image)
        
        # スケール倍率を取得
        if image_type == 'thermal':
            scale_multiplier = self._get_annotation_scale('thermal')
        elif image_type == 'visible':
            scale_multiplier = self._get_annotation_scale('visible')
        else:
            scale_multiplier = self._get_annotation_scale('overall')
        
        # アノテーションアイコン/シェイプを描画（オフセット適用）
        icon_height = self.draw_annotation_icon_on_image(
            annotated_image,
            draw,
            x, y,
            defect_type,
            color,
            shape,
            scale_multiplier,
            image_type=image_type  # オフセット適用のためimage_typeを指定
        )
        
        # ID番号を描画（オフセット適用）
        self._draw_id_label_on_image(
            draw,
            x, y,
            annotation_id,
            color,
            annotated_image.size,
            scale_multiplier,
            icon_height,
            image_type=image_type  # オフセット適用のためimage_typeを指定
        )
        
        return annotated_image

    def save_offset_settings(self):
        """アノテーション位置オフセット設定をJSONファイルに保存"""
        try:
            # プロジェクトパスが設定されていない場合はスキップ
            if not self.project_path:
                return
            
            settings_folder = os.path.join(self.project_path, "アノテーション設定フォルダ")
            os.makedirs(settings_folder, exist_ok=True)
            
            settings_file = os.path.join(settings_folder, "offset_settings.json")
            
            offset_data = {
                "thermal_offset_x": self.thermal_offset_x,
                "thermal_offset_y": self.thermal_offset_y,
                "visible_offset_x": self.visible_offset_x,
                "visible_offset_y": self.visible_offset_y,
                "ortho_offset_x": self.ortho_offset_x,
                "ortho_offset_y": self.ortho_offset_y,
                "updated_date": datetime.now().isoformat()
            }
            
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(offset_data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"オフセット設定の保存に失敗しました: {str(e)}")

    def load_offset_settings(self):
        """アノテーション位置オフセット設定をJSONファイルから読み込み"""
        try:
            # プロジェクトパスが設定されていない場合はスキップ
            if not self.project_path:
                return
            
            settings_file = os.path.join(self.project_path, "アノテーション設定フォルダ", "offset_settings.json")
            
            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                self.thermal_offset_x = data.get('thermal_offset_x', 0)
                self.thermal_offset_y = data.get('thermal_offset_y', 0)
                self.visible_offset_x = data.get('visible_offset_x', 0)
                self.visible_offset_y = data.get('visible_offset_y', 0)
                self.ortho_offset_x = data.get('ortho_offset_x', 0)
                self.ortho_offset_y = data.get('ortho_offset_y', 0)
                
        except Exception as e:
            print(f"オフセット設定の読み込みに失敗しました: {str(e)}")

    def load_annotations(self):
        """アノテーションを読み込み"""
        try:
            annotation_file = os.path.join(self.project_path, "アノテーション設定フォルダ", "annotations.json")
            if os.path.exists(annotation_file):
                with open(annotation_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                self.annotations = data.get('annotations', [])
                for ann in self.annotations:
                    try:
                        ann['id'] = int(ann.get('id', 0))
                    except (TypeError, ValueError):
                        ann['id'] = 0
                self.defect_types.update(data.get('defect_types', {}))
                self.annotation_shapes.update(data.get('annotation_shapes', {}))
                self.image_path = data.get('image_path', '')

                self.initialize_annotation_icons(reset_warning=False)

                if hasattr(self, 'defect_combo'):
                    defect_keys = list(self.defect_types.keys()) or [self.defect_var.get()]
                    self.defect_combo['values'] = defect_keys
                    if self.defect_var.get() not in self.defect_types and defect_keys:
                        self.defect_var.set(defect_keys[0])
                if hasattr(self, 'shape_combo'):
                    shape_keys = list(self.annotation_shapes.keys()) or [self.shape_var.get()]
                    self.shape_combo['values'] = shape_keys
                    if self.shape_var.get() not in self.annotation_shapes and shape_keys:
                        self.shape_var.set(shape_keys[0])

                # 次のIDを設定
                if self.annotations:
                    max_id = max(ann['id'] for ann in self.annotations)
                    self.next_id = max_id + 1
                else:
                    self.next_id = 1
                
                # オフセット設定を読み込み
                self.load_offset_settings()

                self.update_table()
                if self.current_image:
                    self.draw_annotations()
                
                # 差分保存: 読み込み後にスナップショットを初期化
                self.update_snapshot()
                self.has_unsaved_changes = False
                self.changed_annotation_ids.clear()

        except Exception as e:
            messagebox.showerror("エラー", f"アノテーションの読み込みに失敗しました: {str(e)}")

    # ============================================================
    # 差分保存機能（Requirement 5）
    # ============================================================

    def create_snapshot(self):
        """
        現在の状態のスナップショットを作成
        
        Returns:
            dict: スナップショットデータ
        """
        return {
            'annotations': copy.deepcopy(self.annotations),
            'offsets': {
                'thermal_x': self.thermal_offset_x,
                'thermal_y': self.thermal_offset_y,
                'visible_x': self.visible_offset_x,
                'visible_y': self.visible_offset_y,
                'ortho_x': self.ortho_offset_x,
                'ortho_y': self.ortho_offset_y,
            },
            'colors': copy.deepcopy(self.defect_types),
        }

    def update_snapshot(self):
        """スナップショットを最新状態に更新"""
        self.last_saved_snapshot = self.create_snapshot()

    def detect_changes(self):
        """
        変更されたアノテーションを検出
        
        Returns:
            dict: {
                'changed_ids': set,              # 変更されたアノテーションID
                'deleted_ids': set,              # 削除されたアノテーションID
                'added_ids': set,                # 新規追加されたアノテーションID
                'need_full_redraw': bool,        # 全体図の再描画が必要か
                'need_individual_redraw': set,   # 個別全体図の再描画が必要なID
                'has_visual_changes': bool,      # 画像に影響する変更があるか
                'has_text_only_changes': bool,   # テキストのみの変更があるか
            }
        """
        if not self.last_saved_snapshot:
            # 初回保存時は全て変更扱い
            all_ids = set(a['id'] for a in self.annotations)
            return {
                'changed_ids': all_ids,
                'deleted_ids': set(),
                'added_ids': all_ids,
                'need_full_redraw': True,
                'need_individual_redraw': all_ids,
                'has_visual_changes': True,
                'has_text_only_changes': False,
            }
        
        current = {a['id']: a for a in self.annotations}
        previous = {a['id']: a for a in self.last_saved_snapshot['annotations']}
        
        current_ids = set(current.keys())
        previous_ids = set(previous.keys())
        
        added_ids = current_ids - previous_ids
        deleted_ids = previous_ids - current_ids
        potential_changed = current_ids & previous_ids
        
        visual_changed_ids = set()      # 画像に影響する変更
        text_only_changed_ids = set()   # テキストのみの変更
        
        for aid in potential_changed:
            change_type = self._annotation_changed(previous[aid], current[aid])
            if change_type == 'visual':
                visual_changed_ids.add(aid)
            elif change_type == 'text':
                text_only_changed_ids.add(aid)
        
        # オフセット・色の変更チェック
        offset_changed = self._offset_changed()
        color_changed = self._color_changed()
        
        # 画像に影響する変更があるか
        has_visual_changes = bool(visual_changed_ids or added_ids or deleted_ids or 
                                  offset_changed or color_changed)
        
        # テキストのみの変更があるか
        has_text_only_changes = bool(text_only_changed_ids) and not has_visual_changes
        
        # 全体図の再描画が必要か
        need_full_redraw = has_visual_changes
        
        # 個別全体図の再描画が必要なID
        need_individual_redraw = visual_changed_ids | added_ids
        
        return {
            'changed_ids': visual_changed_ids,
            'deleted_ids': deleted_ids,
            'added_ids': added_ids,
            'need_full_redraw': need_full_redraw,
            'need_individual_redraw': need_individual_redraw,
            'has_visual_changes': has_visual_changes,
            'has_text_only_changes': has_text_only_changes,
            'text_only_changed_ids': text_only_changed_ids,
        }

    def _annotation_changed(self, old, new):
        """
        アノテーションが変更されたか判定
        
        Args:
            old: 前回のアノテーションデータ
            new: 現在のアノテーションデータ
        
        Returns:
            str: 'visual' (画像に影響), 'text' (テキストのみ), None (変更なし)
        """
        # 画像に影響する変更を検出
        visual_keys = ['x', 'y', 'defect_type', 'shape', 
                       'thermal_image', 'visible_image', 
                       'thermal_overlays', 'visible_overlays']
        
        for key in visual_keys:
            if old.get(key) != new.get(key):
                return 'visual'
        
        # テキストのみの変更を検出
        text_keys = ['area_no', 'pcs_no', 'junction_box_no', 'circuit_no', 
                     'array_no', 'module_position', 'serial_no', 'remarks', 
                     'report_date', 'management_level']
        
        for key in text_keys:
            if old.get(key) != new.get(key):
                return 'text'
        
        return None

    def _offset_changed(self):
        """オフセット設定が変更されたか判定"""
        if not self.last_saved_snapshot:
            return False
        
        prev_offsets = self.last_saved_snapshot['offsets']
        return (
            self.thermal_offset_x != prev_offsets['thermal_x'] or
            self.thermal_offset_y != prev_offsets['thermal_y'] or
            self.visible_offset_x != prev_offsets['visible_x'] or
            self.visible_offset_y != prev_offsets['visible_y'] or
            self.ortho_offset_x != prev_offsets['ortho_x'] or
            self.ortho_offset_y != prev_offsets['ortho_y']
        )

    def _color_changed(self):
        """色設定が変更されたか判定"""
        if not self.last_saved_snapshot:
            return False
        
        prev_colors = self.last_saved_snapshot['colors']
        return self.defect_types != prev_colors

    def copy_related_images_incremental(self, target_ids):
        """
        指定されたIDのアノテーションのみ関連画像を生成（差分保存）
        
        Args:
            target_ids: set of int - 処理対象のアノテーションID
        """
        for annotation in self.annotations:
            if annotation['id'] not in target_ids:
                continue
            
            annotation_id = annotation['id']
            defect_type = annotation.get('defect_type', '不具合')
            shape = annotation.get('shape', 'cross')
            color = self.defect_types.get(defect_type, '#FF0000')
            
            # サーモ画像の処理
            if annotation.get('thermal_image'):
                src_path = annotation['thermal_image']
                thermal_overlays = annotation.get('thermal_overlays', [])
                
                if os.path.exists(src_path) and thermal_overlays:
                    try:
                        thermal_image = Image.open(src_path)
                        
                        for overlay_point in thermal_overlays:
                            x = overlay_point['x']
                            y = overlay_point['y']
                            
                            thermal_image = self._draw_annotation_on_related_image(
                                thermal_image, x, y, annotation_id,
                                defect_type, color, shape, image_type='thermal'
                            )
                        
                        annotated_thermal = thermal_image
                        ext = os.path.splitext(src_path)[1]
                        filename = f"ID{annotation_id:03d}_{defect_type}_サーモ異常{ext}"
                        dst_path = os.path.join(self.project_path, "サーモ画像フォルダ", filename)
                        
                        if ext.lower() in ['.jpg', '.jpeg']:
                            if annotated_thermal.mode == 'RGBA':
                                rgb_image = Image.new('RGB', annotated_thermal.size, (255, 255, 255))
                                rgb_image.paste(annotated_thermal, mask=annotated_thermal.split()[3])
                                annotated_thermal = rgb_image
                            elif annotated_thermal.mode != 'RGB':
                                annotated_thermal = annotated_thermal.convert('RGB')
                            annotated_thermal.save(dst_path, 'JPEG', quality=95)
                        else:
                            annotated_thermal.save(dst_path)
                            
                    except Exception as e:
                        print(f"[WARN] サーモ画像の処理に失敗 (ID{annotation_id}): {e}")
            
            # 可視画像の処理
            if annotation.get('visible_image'):
                src_path = annotation['visible_image']
                visible_overlays = annotation.get('visible_overlays', [])
                
                if os.path.exists(src_path) and visible_overlays:
                    try:
                        visible_image = Image.open(src_path)
                        
                        for overlay_point in visible_overlays:
                            x = overlay_point['x']
                            y = overlay_point['y']
                            
                            visible_image = self._draw_annotation_on_related_image(
                                visible_image, x, y, annotation_id,
                                defect_type, color, shape, image_type='visible'
                            )
                        
                        annotated_visible = visible_image
                        ext = os.path.splitext(src_path)[1]
                        filename = f"ID{annotation_id:03d}_{defect_type}_可視異常{ext}"
                        dst_path = os.path.join(self.project_path, "可視画像フォルダ", filename)
                        
                        if ext.lower() in ['.jpg', '.jpeg']:
                            if annotated_visible.mode == 'RGBA':
                                rgb_image = Image.new('RGB', annotated_visible.size, (255, 255, 255))
                                rgb_image.paste(annotated_visible, mask=annotated_visible.split()[3])
                                annotated_visible = rgb_image
                            elif annotated_visible.mode != 'RGB':
                                annotated_visible = annotated_visible.convert('RGB')
                            annotated_visible.save(dst_path, 'JPEG', quality=95)
                        else:
                            annotated_visible.save(dst_path)
                            
                    except Exception as e:
                        print(f"[WARN] 可視画像の処理に失敗 (ID{annotation_id}): {e}")

    def delete_related_images(self, deleted_ids):
        """
        削除されたアノテーションの画像ファイルを削除
        
        Args:
            deleted_ids: set of int - 削除されたアノテーションID
        """
        import glob
        
        for aid in deleted_ids:
            # サーモ画像の削除
            thermal_pattern = f"ID{aid:03d}_*_サーモ異常.*"
            thermal_folder = os.path.join(self.project_path, "サーモ画像フォルダ")
            if os.path.exists(thermal_folder):
                for file in glob.glob(os.path.join(thermal_folder, thermal_pattern)):
                    try:
                        os.remove(file)
                        print(f"[INFO] 削除: {file}")
                    except Exception as e:
                        print(f"[WARN] ファイル削除失敗: {file}, {e}")
            
            # 可視画像の削除
            visible_pattern = f"ID{aid:03d}_*_可視異常.*"
            visible_folder = os.path.join(self.project_path, "可視画像フォルダ")
            if os.path.exists(visible_folder):
                for file in glob.glob(os.path.join(visible_folder, visible_pattern)):
                    try:
                        os.remove(file)
                        print(f"[INFO] 削除: {file}")
                    except Exception as e:
                        print(f"[WARN] ファイル削除失敗: {file}, {e}")
            
            # 個別全体図の削除
            individual_pattern = f"ID{aid:03d}_*.*"
            individual_folder = os.path.join(self.project_path, "個別全体図フォルダ")
            if os.path.exists(individual_folder):
                for file in glob.glob(os.path.join(individual_folder, individual_pattern)):
                    try:
                        os.remove(file)
                        print(f"[INFO] 削除: {file}")
                    except Exception as e:
                        print(f"[WARN] ファイル削除失敗: {file}, {e}")

    def save_individual_annotated_images_incremental(self, target_ids):
        """
        指定されたIDの個別全体図のみを生成（差分保存）
        
        Args:
            target_ids: set of int - 処理対象のアノテーションID
        """
        if not self.current_image:
            return
        
        for annotation in sorted(self.annotations, key=lambda x: x['id']):
            if annotation['id'] not in target_ids:
                continue
            
            annotation_id = annotation['id']
            x = annotation['x']
            y = annotation['y']
            defect_type = annotation.get('defect_type', '不具合')
            color = self.defect_types.get(defect_type, '#FF0000')
            shape = annotation.get('shape', 'cross')
            
            # 画像をコピー
            annotated_image = self.current_image.copy()
            draw = ImageDraw.Draw(annotated_image)
            overall_scale = self._get_annotation_scale("overall")
            
            # 該当アノテーションのみ描画
            annotation_y = y
            icon_height = self.draw_annotation_icon_on_image(
                annotated_image, draw, x, annotation_y,
                defect_type, color, shape, overall_scale,
                image_type='ortho'
            )
            
            self._draw_id_label_on_image(
                draw, x, annotation_y, annotation_id, color,
                annotated_image.size, overall_scale, icon_height,
                image_type='ortho'
            )
            
            # 保存
            individual_folder = os.path.join(self.project_path, "個別全体図フォルダ")
            os.makedirs(individual_folder, exist_ok=True)
            
            base_name, ext = os.path.splitext(os.path.basename(self.image_path))
            filename = f"ID{annotation_id:03d}_{defect_type}{ext}"
            dst_path = os.path.join(individual_folder, filename)
            
            if ext.lower() in ['.jpg', '.jpeg']:
                if annotated_image.mode == 'RGBA':
                    rgb_image = Image.new('RGB', annotated_image.size, (255, 255, 255))
                    rgb_image.paste(annotated_image, mask=annotated_image.split()[3])
                    annotated_image = rgb_image
                elif annotated_image.mode != 'RGB':
                    annotated_image = annotated_image.convert('RGB')
                annotated_image.save(dst_path, 'JPEG', quality=95)
            else:
                annotated_image.save(dst_path)

    def _create_save_message(self, changes, generate_images):
        """
        保存時のメッセージを生成
        
        Args:
            changes: detect_changes()の結果
            generate_images: 画像を生成したかどうか
        
        Returns:
            str: メッセージ
        """
        added_count = len(changes['added_ids'])
        changed_count = len(changes['changed_ids'])
        deleted_count = len(changes['deleted_ids'])
        text_only_count = len(changes.get('text_only_changed_ids', set()))
        
        if not generate_images and changes['has_text_only_changes']:
            # 画像生成をスキップした場合
            message = f"プロジェクト '{self.project_name}' を保存しました。\n\n"
            message += "一覧表のみを更新しました（画像は再生成されませんでした）。\n"
            if text_only_count > 0:
                message += f"- テキスト変更: {text_only_count}件"
            return message
        
        if added_count == 0 and changed_count == 0 and deleted_count == 0 and text_only_count == 0:
            # 変更なし
            return f"プロジェクト '{self.project_name}' を保存しました。\n\n変更がないため、一覧表のみを更新しました。"
        
        # 変更あり
        message = f"プロジェクト '{self.project_name}' を保存しました。\n\n"
        details = []
        if added_count > 0:
            details.append(f"追加: {added_count}件")
        if changed_count > 0:
            details.append(f"変更: {changed_count}件")
        if deleted_count > 0:
            details.append(f"削除: {deleted_count}件")
        if text_only_count > 0 and generate_images:
            details.append(f"テキスト変更: {text_only_count}件")
        
        message += "\n".join(f"- {detail}" for detail in details)
        return message

    def quit_application(self):
        """アプリケーションを終了（未保存変更の警告付き）"""
        # 未保存の変更がある場合は警告
        if self.has_unsaved_changes:
            result = messagebox.askyesnocancel(
                "未保存の変更",
                "保存していない変更があります。保存しますか？\n\n"
                "・はい: 保存して終了\n"
                "・いいえ: 保存せずに終了\n"
                "・キャンセル: 終了を中止"
            )
            if result is None:  # キャンセル
                return
            elif result:  # はい
                self.save_project()
                # save_project でキャンセルされた場合は終了しない
                if self.has_unsaved_changes:
                    return
        
        self.root.quit()
            elif result is False:  # No
                self.root.quit()
            # Cancel の場合は何もしない
        else:
            self.root.quit()

    
def main():
    """メイン関数"""
    root = tk.Tk()
    app = OrthoImageAnnotationSystem(root)
    root.mainloop()


if __name__ == "__main__":
    main()
